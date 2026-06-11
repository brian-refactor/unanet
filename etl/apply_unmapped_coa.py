"""
Apply mappings to the 26 source accounts not yet tied to the master COA.
Updates the Needs Mapping tab in COA_Master_Mapping_v2.xlsx with the chosen
master COA code and name, and removes the red highlight once mapped.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

SRC  = Path(r'E:\Unanet\output\COA_Master_Mapping_v2.xlsx')
DEST = Path(r'E:\Unanet\output\COA_Master_Mapping_v2.xlsx')

# Mappings: (office, src_code) -> (master_code, master_name, rationale)
MAPPINGS = {
    # ── Dallas ────────────────────────────────────────────────────────────────
    ('DAL', 'Design_Income-Accrued'):      (42501, 'Fee Revenue - Accrued / WIP Adj',        'Accrued design income = WIP/accrued revenue'),
    ('DAL', 'Fusion_AE_-_FVC_Bank'):       (10302, 'Cash - FVCBank Checking',                'FVC Bank account maps to Fusion FVCBank checking'),
    ('DAL', 'Fusion_AE_-_Payroll'):        (10201, 'Cash - Fusion Payroll',                  'Payroll funding bank account'),
    ('DAL', 'ID_Studio4,_LLC_-_Sweep'):    (10301, 'Cash - Fusion Sweep',                    'Sweep account rolls to Fusion sweep'),
    ('DAL', 'Investment_in_Reztark_ACi'):  (15001, 'Investment Asset - DAL',                 'Intercompany investment = investment asset'),
    ('DAL', 'NWC_Escrow_-_Reztark'):       (25001, 'NWC Escrow - DAL',                       'Direct match — NWC escrow account'),
    ('DAL', 'Preferred_Equity_Payments'):  (81701, 'Management Fees',                        'Preferred return payments run through management fees'),

    # ── Minneapolis ───────────────────────────────────────────────────────────
    ('MSP', '401'):  (40001, 'Design Income',                'Billed fee revenue = design income'),
    ('MSP', '767'):  (63101, 'Project Related Costs',        'Drafting expenses = direct project costs'),
    ('MSP', '721'):  (72601, 'Equipment Rental - Indirect',  'Equipment rental = indirect overhead'),
    ('MSP', '980'):  (81303, 'State and County Taxes',       'MN minimum fee tax = state tax'),
    ('MSP', '186'):  (12401, 'Other Current Assets - MSP',   'Org/startup costs = other current assets'),
    ('MSP', '172'):  (81701, 'Management Fees',              'Preferred equity payments = management fees'),
    ('MSP', '550'):  (61001, 'Reimbursable - Travel',        'General reimbursable expenses map to reimbursable series (Employee Reimbursable subledger)'),
    ('MSP', '981'):  (81301, 'Business Taxes',               'Sales tax expense = business taxes'),
    ('MSP', '982'):  (81303, 'State and County Taxes',       'WI fee tax = state/county tax'),

    # ── Cincinnati ────────────────────────────────────────────────────────────
    ('CIN', 'AJ-183'): (20303, 'Payroll Taxes Payable - CIN',       'Federal income tax withholding = payroll taxes payable'),
    ('CIN', 'AJ-205'): (12201, 'Other Short Term Receivables - CIN', 'Intercompany due-from = other short-term receivables (no dedicated intercompany asset in master)'),
    ('CIN', 'AJ-206'): (21203, 'Intercompany Payable - CIN',        'Intercompany due-to = intercompany payable'),
    ('CIN', 'AJ-207'): (90101, 'Intercompany Revenue',              'Intercompany other income = intercompany revenue'),
    ('CIN', 'AJ-90'):  (60201, 'Subconsultant Costs - Out-of-Contract', 'Nonbillable consultant expense = out-of-contract subconsultant cost'),
    ('CIN', 'AJ-107'): (73101, 'Dues & Subscriptions',              'Professional registration & dues = dues & subscriptions'),

    # ── Orlando ───────────────────────────────────────────────────────────────
    ('ORL', 'Accrued_Revenue'):                              (20704, 'Accrued Revenue Liability - ORL',  'Direct match — accrued revenue liability'),
    ('ORL', 'Insurance_Auto_Insurance'):                     (74003, 'Insurance - Auto',                 'Direct match — auto insurance'),
    ('ORL', 'Travel_&_Entertainment_Entertainment'):         (76001, 'Meals & Entertainment',            'Entertainment expense = meals & entertainment (overhead)'),
    ('ORL', 'Car/Truck_Expense_Registration_&_License'):     (73201, 'Licenses & Permits',               'Vehicle registration & license = licenses & permits'),
}

wb = openpyxl.load_workbook(SRC)
ws = wb['Needs Mapping']

MAPPED_FILL = PatternFill('solid', fgColor='E2EFDA')  # light green = done
MAPPED_FONT = Font(size=9, color='375623')

updated = 0
not_found = []

for rn in range(2, ws.max_row + 1):
    office   = str(ws.cell(row=rn, column=1).value or '').strip()
    src_code = str(ws.cell(row=rn, column=2).value or '').strip()

    key = (office, src_code)
    if key not in MAPPINGS:
        not_found.append(f'  No mapping for ({office}, {src_code})')
        continue

    master_code, master_name, rationale = MAPPINGS[key]

    # Write master code and name into the highlighted columns
    ws.cell(row=rn, column=6).value = master_code
    ws.cell(row=rn, column=7).value = master_name
    ws.cell(row=rn, column=8).value = rationale

    # Turn row green
    for col in range(1, 9):
        c = ws.cell(row=rn, column=col)
        c.fill = MAPPED_FILL
        c.font = MAPPED_FONT

    updated += 1

wb.save(DEST)
print(f'Mapped {updated} of {ws.max_row - 1} accounts.')
print(f'Saved: {DEST}')
if not_found:
    print('\nNot found in MAPPINGS dict:')
    for s in not_found:
        print(s)
