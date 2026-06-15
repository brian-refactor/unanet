"""
Unanet template writer — reads from Supabase and writes into the Excel upload templates.

Writes one combined workbook per entity (all offices on separate sheets),
PLUS one workbook per office (all offices merged, ready for upload to that
specific Unanet company — though since it is a single Unanet instance, the
merged file is what gets loaded).

Actually: writes a single merged workbook per entity type with all offices
combined, since all 4 offices load into one Unanet instance.

Usage:
    python etl/write_templates.py              # all entities
    python etl/write_templates.py --entity coa # one entity

Output: output/templates/<entity>_merged.xlsx  (ready to upload to Unanet)

Credentials: SUPABASE_URL and SUPABASE_KEY in etl/.env
"""

import argparse
import os
import shutil
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv(Path(__file__).parent / ".env")

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

TEMPLATE_DIR = Path("Unanet-Migration-Files/Data Upload Templates")
OUTPUT_DIR = Path("output/templates")

OFFICES = ["minnesota", "cincinnati", "dallas", "orlando"]

# ---------------------------------------------------------------------------
# Each entity config: (template_file, sheet_name, data_start_row, column_order)
# column_order = ordered list of DB column names matching template columns left→right
# ---------------------------------------------------------------------------

COA_COLS = [
    "base_code", "base_name", "description", "is_active", "is_1099",
    "is_subcontractor", "financial_type", "subledger_type", "metric_type",
    "cost_type", "pm_type", "labor_revenue_type", "expense_revenue_type",
]

CLIENTS_COLS = [
    "firm_code", "firm_name", "is_active", "website", "client_type",
    "specialty", "note", "pay_days", "main_email", "bill_to_phone",
    "bill_to_street1", "bill_to_street2", "bill_to_street3", "bill_to_street4",
    "bill_to_city", "bill_to_state", "bill_to_zip", "bill_to_country",
    "main_contact_prefix", "main_contact_suffix", "main_contact_title",
    "main_contact_first_name", "main_contact_last_name",
    "main_contact_work_phone", "main_contact_cell_phone",
    "main_contact_work_email", "main_contact_home_email",
]

CONTACTS_COLS = [
    "firm_code", "firm_relationship", "prefix", "suffix", "title",
    "first_name", "last_name", "work_phone", "cell_phone",
    "work_email", "home_email",
    "work_address1", "work_address2", "work_address3", "work_address4",
    "work_city", "work_state", "work_zip", "work_country",
    "home_address1", "home_address2", "home_address3", "home_address4",
    "home_city", "home_state", "home_zip", "home_country",
]

VENDORS_COLS = [
    "firm_code", "firm_name", "is_active", "is_consultant", "consultant_type",
    "is_1099", "website", "vendor_type", "note", "net_days", "ein",
    "pay_to_phone", "pay_to_street1", "pay_to_street2", "pay_to_street3",
    "pay_to_street4", "pay_to_city", "pay_to_state", "pay_to_zip",
    "pay_to_country",
    "main_contact_prefix", "main_contact_suffix", "main_contact_title",
    "main_contact_first_name", "main_contact_last_name",
    "main_contact_work_phone", "main_contact_cell_phone",
    "main_contact_work_email", "main_contact_home_email",
    "enable_eft", "company_id", "company_name", "aba_routing",
    "account_number", "savings", "ef_type_sec",
]

PAY_HISTORY_COLS = [
    "employee_code", "employee_name", "pay_rate", "salary_per_pay_period",
    "pay_rate_start_date", "pay_rate_end_date", "is_hourly", "ot_rate", "otmu",
]

EXPENSE_CODES_COLS = [
    "ec_code", "ec_name", "show_in_es", "is_unit", "unit_type_name",
    "ec_type_name", "exp_markup_type_name", "markup", "unit_rate",
    "bill_status_name",
    "direct_base_code", "direct_base_name",
    "oh_base_code", "oh_base_name",
    "billed_direct_base_code", "billed_direct_base_name",
    "billed_markup_base_code", "billed_markup_base_name",
    "unbilled_base_code", "unbilled_base_name",
    "currency_code", "pm_cmt_required", "int_cmt_required", "is_non_reim",
]

PROJECTS_COLS = [
    "client_firm_code",   # 1  ClientCode
    "owning_org",         # 2  OwningOrg
    "project_code",       # 3  ProjectCode
    "project_name",       # 4  ProjectName
    "charge_type",        # 5  ChargeTypeName
    "start_date",         # 6  StartDate
    "end_date",           # 7  EndDate
    "contract_type",      # 8  ContractTypeName
    "project_note",       # 9  ProjectNote
    "po_number",          # 10 PONumber
    "pm_emp_code",        # 11 ProjectManagerEmpCode
    "pic_emp_code",       # 12 PICEmpCode
    "pa_emp_code",        # 13 ProjectAccountEmpCode
    "billing_term_type",  # 14 BillingTermType
    "net_days",           # 15 NetDays
    None,                 # 16 NextInvNum (not tracked)
    "invoice_email",      # 17 InvoiceEmail
    "location_street1",   # 18 ProjectLocationStreet1
    "location_street2",   # 19 ProjectLocationStreet2
    "location_city",      # 20 ProjectLocationCity
    "location_state",     # 21 ProjectLocationState
    "location_zip",       # 22 ProjectLocationZip
    "location_country",   # 23 ProjectLocationCountry
    "use_client_bill_to", # 24 UseClientBillTo
    "bill_to_street1",    # 25 BillToStreet1
    "bill_to_street2",    # 26 BillToStreet2
    "bill_to_city",       # 27 BillToCity
    "bill_to_state",      # 28 BillToState
    "bill_to_zip",        # 29 BillToZip
    "bill_to_country",    # 30 BillToCountry
]

PHASES_COLS = [
    "project_code",        # 1  LevelOneProjectCode
    "contract_type",       # 2  ContractType
    "level2_name",         # 3  Level2ProjectName
    "level2_code",         # 4  Level2ProjectCode
    "level3_name",         # 5  Level3ProjectName
    "level3_code",         # 6  Level3ProjectCode
    "start_date",          # 7  StartDate
    "end_date",            # 8  EndDate
    "org_path",            # 9  OrgPath
    "fixed_fee",           # 10 FixedFee
    "labor_contract_cap",  # 11 LaborContractCap
    "odc_contract_cap",    # 12 ODContractCap
    "occ_contract_cap",    # 13 OCCContractCap
    "icc_fixed_fee",       # 14 ICCFixedFeePortion
    "labor_budget",        # 15 LaborBudget
    "odc_budget",          # 16 ODCBudget
    "occ_budget",          # 17 OCCBudget
    "icc_budget",          # 18 ICCBudget
    "hours_budget",        # 19 HoursBudget
]

ENTITY_CONFIG = {
    "coa": (
        "02-COA_Fusion.xlsx", "Chart of Accounts", 4, "coa_resolved", COA_COLS,
    ),
    "clients": (
        "03a-Clients_Fusion.xlsx", "Clients", 4, "clients_resolved", CLIENTS_COLS,
    ),
    "client_contacts": (
        "03c-ClientContacts_Fusion.xlsx", "Contacts", 3, "client_contacts_resolved", CONTACTS_COLS,
    ),
    "vendors": (
        "04a-Vendors_Fusion.xlsx", "Vendors", 4, "vendors_resolved", VENDORS_COLS,
    ),
    "vendor_contacts": (
        "04c-VendorContacts_Fusion.xlsx", "Contacts", 3, "vendor_contacts_resolved", CONTACTS_COLS,
    ),
    "pay_history": (
        "05b-PayHistory_Fusion.xlsx", "Employee Pay History", 4, "employees_resolved", PAY_HISTORY_COLS,
    ),
    "expense_codes": (
        "06-ExpenseCodes_Fusion.xlsx", "Expense Codes", 4, "expense_codes_resolved", EXPENSE_CODES_COLS,
    ),
}

ENTITIES = list(ENTITY_CONFIG.keys())


def fetch_all(sb: Client, table: str, cols: list[str]) -> list[dict]:
    select_cols = ",".join(["office"] + cols)
    page_size = 1000
    offset = 0
    all_rows = []
    while True:
        result = (
            sb.table(table)
            .select(select_cols)
            .order("office")
            .range(offset, offset + page_size - 1)
            .execute()
        )
        batch = result.data
        all_rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return all_rows


def coerce_val(val):
    """Normalize None and booleans for Excel."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val  # openpyxl writes TRUE/FALSE correctly
    return val


def write_entity(sb: Client, entity: str) -> None:
    template_file, sheet_name, data_start_row, db_table, db_cols = ENTITY_CONFIG[entity]

    # Fetch from Supabase — Cincinnati employees only if has pay rate
    rows = fetch_all(sb, db_table, db_cols)

    # For pay_history, only include rows that actually have a pay rate
    if entity == "pay_history":
        rows = [r for r in rows if r.get("pay_rate") is not None or r.get("salary_per_pay_period") is not None]

    print(f"  {entity}: {len(rows)} rows from Supabase")

    # Copy template to output
    src = TEMPLATE_DIR / template_file
    out_name = entity + "_merged.xlsx"
    dst = OUTPUT_DIR / out_name
    shutil.copy2(src, dst)

    wb = openpyxl.load_workbook(dst)
    ws = wb[sheet_name]

    # Clear everything from data_start_row down
    max_row = ws.max_row
    if max_row >= data_start_row:
        for row in ws.iter_rows(min_row=data_start_row, max_row=max_row):
            for cell in row:
                cell.value = None

    # Write rows
    for r_idx, row in enumerate(rows, start=data_start_row):
        for c_idx, col in enumerate(db_cols, start=1):
            val = coerce_val(row.get(col))
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(dst)
    print(f"          -> saved {dst}")


def write_open_projects(sb: Client) -> None:
    """Write 07a-OpenProjects — both Projects tab and Phases and Tasks tab."""
    template_file = "07a-OpenProjects_Fusion.xlsx"
    out_name = "open_projects_merged.xlsx"
    src = TEMPLATE_DIR / template_file
    dst = OUTPUT_DIR / out_name

    # ── Projects tab ─────────────────────────────────────────────────────────
    proj_cols = [c for c in PROJECTS_COLS if c is not None]
    select_cols = ",".join(["office"] + proj_cols)
    page, offset, proj_rows = 1000, 0, []
    while True:
        batch = (
            sb.table("projects")
            .select(select_cols)
            .eq("is_active", True)
            .order("office")
            .order("project_code")
            .range(offset, offset + page - 1)
            .execute()
        ).data
        proj_rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    print(f"  open_projects (Projects tab): {len(proj_rows)} rows")

    # ── Phases tab ────────────────────────────────────────────────────────────
    phase_cols = ",".join(["office"] + PHASES_COLS)
    page, offset, phase_rows = 1000, 0, []
    while True:
        batch = (
            sb.table("project_phases")
            .select(phase_cols)
            .order("office")
            .order("project_code")
            .order("level2_code")
            .order("level3_code", nullsfirst=True)
            .range(offset, offset + page - 1)
            .execute()
        ).data
        phase_rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    print(f"  open_projects (Phases tab):   {len(phase_rows)} rows")

    # ── Write workbook ────────────────────────────────────────────────────────
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    DATA_ROW = 4

    def clear_and_write(ws, cols, rows):
        if ws.max_row >= DATA_ROW:
            for row in ws.iter_rows(min_row=DATA_ROW, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
        for r_idx, row in enumerate(rows, start=DATA_ROW):
            for c_idx, col in enumerate(cols, start=1):
                ws.cell(row=r_idx, column=c_idx, value=coerce_val(row.get(col) if col else None))

    clear_and_write(wb["Projects"], PROJECTS_COLS, proj_rows)
    clear_and_write(wb["Phases and Tasks"], PHASES_COLS, phase_rows)

    wb.save(dst)
    print(f"          -> saved {dst}")


def main():
    parser = argparse.ArgumentParser(description="Write Unanet upload templates from Supabase")
    parser.add_argument("--entity", choices=ENTITIES + ["open_projects"], help="Write one entity only")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    all_entities = ENTITIES + ["open_projects"]
    entities = [args.entity] if args.entity else all_entities
    for entity in entities:
        if entity == "open_projects":
            write_open_projects(sb)
        else:
            write_entity(sb, entity)

    print("\nDone. Files in output/templates/")


if __name__ == "__main__":
    main()
