"""
Unanet Data Review App
Run: streamlit run etl/review_app.py

Three tabs:
  1. Browse & Edit   — filter, edit cells inline, changes saved to field_overrides
  2. Duplicates      — detect same firm across offices, record merge decisions
  3. Validation      — flag missing required fields, waive or resolve issues
"""

import difflib
import io
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv(Path(__file__).parent / ".env")

SUPABASE_URL = os.environ.get("SUPABASE_URL") or st.secrets.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY") or st.secrets.get("SUPABASE_KEY", "")

OFFICES = ["minnesota", "cincinnati", "dallas", "orlando", "corporate"]
SOURCE_OFFICES = ["minnesota", "cincinnati", "dallas", "orlando"]  # offices with COA/project data

ENTITY_META = {
    "COA":              {"table": "coa",             "view": "coa_resolved",              "key": "coa_key"},
    "Clients":          {"table": "clients",          "view": "clients_resolved",          "key": "firm_code"},
    "Client Contacts":  {"table": "client_contacts",  "view": "client_contacts_resolved",  "key": "record_key"},
    "Vendors":          {"table": "vendors",          "view": "vendors_resolved",          "key": "firm_code"},
    "Vendor Contacts":  {"table": "vendor_contacts",  "view": "vendor_contacts_resolved",  "key": "record_key"},
    "Employees":        {"table": "employees",        "view": "employees",                 "key": "record_key"},
    "Expense Codes":    {"table": "expense_codes",    "view": "expense_codes_resolved",    "key": "ec_code"},
    "Projects":         {"table": "projects",         "view": "projects",                  "key": "project_code"},
    "Project Phases":   {"table": "project_phases",   "view": "project_phases",            "key": "id"},
}

# Fields that are required in Unanet templates
REQUIRED = {
    "clients":      ["firm_name", "bill_to_street1", "bill_to_city", "bill_to_state", "bill_to_zip"],
    "vendors":      ["firm_name", "pay_to_street1", "pay_to_city", "pay_to_state", "pay_to_zip"],
    "coa":          ["base_code", "base_name", "financial_type"],
    "expense_codes":["ec_code", "ec_name", "show_in_es"],
    "projects":     ["project_code", "project_name", "client_firm_code", "charge_type", "start_date"],
}

# Columns to hide in the editor (internal / not editable)
HIDE_COLS = {"id", "source_id", "has_overrides"}


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def get_sb() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def sb_exec(fn):
    """Run a Supabase callable, retrying once on any connection error."""
    for attempt in range(2):
        try:
            return fn()
        except Exception as e:
            if attempt == 0 and any(w in str(e).lower() for w in ("disconnected", "connection", "timeout")):
                continue
            raise


def fetch(view: str, offices: list[str]) -> pd.DataFrame:
    page, offset, rows = 1000, 0, []
    while True:
        def _q(o=offset):
            sb = get_sb()  # fresh client each call — no stale connections
            q = sb.table(view).select("*").order("office")
            if offices:
                q = q.in_("office", offices)
            return q.range(o, o + page - 1).execute().data
        batch = sb_exec(_q)
        rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def record_key(row: dict, entity_label: str) -> str:
    meta = ENTITY_META[entity_label]
    k = meta["key"]
    if k == "coa_key" or entity_label == "Projects":
        return f"{row.get('office', '')}::{row.get('project_code' if entity_label == 'Projects' else 'base_code', '')}"
    return str(row.get(k, ""))


def save_overrides(table: str, orig: pd.DataFrame, edited: pd.DataFrame,
                   entity_label: str) -> int:
    sb = get_sb()
    changes = []
    skip = HIDE_COLS | {"office", "record_key", "coa_key"}

    for idx in orig.index:
        rk = record_key(orig.loc[idx].to_dict(), entity_label)
        for col in orig.columns:
            if col in skip:
                continue
            ov = orig.at[idx, col]
            nv = edited.at[idx, col]
            ov_s = "" if (ov is None or (isinstance(ov, float) and pd.isna(ov))) else str(ov)
            nv_s = "" if (nv is None or (isinstance(nv, float) and pd.isna(nv))) else str(nv)
            if ov_s != nv_s:
                changes.append({
                    "table_name": table,
                    "record_key": rk,
                    "field_name": col,
                    "original_value": ov_s or None,
                    "new_value":      nv_s or None,
                    "changed_by":     "user",
                })
    if changes:
        sb.table("field_overrides").insert(changes).execute()
    return len(changes)


def fetch_overrides_summary() -> dict:
    """Return {table_name: count} of pending overrides."""
    sb = get_sb()
    rows = sb.table("field_overrides").select("table_name").execute().data
    counts: dict = {}
    for r in rows:
        t = r["table_name"]
        counts[t] = counts.get(t, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Tab 1 — Browse & Edit
# ---------------------------------------------------------------------------

def tab_edit():
    c1, c2 = st.columns([2, 3])
    entity_label = c1.selectbox("Entity", list(ENTITY_META.keys()))
    with c2:
        st.caption("Office")
        off_cols = st.columns(len(OFFICES))
        office_sel = [o for i, o in enumerate(OFFICES)
                      if off_cols[i].checkbox(o.capitalize(), value=True, key=f"edit_off_{o}")]

    meta = ENTITY_META[entity_label]
    load_key = f"df_{entity_label}_{'_'.join(office_sel)}"

    if st.button("Load data", key="load_btn") or load_key in st.session_state:
        if load_key not in st.session_state:
            with st.spinner("Loading from Supabase..."):
                df = fetch(meta["view"], office_sel)
            st.session_state[load_key] = df
            st.session_state[f"orig_{load_key}"] = df.copy()

        df = st.session_state[load_key]

        if df.empty:
            st.info("No data found.")
            return

        # Summary metrics
        m1, m2, m3 = st.columns(3)
        m1.metric("Total rows", len(df))
        if "has_overrides" in df.columns:
            m2.metric("Rows with edits", int(df["has_overrides"].sum()))
        if "is_active" in df.columns:
            m3.metric("Inactive", int((df["is_active"] == False).sum()))

        # Search filter
        search = st.text_input("Search (name / code)", key=f"search_{entity_label}")
        if search:
            mask = df.apply(
                lambda col: col.astype(str).str.contains(search, case=False, na=False)
            ).any(axis=1)
            view_df = df[mask].copy()
        else:
            view_df = df.copy()

        # Display columns — drop internal cols, put has_overrides last
        display_cols = [c for c in view_df.columns if c not in HIDE_COLS]
        if "has_overrides" in view_df.columns:
            display_cols = [c for c in display_cols if c != "has_overrides"] + ["has_overrides"]

        # Colour-code rows with existing overrides
        st.caption(f"Showing {len(view_df):,} rows — yellow rows have been manually edited")

        edited = st.data_editor(
            view_df[display_cols].reset_index(drop=True),
            use_container_width=True,
            num_rows="fixed",
            key=f"editor_{load_key}_{search}",
            column_config={
                "has_overrides": st.column_config.CheckboxColumn("Edited?", disabled=True),
                "is_active":     st.column_config.CheckboxColumn("Active"),
                "is_1099":       st.column_config.CheckboxColumn("1099"),
                "is_consultant": st.column_config.CheckboxColumn("Consultant"),
                "is_hourly":     st.column_config.CheckboxColumn("Hourly"),
                "show_in_es":    st.column_config.CheckboxColumn("Show in ES"),
                "is_unit":       st.column_config.CheckboxColumn("Is Unit"),
                "is_non_reim":   st.column_config.CheckboxColumn("Non-Reimb"),
                "enable_eft":    st.column_config.CheckboxColumn("EFT"),
            },
        )

        col_save, col_reset, col_dl, _ = st.columns([1, 1, 2, 2])

        if col_save.button("Save changes", type="primary"):
            orig = st.session_state[f"orig_{load_key}"]
            # Re-align edited to original index
            orig_display = orig[display_cols].reset_index(drop=True)
            n = save_overrides(meta["table"], orig_display, edited, entity_label)
            if n:
                st.success(f"Saved {n} field change(s) to Supabase.")
                # Invalidate cache so next load reflects overrides
                del st.session_state[load_key]
                del st.session_state[f"orig_{load_key}"]
            else:
                st.info("No changes detected.")

        if col_reset.button("Clear & reload"):
            for k in [load_key, f"orig_{load_key}"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

        offices_label = "_".join(office_sel) if office_sel else "all"
        col_dl.download_button(
            label="Download Unanet template",
            data=build_template_excel(entity_label, view_df),
            file_name=f"{entity_label.replace(' ', '_')}_{offices_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Override audit log
    with st.expander("Override history (all entities)"):
        sb = get_sb()
        rows = sb_exec(lambda: sb.table("field_overrides")
                  .select("*")
                  .order("changed_at", desc=True)
                  .limit(200)
                  .execute().data)
        if rows:
            st.dataframe(pd.DataFrame(rows), use_container_width=True)
        else:
            st.info("No overrides recorded yet.")


# ---------------------------------------------------------------------------
# Tab 2 — Duplicates
# ---------------------------------------------------------------------------

def normalize(name: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())


def tab_duplicates():

    entity_label = st.radio("Entity", ["Clients", "Vendors"], horizontal=True)
    meta = ENTITY_META[entity_label]
    table = meta["table"]

    st.markdown("**Step 1 — Exact name matches across offices**")

    if st.button("Find exact duplicates"):
        with st.spinner("Scanning..."):
            df = fetch(meta["view"], [])
        if df.empty:
            st.info("No data.")
            return

        name_col = "firm_name"
        df["_norm"] = df[name_col].apply(normalize)
        dupes = df[df["_norm"].duplicated(keep=False) & (df["_norm"] != "")]
        dupes = dupes.sort_values(["_norm", "office"])

        st.metric("Candidate duplicate groups", dupes["_norm"].nunique())

        if dupes.empty:
            st.success("No exact name duplicates found.")
        else:
            groups = dupes.groupby("_norm")
            for g_idx, (norm_name, grp) in enumerate(groups):
                grp = grp.reset_index(drop=True)
                with st.expander(f"{grp[name_col].iloc[0]}  ({len(grp)} records)"):
                    key_col = "firm_code"
                    show_cols = [key_col, "office", name_col,
                                 "bill_to_city" if entity_label == "Clients" else "pay_to_city",
                                 "is_active"]
                    show_cols = [c for c in show_cols if c in grp.columns]
                    st.dataframe(grp[show_cols], use_container_width=True, hide_index=True)

                    # Deduplicate codes and generate unique pairs
                    codes = list(dict.fromkeys(grp[key_col].tolist()))  # preserve order, remove dupes
                    for i, keep in enumerate(codes):
                        for j, drop in enumerate(codes):
                            if j <= i:
                                continue
                            if keep == drop:
                                continue
                            btn_key = f"g{g_idx}_i{i}_j{j}"
                            c1, c2, c3 = st.columns([3, 1, 1])
                            c1.write(f"Keep **{keep}** · drop **{drop}**")
                            if c2.button("Approve", key=f"apr_{btn_key}"):
                                _save_merge(table, keep, drop, "approved")
                                st.success(f"Merge decision saved: keep {keep}, drop {drop}")
                            if c3.button("Not a dup", key=f"rej_{btn_key}"):
                                _save_merge(table, keep, drop, "rejected")
                                st.info("Marked as not a duplicate.")

    st.divider()
    st.markdown("**Step 2 — Fuzzy search** (find similar names to a specific firm)")
    search_name = st.text_input("Firm name to search for", key="fuzzy_search")
    threshold = st.slider("Similarity threshold", 0.6, 1.0, 0.85, 0.05)

    if search_name and st.button("Search similar"):
        with st.spinner("Loading and scoring..."):
            df = fetch(meta["view"], [])
        if df.empty:
            return
        norm_target = normalize(search_name)
        df["_score"] = df["firm_name"].apply(
            lambda n: difflib.SequenceMatcher(None, norm_target, normalize(n)).ratio()
        )
        matches = df[df["_score"] >= threshold].sort_values("_score", ascending=False)
        if matches.empty:
            st.info("No similar firms found.")
        else:
            show = [c for c in ["firm_code", "office", "firm_name", "is_active", "_score"] if c in matches.columns]
            st.dataframe(matches[show].reset_index(drop=True), use_container_width=True)

    st.divider()
    st.markdown("**Existing merge decisions**")
    _show_merge_decisions(table)


def _save_merge(entity_type: str, keep: str, drop: str, status: str):
    sb = get_sb()
    sb.table("merge_decisions").upsert({
        "entity_type":    entity_type,
        "firm_code_keep": keep,
        "firm_code_drop": drop,
        "status":         status,
        "decided_by":     "user",
    }, on_conflict="entity_type,firm_code_keep,firm_code_drop").execute()


def _show_merge_decisions(entity_type: str):
    sb = get_sb()
    rows = (sb.table("merge_decisions")
              .select("*")
              .eq("entity_type", entity_type)
              .order("decided_at", desc=True)
              .execute().data)
    if not rows:
        st.info("No merge decisions recorded yet.")
        return
    df = pd.DataFrame(rows)
    st.dataframe(df[["firm_code_keep", "firm_code_drop", "status", "decided_at", "note"]],
                 use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# Tab 3 — Validation
# ---------------------------------------------------------------------------

def tab_validation():

    run_col, _ = st.columns([1, 3])
    if run_col.button("Run validation checks", type="primary"):
        sb = get_sb()
        issues: list[dict] = []

        for entity_label, req_fields in [
            ("Clients", REQUIRED["clients"]),
            ("Vendors", REQUIRED["vendors"]),
            ("COA",     REQUIRED["coa"]),
            ("Expense Codes", REQUIRED["expense_codes"]),
        ]:
            meta = ENTITY_META[entity_label]
            with st.spinner(f"Checking {entity_label}..."):
                df = fetch(meta["view"], [])
            if df.empty:
                continue

            # Determine natural key column for display
            key_col = meta["key"]
            if key_col == "coa_key":
                df["coa_key"] = df["office"].astype(str) + "::" + df["base_code"].astype(str)

            for field in req_fields:
                if field not in df.columns:
                    continue
                missing = df[df[field].isna() | (df[field].astype(str).str.strip() == "")]
                for _, row in missing.iterrows():
                    rk = record_key(row.to_dict(), entity_label)
                    issues.append({
                        "table_name":  meta["table"],
                        "record_key":  rk,
                        "field_name":  field,
                        "issue_type":  "missing_required",
                        "description": f"{entity_label} {rk}: '{field}' is blank (required by Unanet)",
                    })

        # Employees with no pay data (excluding CIN which has history)
        meta_emp = ENTITY_META["Employees"]
        df_emp = fetch(meta_emp["view"], [])
        if not df_emp.empty:
            no_pay = df_emp[df_emp["pay_rate"].isna() & df_emp["salary_per_pay_period"].isna()]
            for _, row in no_pay.iterrows():
                issues.append({
                    "table_name":  "employees",
                    "record_key":  str(row.get("record_key", "")),
                    "field_name":  "pay_rate",
                    "issue_type":  "missing_required",
                    "description": f"Employee {row.get('employee_name', row.get('record_key', ''))} has no pay rate",
                })

        # Upsert issues (skip ones that already exist with same record_key+field)
        if issues:
            existing = sb.table("validation_flags").select("record_key,field_name,table_name").execute().data
            existing_set = {(r["table_name"], r["record_key"], r["field_name"]) for r in existing}
            new_issues = [i for i in issues
                          if (i["table_name"], i["record_key"], i.get("field_name")) not in existing_set]
            if new_issues:
                for chunk in [new_issues[i:i+500] for i in range(0, len(new_issues), 500)]:
                    sb.table("validation_flags").insert(chunk).execute()
            st.success(f"Found {len(issues)} issues — {len(new_issues)} new, {len(issues)-len(new_issues)} already flagged.")
        else:
            st.success("No issues found.")

    st.divider()

    # Show flags with filter
    status_filter = st.selectbox("Show flags with status", ["open", "waived", "resolved", "all"])
    office_filter = st.multiselect("Filter by office (prefix in record_key)", OFFICES, default=[])

    sb = get_sb()
    q = sb.table("validation_flags").select("*").order("created_at", desc=True)
    if status_filter != "all":
        q = q.eq("status", status_filter)
    rows = q.limit(500).execute().data

    if not rows:
        st.info("No validation flags found.")
        return

    df_flags = pd.DataFrame(rows)

    if office_filter:
        pattern = "|".join(office_filter)
        df_flags = df_flags[df_flags["record_key"].str.contains(pattern, case=False, na=False)]

    st.metric("Flags shown", len(df_flags))
    st.dataframe(
        df_flags[["id", "table_name", "record_key", "field_name", "description", "status", "note"]],
        use_container_width=True,
        hide_index=True,
    )

    st.divider()
    st.markdown("**Resolve or waive a flag**")

    flag_id = st.number_input("Flag ID", min_value=1, step=1)
    new_status = st.radio("New status", ["resolved", "waived"], horizontal=True)
    resolution_note = st.text_input("Note (optional)")

    if st.button("Update flag"):
        sb.table("validation_flags").update({
            "status":      new_status,
            "resolved_at": "now()",
            "resolved_by": "user",
            "note":        resolution_note or None,
        }).eq("id", int(flag_id)).execute()
        st.success(f"Flag {flag_id} marked as {new_status}.")
        st.rerun()


# ---------------------------------------------------------------------------
# Tab 4 — COA Mapping
# ---------------------------------------------------------------------------

def load_master_coa() -> pd.DataFrame:
    """Always fetch fresh — called sparingly, cached in session_state by callers."""
    sb = get_sb()
    try:
        rows = sb.table("coa_master").select("*").execute().data
        FIN_ORDER = {"Asset": 0, "Liability": 1, "Equity": 2, "Revenue": 3, "Expense": 4}
        rows.sort(key=lambda r: (
            FIN_ORDER.get(r.get("financial_type", ""), 9),
            str(r.get("master_code", "")),
        ))
        return pd.DataFrame(rows)
    except Exception as e:
        st.error(f"Could not load master COA: {e}")
        return pd.DataFrame()


def get_master_coa() -> pd.DataFrame:
    if "master_coa_df" not in st.session_state or st.session_state["master_coa_df"].empty:
        st.session_state["master_coa_df"] = load_master_coa()
    return st.session_state["master_coa_df"]


def invalidate_master_coa():
    st.session_state.pop("master_coa_df", None)


def get_crosswalk(office: str) -> dict:
    """Returns {source_base_code: master_code} for the given office."""
    key = f"crosswalk_{office}"
    if key not in st.session_state:
        sb = get_sb()
        rows = sb.table("coa_crosswalk").select("source_base_code,master_code").eq("office", office).execute().data
        st.session_state[key] = {r["source_base_code"]: r["master_code"] for r in rows}
    return st.session_state[key]


def invalidate_crosswalk(office: str):
    st.session_state.pop(f"crosswalk_{office}", None)


def upsert_mapping(office: str, source_code: str, source_name: str, master_code: str) -> None:
    get_sb().table("coa_crosswalk").upsert({
        "office": office, "source_base_code": source_code,
        "source_base_name": source_name, "master_code": master_code, "mapped_by": "user",
    }, on_conflict="office,source_base_code").execute()


def delete_mapping(office: str, source_code: str) -> None:
    get_sb().table("coa_crosswalk").delete().eq("office", office).eq("source_base_code", source_code).execute()


def best_master_match(source_name: str, master_df: pd.DataFrame, threshold: float = 0.55) -> str | None:
    best_code, best_score = None, 0.0
    src_norm = normalize(source_name)
    for _, row in master_df.iterrows():
        score = difflib.SequenceMatcher(None, src_norm, normalize(row["master_name"])).ratio()
        if score > best_score:
            best_score, best_code = score, row["master_code"]
    return best_code if best_score >= threshold else None


def tab_coa_mapping():
    st.subheader("Chart of Accounts — Master Mapping")
    subtab_master, subtab_map, subtab_crosswalk = st.tabs(
        ["Master COA Editor", "Mapping Workbench", "Crosswalk Export"]
    )

    # ------------------------------------------------------------------
    # Sub-tab A: Master COA Editor
    # ------------------------------------------------------------------
    with subtab_master:
        st.caption("Define and edit the master chart of accounts. Changes here affect all offices.")

        if st.button("Reload master COA", key="reload_master"):
            invalidate_master_coa()

        master_df = get_master_coa()

        if master_df.empty:
            st.error("Master COA is empty or failed to load. Check your Supabase connection.")
            if st.button("Retry load", key="retry_master"):
                invalidate_master_coa()
                st.rerun()
            return

        st.success(f"{len(master_df)} master accounts loaded.")

        section_filter = st.selectbox(
            "Filter by section",
            ["All"] + sorted(master_df["section"].dropna().unique().tolist()),
            key="master_section_filter",
        )
        disp = master_df if section_filter == "All" else master_df[master_df["section"] == section_filter]
        FIN_ORDER_MAP = {"Asset": 0, "Liability": 1, "Equity": 2, "Revenue": 3, "Expense": 4}
        disp = disp.iloc[disp["financial_type"].map(lambda v: FIN_ORDER_MAP.get(v, 9)).argsort(kind="stable")]

        edit_cols = ["master_code", "master_name", "section", "financial_type",
                     "subledger_type", "metric_type", "cost_type", "pm_type",
                     "is_active", "is_1099", "is_subcontractor", "description", "notes"]
        edit_cols = [c for c in edit_cols if c in disp.columns]

        edited_master = st.data_editor(
            disp[edit_cols].reset_index(drop=True),
            use_container_width=True,
            num_rows="dynamic",
            key=f"master_editor_{section_filter}",
            column_config={
                "is_active":        st.column_config.CheckboxColumn("Active"),
                "is_1099":          st.column_config.CheckboxColumn("1099"),
                "is_subcontractor": st.column_config.CheckboxColumn("Subcon"),
            },
        )

        if st.button("Save master COA changes", type="primary", key="save_master"):
            sb = get_sb()
            saved = 0
            for r in edited_master.to_dict("records"):
                r = {k: (None if (v == "" or (isinstance(v, float) and pd.isna(v))) else v)
                     for k, v in r.items()}
                if r.get("master_code"):
                    sb.table("coa_master").upsert(r, on_conflict="master_code").execute()
                    saved += 1
            invalidate_master_coa()
            st.success(f"Saved {saved} master accounts.")
            st.rerun()

    # ------------------------------------------------------------------
    # Sub-tab B: Mapping Workbench
    # ------------------------------------------------------------------
    with subtab_map:
        c1, c2, c3 = st.columns([2, 2, 2])
        office      = c1.selectbox("Office", SOURCE_OFFICES, key="coa_office")
        fin_filter  = c2.selectbox("Account type", ["All","Asset","Liability","Equity","Revenue","Expense"], key="coa_fin_filter")
        unmapped_only = c3.checkbox("Unmapped only", key="coa_unmapped_only")

        master_df = get_master_coa()
        if master_df.empty:
            st.warning("Load the Master COA first (use the Master COA Editor tab).")
            return

        # Source accounts for this office
        source_rows = get_sb().table("coa_resolved").select(
            "base_code,base_name,financial_type"
        ).eq("office", office).execute().data
        if not source_rows:
            st.info("No COA accounts found for this office.")
            return

        crosswalk = get_crosswalk(office)

        # Coverage bar
        total  = len(source_rows)
        mapped = len(crosswalk)
        m1, m2, m3 = st.columns(3)
        m1.metric("Total", total)
        m2.metric("Mapped", mapped)
        m3.metric("Unmapped", max(0, total - mapped))
        st.progress(min(mapped / total, 1.0) if total else 0)
        st.divider()

        # Apply filters
        display_rows = source_rows
        if fin_filter != "All":
            display_rows = [r for r in display_rows if r.get("financial_type") == fin_filter]
        if unmapped_only:
            display_rows = [r for r in display_rows if r["base_code"] not in crosswalk]

        # Build master option list
        master_options = ["(unmapped)"] + [
            f"{r['master_code']}  {r['master_name']}  [{r['section']}]"
            for _, r in master_df.iterrows()
        ]
        # Map option string → master_code
        opt_to_code = {opt: opt.split("  ")[0] for opt in master_options}
        opt_to_code["(unmapped)"] = None
        code_to_opt  = {v: k for k, v in opt_to_code.items() if v}

        # --- Auto-suggest ---
        col_as, col_save_as, _ = st.columns([2, 2, 4])
        if col_as.button("Auto-suggest all unmapped", key="coa_autosuggest"):
            unmapped_rows = [r for r in source_rows if r["base_code"] not in crosswalk]
            suggestions: dict = {}
            bar = st.progress(0, text="Running fuzzy match...")
            for i, row in enumerate(unmapped_rows):
                code = best_master_match(row["base_name"] or "", master_df)
                if code:
                    suggestions[row["base_code"]] = (code, row["base_name"] or "")
                bar.progress((i + 1) / max(len(unmapped_rows), 1),
                              text=f"Matched {i+1} of {len(unmapped_rows)}")
            bar.empty()
            st.session_state[f"suggestions_{office}"] = suggestions
            st.info(f"Suggested {len(suggestions)} mappings — review in the table below, then click Save Suggestions.")

        suggestions = st.session_state.get(f"suggestions_{office}", {})

        if suggestions:
            if col_save_as.button(f"Save {len(suggestions)} suggestions", type="primary", key="save_suggestions"):
                for src_code, (mcode, sname) in suggestions.items():
                    upsert_mapping(office, src_code, sname, mcode)
                st.session_state.pop(f"suggestions_{office}", None)
                invalidate_crosswalk(office)
                st.success(f"Saved {len(suggestions)} suggested mappings.")
                st.rerun()

        st.divider()

        # --- Build editable mapping dataframe ---
        # Columns: source_code | source_name | type | current_master | suggested
        table_data = []
        for r in display_rows:
            code = r["base_code"]
            cur  = crosswalk.get(code, "")
            sug  = suggestions.get(code, (None, ""))[0] if code in suggestions else None
            table_data.append({
                "source_code":    code,
                "source_name":    r.get("base_name", ""),
                "type":           r.get("financial_type", ""),
                "master_code":    code_to_opt.get(cur, "") if cur else "",
                "suggested":      code_to_opt.get(sug, "") if sug else "",
            })

        map_df    = pd.DataFrame(table_data)
        orig_df   = map_df.copy()
        st.caption("Select a **Master Account** in the dropdown, then click Save Mappings.")

        edited_map = st.data_editor(
            map_df,
            use_container_width=True,
            num_rows="fixed",
            key=f"map_editor_{office}_{fin_filter}_{unmapped_only}",
            disabled=["source_code", "source_name", "type", "suggested"],
            column_config={
                "master_code": st.column_config.SelectboxColumn(
                    "Master Account",
                    options=master_options,
                    required=False,
                ),
                "suggested": st.column_config.TextColumn("Auto-suggest", disabled=True),
            },
        )

        if st.button("Save mappings", type="primary", key="save_map"):
            saved, cleared = 0, 0
            src_name_map = {r["base_code"]: r.get("base_name", "") for r in source_rows}
            for _, row in edited_map.iterrows():
                src_code  = row["source_code"]
                raw_new   = (row["master_code"] or "").strip()
                new_code  = opt_to_code.get(raw_new, raw_new) or ""
                if new_code == "(unmapped)":
                    new_code = ""
                raw_orig  = (orig_df.loc[orig_df["source_code"] == src_code, "master_code"].values[0] or "").strip()
                orig_code = opt_to_code.get(raw_orig, raw_orig) or ""
                if new_code and new_code != orig_code:
                    upsert_mapping(office, src_code, src_name_map.get(src_code, ""), new_code)
                    saved += 1
                elif not new_code and orig_code:
                    delete_mapping(office, src_code)
                    cleared += 1
            invalidate_crosswalk(office)
            msg = []
            if saved:  msg.append(f"{saved} saved")
            if cleared: msg.append(f"{cleared} cleared")
            if msg:
                st.success(f"Mappings updated: {', '.join(msg)}.")
                st.rerun()
            else:
                st.info("No changes detected.")

    # ------------------------------------------------------------------
    # Sub-tab C: Crosswalk Export
    # ------------------------------------------------------------------
    with subtab_crosswalk:
        st.caption("Full crosswalk — every source account and its master assignment.")
        office_filter = st.multiselect("Offices", SOURCE_OFFICES, default=SOURCE_OFFICES, key="xwalk_offices")
        sb = get_sb()
        rows = []
        for off in office_filter:
            rows.extend(sb.table("coa_crosswalk").select("*").eq("office", off).execute().data)

        if not rows:
            st.info("No mappings recorded yet.")
        else:
            xwalk_df  = pd.DataFrame(rows)
            master_df = get_master_coa()
            mlookup   = master_df.set_index("master_code")[
                ["master_name","section","financial_type","metric_type","cost_type"]
            ].to_dict("index")
            for col, attr in [("master_name","master_name"),("section","section"),
                              ("master_fin_type","financial_type"),("metric_type","metric_type"),
                              ("cost_type","cost_type")]:
                xwalk_df[col] = xwalk_df["master_code"].map(lambda c: mlookup.get(c, {}).get(attr, ""))

            show = ["office","source_base_code","source_base_name","master_code",
                    "master_name","section","master_fin_type","metric_type","cost_type","mapped_at","notes"]
            show = [c for c in show if c in xwalk_df.columns]
            st.metric("Mapped accounts", len(xwalk_df))
            st.dataframe(xwalk_df[show].sort_values(["office","master_code"]),
                         use_container_width=True, hide_index=True)
            st.download_button("Download crosswalk CSV",
                               xwalk_df[show].to_csv(index=False),
                               file_name="coa_crosswalk.csv", mime="text/csv")

            st.divider()
            st.markdown("**Unmapped accounts by office**")
            for off in [o for o in office_filter if o in SOURCE_OFFICES]:
                mapped_codes = set(xwalk_df[xwalk_df["office"] == off]["source_base_code"])
                all_src = sb.table("coa_resolved").select("base_code,base_name,financial_type").eq("office", off).execute().data
                unmapped = [r for r in all_src if r["base_code"] not in mapped_codes]
                with st.expander(f"{off.upper()} — {len(unmapped)} unmapped of {len(all_src)}"):
                    if unmapped:
                        st.dataframe(pd.DataFrame(unmapped), use_container_width=True, hide_index=True)
                    else:
                        st.success("All accounts mapped!")


# ---------------------------------------------------------------------------
# Unanet template Excel download
# ---------------------------------------------------------------------------

TEMPLATE_COLS: dict[str, dict] = {
    "COA": {
        "sheet": "Chart of Accounts",
        "headers": ["BaseCode","BaseName","Description","IsActive","Is1099","IsSubcontractor",
                    "FinancialType","SubledgerType","MetricType","CostType","PMType",
                    "LaborRevenueType","ExpenseRevenueType"],
        "db_cols": ["base_code","base_name","description","is_active","is_1099","is_subcontractor",
                    "financial_type","subledger_type","metric_type","cost_type","pm_type",
                    "labor_revenue_type","expense_revenue_type"],
    },
    "Clients": {
        "sheet": "Clients",
        "headers": ["FirmCode","FirmName","IsActive","Website","ClientType","Specialty","Note",
                    "PayDays","MainEmail","BillToAddress_Phone","BillToAddress_Street1",
                    "BillToAddress_Street2","BillToAddress_Street3","BillToAddress_Street4",
                    "BillToAddress_City","BillToAddress_State","BillToAddress_Zip","BillToAddress_Country",
                    "MainContact_Prefix","MainContact_Suffix","MainContact_Title",
                    "MainContact_FirstName","MainContact_LastName","MainContact_WorkPhone",
                    "MainContact_CellPhone","MainContact_WorkEmail","MainContact_HomeEmail"],
        "db_cols": ["firm_code","firm_name","is_active","website","client_type","specialty","note",
                    "pay_days","main_email","bill_to_phone","bill_to_street1","bill_to_street2",
                    "bill_to_street3","bill_to_street4","bill_to_city","bill_to_state","bill_to_zip",
                    "bill_to_country","main_contact_prefix","main_contact_suffix","main_contact_title",
                    "main_contact_first_name","main_contact_last_name","main_contact_work_phone",
                    "main_contact_cell_phone","main_contact_work_email","main_contact_home_email"],
    },
    "Client Contacts": {
        "sheet": "Contacts",
        "headers": ["FirmCode","FirmRelationship","Prefix","Suffix","Title","FirstName","LastName",
                    "WorkPhone","CellPhone","WorkEmail","HomeEmail",
                    "WorkAddress1","WorkAddress2","WorkAddress3","WorkAddress4",
                    "WorkCity","WorkState","WorkZip","WorkCountry",
                    "HomeAddress1","HomeAddress2","HomeAddress3","HomeAddress4",
                    "HomeCity","HomeState","HomeZip","HomeCountry"],
        "db_cols": ["firm_code","firm_relationship","prefix","suffix","title","first_name","last_name",
                    "work_phone","cell_phone","work_email","home_email",
                    "work_address1","work_address2","work_address3","work_address4",
                    "work_city","work_state","work_zip","work_country",
                    "home_address1","home_address2","home_address3","home_address4",
                    "home_city","home_state","home_zip","home_country"],
    },
    "Vendors": {
        "sheet": "Vendors",
        "headers": ["FirmCode","FirmName","IsActive","IsConsultant","ConsultantType","Is1099",
                    "Website","VendorType","Note","NetDays","EIN",
                    "PayToAddress_Phone","PayToAddress_Street1","PayToAddress_Street2",
                    "PayToAddress_Street3","PayToAddress_Street4","PayToAddress_City",
                    "PayToAddress_State","PayToAddress_Zip","PayToAddress_Country",
                    "MainContact_Prefix","MainContact_Suffix","MainContact_Title",
                    "MainContact_FirstName","MainContact_LastName","MainContact_WorkPhone",
                    "MainContact_CellPhone","MainContact_WorkEmail","MainContact_HomeEmail"],
        "db_cols": ["firm_code","firm_name","is_active","is_consultant","consultant_type","is_1099",
                    "website","vendor_type","note","net_days","ein",
                    "pay_to_phone","pay_to_street1","pay_to_street2","pay_to_street3","pay_to_street4",
                    "pay_to_city","pay_to_state","pay_to_zip","pay_to_country",
                    "main_contact_prefix","main_contact_suffix","main_contact_title",
                    "main_contact_first_name","main_contact_last_name","main_contact_work_phone",
                    "main_contact_cell_phone","main_contact_work_email","main_contact_home_email"],
    },
    "Vendor Contacts": {
        "sheet": "Contacts",
        "headers": ["FirmCode","FirmRelationship","Prefix","Suffix","Title","FirstName","LastName",
                    "WorkPhone","CellPhone","WorkEmail","HomeEmail",
                    "WorkAddress1","WorkAddress2","WorkAddress3","WorkAddress4",
                    "WorkCity","WorkState","WorkZip","WorkCountry",
                    "HomeAddress1","HomeAddress2","HomeAddress3","HomeAddress4",
                    "HomeCity","HomeState","HomeZip","HomeCountry"],
        "db_cols": ["firm_code","firm_relationship","prefix","suffix","title","first_name","last_name",
                    "work_phone","cell_phone","work_email","home_email",
                    "work_address1","work_address2","work_address3","work_address4",
                    "work_city","work_state","work_zip","work_country",
                    "home_address1","home_address2","home_address3","home_address4",
                    "home_city","home_state","home_zip","home_country"],
    },
    "Employees": {
        "sheet": "Employee Pay History",
        "headers": ["EmployeeCode","EmployeeName","PayRate","salaryperpayperiod",
                    "PayRateStartDate","PayRateEndDate","IsHourly","OTRate","OTMU"],
        "db_cols": ["employee_code","employee_name","pay_rate","salary_per_pay_period",
                    "pay_rate_start_date","pay_rate_end_date","is_hourly","ot_rate","otmu"],
    },
    "Expense Codes": {
        "sheet": "Expense Codes",
        "headers": ["ECCode","ECName","ShowInES","IsUnit","UnitTypename","ECTypename",
                    "ExpMarkupTypename","Markup","UnitRate","BillStatusname",
                    "DirectBaseCode","DirectBasename","OHBaseCode","OHBasename",
                    "BilledDirectBaseCode","BilledDirectBasename","BilledMarkupBaseCode",
                    "BilledMarkupBasename","UnBilledBaseCode","UnBilledBasename",
                    "CurrencyCode","PMCmtRequired","IntCmtRequired","IsNonReim"],
        "db_cols": ["ec_code","ec_name","show_in_es","is_unit","unit_type_name","ec_type_name",
                    "exp_markup_type_name","markup","unit_rate","bill_status_name",
                    "direct_base_code","direct_base_name","oh_base_code","oh_base_name",
                    "billed_direct_base_code","billed_direct_base_name","billed_markup_base_code",
                    "billed_markup_base_name","unbilled_base_code","unbilled_base_name",
                    "currency_code","pm_cmt_required","int_cmt_required","is_non_reim"],
    },
    "Project Phases": {
        "sheet": "Phases and Tasks",
        "headers": ["LevelOneProjectCode","ContractType","Level2ProjectName","Level2ProjectCode",
                    "Level3ProjectName","Level3ProjectCode","StartDate","EndDate","OrgPath",
                    "FixedFee","LaborContractCap","ODContractCap","OCCContractCap",
                    "ICCFixedFeePortion","LaborBudget","ODCBudget","OCCBudget","ICCBudget","HoursBudget"],
        "db_cols": ["project_code","contract_type","level2_name","level2_code",
                    "level3_name","level3_code","start_date","end_date","org_path",
                    "fixed_fee","labor_contract_cap","odc_contract_cap","occ_contract_cap",
                    "icc_fixed_fee","labor_budget","odc_budget","occ_budget","icc_budget","hours_budget"],
    },
    "Projects": {
        "sheet": "Projects",
        "headers": ["ClientCode","OwningOrg","ProjectCode","ProjectName","ChargeTypeName",
                    "StartDate","EndDate","ContractTypeName","ProjectNote","PONumber",
                    "ProjectManagerEmpCode","PICEmpCode","ProjectAccountEmpCode",
                    "BillingTermType","NetDays","NextInvNum","InvoiceEmail",
                    "ProjectLocationStreet1","ProjectLocationStreet2","ProjectLocationCity",
                    "ProjectLocationState","ProjectLocationZip","ProjectLocationCountry",
                    "UseClientBillTo","BillToStreet1","BillToStreet2","BillToCity",
                    "BillToState","BillToZip","BillToCountry"],
        "db_cols": ["client_firm_code","owning_org","project_code","project_name","charge_type",
                    "start_date","end_date","contract_type","project_note","po_number",
                    "pm_emp_code","pic_emp_code","pa_emp_code",
                    "billing_term_type","net_days",None,"invoice_email",
                    "location_street1","location_street2","location_city",
                    "location_state","location_zip","location_country",
                    "use_client_bill_to","bill_to_street1","bill_to_street2","bill_to_city",
                    "bill_to_state","bill_to_zip","bill_to_country"],
    },
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


def build_template_excel(entity_label: str, df: pd.DataFrame) -> bytes:
    """Return an in-memory Excel file in Unanet upload template format."""
    config = TEMPLATE_COLS.get(entity_label)
    if not config:
        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        return buf.getvalue()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["sheet"]

    # Header row styled to match Unanet template
    for c, h in enumerate(config["headers"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col in enumerate(config["db_cols"], 1):
            val = row.get(col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = None
            ws.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_open_projects_excel(projects_df: pd.DataFrame, phases_df: pd.DataFrame) -> bytes:
    """Build 07a-OpenProjects workbook with Projects tab + Phases and Tasks tab."""
    wb = openpyxl.Workbook()

    def write_sheet(ws, config, df):
        ws.title = config["sheet"]
        for c, h in enumerate(config["headers"], 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, col in enumerate(config["db_cols"], 1):
                val = row.get(col) if col else None
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    val = None
                ws.cell(row=r_idx, column=c_idx, value=val)

    ws_proj = wb.active
    write_sheet(ws_proj, TEMPLATE_COLS["Projects"], projects_df)
    ws_phases = wb.create_sheet()
    write_sheet(ws_phases, TEMPLATE_COLS["Project Phases"], phases_df)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tab 5 — Projects
# ---------------------------------------------------------------------------

def tab_projects():
    c1, c2 = st.columns([3, 2])
    with c1:
        st.caption("Office")
        off_cols = st.columns(len(OFFICES))
        office_sel = [o for i, o in enumerate(OFFICES)
                      if off_cols[i].checkbox(o.capitalize(), value=True, key=f"proj_off_{o}")]
    status_sel = c2.radio("Status", ["Active only", "All", "Inactive only"],
                          horizontal=True, index=0, key="proj_status")

    if not office_sel:
        st.info("Select at least one office.")
        return

    cache_key = f"proj_df_{'_'.join(sorted(office_sel))}"

    if st.button("Load / Refresh", key="proj_load") or cache_key in st.session_state:
        if cache_key not in st.session_state:
            with st.spinner(f"Loading projects for {', '.join(office_sel)}..."):
                df = fetch("projects", office_sel)
            st.session_state[cache_key] = df

        df = st.session_state[cache_key]

        if df.empty:
            st.info("No projects found.")
            return

        # Per-office summary tiles
        off_list = sorted(df["office"].unique().tolist()) if "office" in df.columns else []
        if off_list:
            m_cols = st.columns(len(off_list))
            for i, off in enumerate(off_list):
                sub = df[df["office"] == off]
                active = int(sub["is_active"].sum()) if "is_active" in sub.columns else len(sub)
                no_cli = int(
                    (sub["client_firm_code"].isna() | sub["client_firm_code"].eq("")).sum()
                ) if "client_firm_code" in sub.columns else 0
                no_pm = int(sub["pm_emp_code"].isna().sum()) if "pm_emp_code" in sub.columns else 0
                m_cols[i].metric(off[:3].upper(), f"{active} active / {len(sub)} total")
                m_cols[i].caption(f"No client: {no_cli}  |  No PM: {no_pm}")

        st.divider()

        # Apply status filter
        view_df = df.copy()
        if "is_active" in view_df.columns:
            if status_sel == "Active only":
                view_df = view_df[view_df["is_active"] == True]
            elif status_sel == "Inactive only":
                view_df = view_df[view_df["is_active"] == False]

        sf1, sf2, sf3 = st.columns([3, 1, 1])
        search = sf1.text_input("Search code / name / client", key="proj_search")
        missing_cli = sf2.checkbox("No client code", key="proj_no_cli")
        missing_pm  = sf3.checkbox("No PM code", key="proj_no_pm")

        if search:
            mask = view_df.apply(
                lambda col: col.astype(str).str.contains(search, case=False, na=False)
            ).any(axis=1)
            view_df = view_df[mask]
        if missing_cli and "client_firm_code" in view_df.columns:
            view_df = view_df[
                view_df["client_firm_code"].isna() | (view_df["client_firm_code"] == "")
            ]
        if missing_pm and "pm_emp_code" in view_df.columns:
            view_df = view_df[view_df["pm_emp_code"].isna()]

        st.caption(f"Showing {len(view_df):,} of {len(df):,} projects")

        display_cols = [c for c in view_df.columns if c not in HIDE_COLS]
        st.dataframe(
            view_df[display_cols].reset_index(drop=True),
            use_container_width=True,
            column_config={
                "is_active":          st.column_config.CheckboxColumn("Active"),
                "use_client_bill_to": st.column_config.CheckboxColumn("Use Client Bill-To"),
            },
            hide_index=True,
        )

        offices_label = "_".join(sorted(office_sel)) if office_sel else "all"
        st.download_button(
            label="Download 07a upload template",
            data=build_template_excel("Projects", view_df),
            file_name=f"projects_{offices_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="proj_dl",
        )
        st.caption("Phases & Tasks sheet will be blank until WBS data is loaded.")


# ---------------------------------------------------------------------------
# Tab 6 — Project Phases
# ---------------------------------------------------------------------------

def tab_phases():
    c1, c2 = st.columns([3, 2])
    with c1:
        st.caption("Office")
        off_cols = st.columns(len(SOURCE_OFFICES))
        office_sel = [o for i, o in enumerate(SOURCE_OFFICES)
                      if off_cols[i].checkbox(o.capitalize(), value=True, key=f"ph_off_{o}")]

    if not office_sel:
        st.info("Select at least one office.")
        return

    cache_key = f"phases_df_{'_'.join(sorted(office_sel))}"

    if st.button("Load / Refresh", key="ph_load") or cache_key in st.session_state:
        if cache_key not in st.session_state:
            with st.spinner("Loading phases from Supabase..."):
                df = fetch("project_phases", office_sel)
            st.session_state[cache_key] = df

        df = st.session_state[cache_key]
        if df.empty:
            st.info("No phase data found.")
            return

        # Contract type filter — built dynamically from actual data values
        all_contract_types = sorted(df["contract_type"].dropna().unique()) if "contract_type" in df.columns else []
        contract_sel = c2.multiselect("Contract Type", all_contract_types,
                                      default=all_contract_types, key="ph_contract")

        # Apply contract type filter
        view_df = df.copy()
        if contract_sel and "contract_type" in view_df.columns:
            view_df = view_df[view_df["contract_type"].isin(contract_sel)]

        # Summary metrics
        offices_in_data = sorted(view_df["office"].unique()) if "office" in view_df.columns else []
        if offices_in_data:
            m_cols = st.columns(len(offices_in_data))
            for i, off in enumerate(offices_in_data):
                sub = view_df[view_df["office"] == off]
                proj_count = sub["project_code"].nunique() if "project_code" in sub.columns else 0
                m_cols[i].metric(off[:3].upper(), f"{len(sub):,} phases")
                m_cols[i].caption(f"{proj_count} projects")

        st.divider()

        # Phase distribution
        if "level2_name" in view_df.columns:
            with st.expander("Phase distribution", expanded=False):
                dist = (view_df.groupby(["level2_name", "contract_type"])
                               .size()
                               .reset_index(name="count")
                               .sort_values("count", ascending=False))
                st.dataframe(dist, use_container_width=True, hide_index=True)

        # Search
        sf1, sf2 = st.columns([3, 2])
        search     = sf1.text_input("Search project code / phase name", key="ph_search")
        has_l3     = sf2.checkbox("L3 sub-phases only", key="ph_has_l3")

        if search:
            mask = view_df.apply(
                lambda col: col.astype(str).str.contains(search, case=False, na=False)
            ).any(axis=1)
            view_df = view_df[mask]
        if has_l3:
            view_df = view_df[view_df["level3_name"].notna() & (view_df["level3_name"] != "")]

        hide = {"id", "office"}
        display_cols = [c for c in view_df.columns if c not in hide]

        st.caption(f"Showing {len(view_df):,} phase rows")
        st.dataframe(
            view_df[display_cols].reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
            column_config={
                "fixed_fee":          st.column_config.NumberColumn("Fixed Fee", format="$%.2f"),
                "labor_contract_cap": st.column_config.NumberColumn("Labor Cap", format="$%.2f"),
                "icc_fixed_fee":      st.column_config.NumberColumn("ICC Fee", format="$%.2f"),
                "labor_budget":       st.column_config.NumberColumn("Labor Budget", format="$%.2f"),
                "hours_budget":       st.column_config.NumberColumn("Hours Budget", format="%.1f"),
            },
        )

        dl1, dl2, _ = st.columns([2, 2, 3])
        dl1.download_button(
            label="Download phases CSV",
            data=view_df[display_cols].to_csv(index=False),
            file_name=f"project_phases_{'_'.join(sorted(office_sel))}.csv",
            mime="text/csv",
            key="ph_dl_csv",
        )
        dl2.download_button(
            label="Download phases Excel",
            data=build_template_excel("Project Phases", view_df),
            file_name=f"project_phases_{'_'.join(sorted(office_sel))}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ph_dl_xlsx",
        )


# ---------------------------------------------------------------------------
# Tab 7 — Export
# ---------------------------------------------------------------------------

def tab_export():
    st.markdown("Generate Unanet upload-ready Excel files. Each download pulls live data from Supabase.")
    st.divider()

    with st.expander("Office filter", expanded=True):
        off_cols = st.columns(len(OFFICES))
        office_sel = [o for i, o in enumerate(OFFICES)
                      if off_cols[i].checkbox(o.capitalize(), value=True, key=f"exp_off_{o}")]
    if not office_sel:
        st.warning("Select at least one office.")
        return

    offices_label = "_".join(sorted(office_sel))

    # Helper: fetch + build + return bytes
    def _fetch_bytes(entity_label: str) -> bytes:
        meta = ENTITY_META[entity_label]
        df = fetch(meta["view"], office_sel)
        if entity_label == "Projects":
            df = df[df.get("is_active", pd.Series([True]*len(df)))==True] if "is_active" in df.columns else df
        return build_template_excel(entity_label, df)

    # ── 07a Open Projects (combined) ─────────────────────────────────────────
    st.subheader("07a — Open Projects")
    st.caption("Generates a two-tab workbook: Projects + Phases and Tasks")

    if st.button("Build 07a export", type="primary", key="exp_07a_build"):
        with st.spinner("Fetching projects and phases..."):
            proj_df   = fetch("projects", office_sel)
            phases_df = fetch("project_phases", office_sel)
            if "is_active" in proj_df.columns:
                proj_df = proj_df[proj_df["is_active"] == True]
            proj_df   = proj_df.sort_values(["office", "project_code"])
            phases_df = phases_df.sort_values(["office", "project_code", "level2_code"])
            data = build_open_projects_excel(proj_df, phases_df)
        st.session_state["export_07a_data"] = data
        st.session_state["export_07a_label"] = offices_label
        st.success(f"Built: {len(proj_df):,} projects, {len(phases_df):,} phase rows.")

    if "export_07a_data" in st.session_state:
        st.download_button(
            label="Download 07a-OpenProjects_merged.xlsx",
            data=st.session_state["export_07a_data"],
            file_name=f"07a-OpenProjects_{st.session_state['export_07a_label']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="exp_07a_dl",
        )

    st.divider()

    # ── All other entities ───────────────────────────────────────────────────
    EXPORT_ENTITIES = [
        ("02 — Chart of Accounts",  "COA",             "02-COA"),
        ("03a — Clients",           "Clients",          "03a-Clients"),
        ("03c — Client Contacts",   "Client Contacts",  "03c-ClientContacts"),
        ("04a — Vendors",           "Vendors",          "04a-Vendors"),
        ("04c — Vendor Contacts",   "Vendor Contacts",  "04c-VendorContacts"),
        ("06 — Expense Codes",      "Expense Codes",    "06-ExpenseCodes"),
    ]

    st.subheader("Other Upload Templates")
    cols = st.columns(2)
    for i, (label, entity_label, file_prefix) in enumerate(EXPORT_ENTITIES):
        with cols[i % 2]:
            st.markdown(f"**{label}**")
            if st.button(f"Build {file_prefix}", key=f"exp_build_{entity_label}"):
                with st.spinner(f"Fetching {entity_label}..."):
                    data = _fetch_bytes(entity_label)
                st.session_state[f"export_{entity_label}"] = (data, offices_label)
            if f"export_{entity_label}" in st.session_state:
                d, lbl = st.session_state[f"export_{entity_label}"]
                st.download_button(
                    label=f"Download {file_prefix}_{lbl}.xlsx",
                    data=d,
                    file_name=f"{file_prefix}_{lbl}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"exp_dl_{entity_label}",
                )
            st.divider()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    st.set_page_config(
        page_title="Unanet Data Review",
        page_icon="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg'/>",
        layout="wide",
    )

    # Password gate — set APP_PASSWORD in .streamlit/secrets.toml or Streamlit Cloud secrets
    _app_pw = st.secrets.get("APP_PASSWORD", "") if hasattr(st, "secrets") else ""
    if _app_pw:
        entered = st.text_input("Enter password to continue", type="password", key="_gate_pw")
        if entered != _app_pw:
            st.stop()

    st.markdown("### Unanet Migration — Data Review")
    st.divider()

    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(
        ["Browse & Edit", "Duplicates", "Validation", "COA Mapping", "Projects", "Project Phases", "Export"]
    )
    with tab1:
        tab_edit()
    with tab2:
        tab_duplicates()
    with tab3:
        tab_validation()
    with tab4:
        tab_coa_mapping()
    with tab5:
        tab_projects()
    with tab6:
        tab_phases()
    with tab7:
        tab_export()


if __name__ == "__main__":
    main()
