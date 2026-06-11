"""
Export the Unanet COA upload file from coa_master in Supabase.
Reads all 389 master accounts and writes them into the Unanet COA template.

Output: output/COA_Upload_<YYYYMMDD>.xlsx
"""
import os, shutil
from datetime import date
from pathlib import Path

from dotenv import load_dotenv
from supabase import create_client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv(Path(__file__).parent / '.env')
sb = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])

TEMPLATE  = Path(r'E:\Unanet\Documentation\OneDrive_1_4-24-2026\Data Upload Templates\02-COA_Fusion.xlsx')
OUTPUT    = Path(r'E:\Unanet\output') / f'COA_Upload_{date.today().strftime("%Y%m%d")}.xlsx'
DATA_ROW  = 4   # first data row in the template (rows 1-2: headers, row 3: field names)

# Column order matches the template left→right
COLUMNS = [
    'master_code', 'master_name', 'description',
    'is_active', 'is_1099', 'is_subcontractor',
    'financial_type', 'subledger_type', 'metric_type',
    'cost_type', 'pm_type', 'labor_revenue_type', 'expense_revenue_type',
]

# ── Pull from Supabase ────────────────────────────────────────────────────────
print('Fetching coa_master from Supabase...')
rows, offset, page = [], 0, 1000
while True:
    batch = sb.table('coa_master').select('*').order('sort_order').range(offset, offset + page - 1).execute().data
    rows.extend(batch)
    if len(batch) < page:
        break
    offset += page
print(f'  {len(rows)} accounts retrieved')

# ── Write into template ───────────────────────────────────────────────────────
shutil.copy(TEMPLATE, OUTPUT)
wb = openpyxl.load_workbook(OUTPUT)
ws = wb['Chart of Accounts']

# Clear any existing sample/demo data rows (row 4 onward)
for rn in range(DATA_ROW, ws.max_row + 1):
    for col in range(1, 14):
        ws.cell(row=rn, column=col).value = None

# Write data
for i, row in enumerate(rows):
    rn = DATA_ROW + i
    for col_idx, field in enumerate(COLUMNS, 1):
        val = row.get(field)
        # Convert booleans to True/False (Excel-native)
        if isinstance(val, bool):
            pass  # keep as bool
        elif val == '':
            val = None
        ws.cell(row=rn, column=col_idx).value = val

wb.save(OUTPUT)
print(f'Saved: {OUTPUT}')
print(f'  {len(rows)} accounts written starting at row {DATA_ROW}')

# ── Quick validation ─────────────────────────────────────────────────────────
wb2   = openpyxl.load_workbook(OUTPUT)
ws2   = wb2['Chart of Accounts']
data  = [r for r in ws2.iter_rows(min_row=DATA_ROW, values_only=True) if r[0]]
blank_ft  = [r for r in data if not r[6]]
blank_name = [r for r in data if not r[1]]
print(f'\nValidation:')
print(f'  Total rows:            {len(data)}')
print(f'  Missing FinancialType: {len(blank_ft)}')
print(f'  Missing BaseName:      {len(blank_name)}')
if blank_ft:
    for r in blank_ft:
        print(f'    {r[0]}  {r[1]}')
