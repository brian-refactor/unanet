"""
Ajera Employee Utilization Report — API Version
Pulls timesheet data via Ajera v2 API (trailing 12 months) without a manual export.

Usage:
    python etl/ajera_utilization_api.py

Output:
    output/cincinnati/cincinnati_utilization_api_<MonthYYYY>.xlsx
"""

import csv
import os
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE          = Path(__file__).parent
OUTPUT_DIR    = HERE.parent / 'output' / 'cincinnati'
PAY_RATES_CSV     = OUTPUT_DIR / 'cincinnati_Employees.csv'
BILLING_RATES_CSV = OUTPUT_DIR / 'cincinnati_billing_rates.csv'

HEADERS       = {'Content-Type': 'application/json'}
BATCH_SIZE    = 20
DEFAULT_TARGET = 85.0

# Overhead label (lowercase) → category key
OVERHEAD_CAT = {
    'vacation':             'vacation',
    'holiday':              'holiday',
    'sick':                 'sick',
    'marketing':            'marketing',
    'office meetings':      'meetings',
    'meetings':             'meetings',
    'continuing education':      'cont_ed',
    'm -bereavement':            'other',
    'maternity/paternity leave': 'other',
    'general':              'admin',
    'it':                   'admin',
    'administration':       'admin',
    'admin':                'admin',
}

CAT_KEYS   = ['admin', 'marketing', 'vacation', 'meetings', 'holiday', 'sick', 'cont_ed', 'other']
CAT_LABELS = {
    'admin': 'Administration', 'marketing': 'Marketing',
    'vacation': 'Vacation',    'meetings': 'Meetings',
    'holiday': 'Holiday',      'sick': 'Sick',
    'cont_ed': 'Cont. Education', 'other': 'Other',
}

_DARK  = '1F4E79'; _BLUE  = '2E75B6'; _LBLUE = 'DEEAF1'
_GREEN = 'E2EFDA'; _AMBER = 'FFF2CC'; _RED   = 'FCE4D6'
_RFONT = 'C00000'; _WHITE = 'FFFFFF'


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def _trail_dates():
    """Trailing 12 months: (trail_start, cur_month_start, today)."""
    today = date.today()
    cur   = today.replace(day=1)
    m, y  = cur.month - 11, cur.year
    if m <= 0:
        m += 12; y -= 1
    return date(y, m, 1), cur, today


def _last_quarter_dates():
    """Last fully-completed calendar quarter: (lq_start, lq_end, label e.g. 'Q1 2026')."""
    today   = date.today()
    cur_q   = (today.month - 1) // 3 + 1   # 1-4
    lq      = cur_q - 1 if cur_q > 1 else 4
    lq_year = today.year if cur_q > 1 else today.year - 1
    lq_start_month = (lq - 1) * 3 + 1
    lq_end_month   = lq * 3
    # last day of end month
    import calendar
    lq_end_day = calendar.monthrange(lq_year, lq_end_month)[1]
    return (date(lq_year, lq_start_month, 1),
            date(lq_year, lq_end_month, lq_end_day),
            f'Q{lq} {lq_year}')


def _parse_date(s):
    try:
        return datetime.fromisoformat(str(s)[:10]).date()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def _connect(api_url, username, password):
    resp = requests.post(api_url, json={
        'Method': 'CreateAPISession', 'Username': username,
        'Password': password, 'APIVersion': 2, 'UseSessionCookie': False,
    }, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data  = resp.json()
    token = data.get('Content', {}).get('SessionToken')
    if not token:
        raise SystemExit(f'Auth failed: {data.get("Errors")}')
    info = data['Content']
    print(f'Connected — {info.get("CompanyName")} (Ajera {info.get("AjeraVersion")})')
    return token


def _disconnect(api_url, token):
    try:
        requests.post(api_url, json={'Method': 'EndAPISession', 'SessionToken': token},
                      headers=HEADERS, timeout=15)
    except Exception:
        pass


def _call(api_url, token, method, args=None):
    payload = {'Method': method, 'SessionToken': token, 'MethodArguments': args or {}}
    data    = requests.post(api_url, json=payload, headers=HEADERS, timeout=90).json()
    return data.get('Content', {})


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def _entry_hours(entry):
    return sum((entry.get(f'D{i}') or 0) for i in range(1, 8))


def fetch_details(api_url, token, trail_start):
    """Return list of GetTimesheets detail records for trailing 12 months."""

    # 1. Discover active employee keys from most-recent batch
    print('  Discovering employees...')
    content   = _call(api_url, token, 'ListTimesheets')
    emp_keys  = list({t['Employee Key'] for t in content.get('Timesheets', [])
                      if t.get('Employee Key')})
    print(f'  Found {len(emp_keys)} employees')

    # 2. Per-employee timesheet key collection
    print(f'  Collecting timesheet keys since {trail_start}...')
    all_ts_keys = set()
    for ek in emp_keys:
        content = _call(api_url, token, 'ListTimesheets', {
            'FilterByEmployee': [ek],
            'FilterByEarliestTimesheetDate': trail_start.isoformat(),
        })
        for t in content.get('Timesheets', []):
            if t.get('Timesheet Key'):
                all_ts_keys.add(t['Timesheet Key'])
    print(f'  {len(all_ts_keys)} timesheets in period')

    # 3. GetTimesheets detail in batches
    print('  Fetching detail...')
    keys    = list(all_ts_keys)
    details = []
    for i in range(0, len(keys), BATCH_SIZE):
        batch = keys[i:i + BATCH_SIZE]
        content = _call(api_url, token, 'GetTimesheets', {'RequestedTimesheets': batch})
        details.extend(content.get('Timesheets', []))
        print(f'    {min(i + BATCH_SIZE, len(keys))}/{len(keys)}', end='\r')
    print()
    return details


# ---------------------------------------------------------------------------
# Data parsing
# ---------------------------------------------------------------------------

def _empty_hours():
    return {'billable': 0.0, 'indirect': 0.0, 'total': 0.0,
            **{k: 0.0 for k in CAT_KEYS}}


def parse_details(details, cur_month_start, lq_start, lq_end):
    """Aggregate detail records into per-employee dicts matching old report format."""
    emps = {}
    unknown_overhead = defaultdict(float)  # label → total hours (unmapped categories)

    for ts in details:
        ek     = ts.get('EmployeeKey')
        name   = ts.get('Employee', '')
        status = ts.get('EmployeeStatus', 'Active')
        target = float(ts.get('TargetBillablePercent') or DEFAULT_TARGET)
        ts_dt  = _parse_date(ts.get('TimesheetDate', ''))
        in_cur = (ts_dt and ts_dt.year == cur_month_start.year
                  and ts_dt.month == cur_month_start.month)
        in_lq  = (ts_dt and lq_start <= ts_dt <= lq_end)

        if ek not in emps:
            emps[ek] = {
                'name':            name,
                'type':            '',
                'status':          status,
                'hire_date':       '',
                'target_pct':      target,
                'ytd_hours':       _empty_hours(),
                'lq_hours':        _empty_hours(),
                'current_hours':   _empty_hours(),
                'ytd_amounts':     _empty_hours(),
                'current_amounts': _empty_hours(),
            }
        emp = emps[ek]

        # Project rows → billable
        for entry in (ts.get('Project') or {}).get('Detail', []):
            h = _entry_hours(entry)
            if h <= 0:
                continue
            if not emp['type'] and entry.get('Employee Type'):
                emp['type'] = entry['Employee Type']
            emp['ytd_hours']['billable'] += h
            emp['ytd_hours']['total']    += h
            if in_lq:
                emp['lq_hours']['billable'] += h
                emp['lq_hours']['total']    += h
            if in_cur:
                emp['current_hours']['billable'] += h
                emp['current_hours']['total']    += h

        # Overhead rows → categorized indirect
        for entry in (ts.get('Overhead') or {}).get('Detail', []):
            h = _entry_hours(entry)
            if h <= 0:
                continue
            label = str(entry.get('Timesheet Overhead Group Detail', '')).lower().strip()
            cat   = OVERHEAD_CAT.get(label, 'other')
            if cat == 'other' and label and label not in OVERHEAD_CAT:
                unknown_overhead[label] += h
            emp['ytd_hours'][cat]        += h
            emp['ytd_hours']['indirect'] += h
            emp['ytd_hours']['total']    += h
            if in_lq:
                emp['lq_hours'][cat]        += h
                emp['lq_hours']['indirect'] += h
                emp['lq_hours']['total']    += h
            if in_cur:
                emp['current_hours'][cat]        += h
                emp['current_hours']['indirect'] += h
                emp['current_hours']['total']    += h

    if unknown_overhead:
        print('\n  Overhead labels mapped to "other" (review for recategorization):')
        for label, hours in sorted(unknown_overhead.items(), key=lambda x: -x[1]):
            print(f'    {hours:>8.1f}h  "{label}"')

    return list(emps.values())


# ---------------------------------------------------------------------------
# Pay / billing rate helpers  (same logic as ajera_utilization.py)
# ---------------------------------------------------------------------------

def _norm(s):
    s = s.lstrip('*').strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\.', '', s)
    return re.sub(r'\s+', ' ', s).lower().strip()

def _fl(n):
    p = n.split()
    return f'{p[0]} {p[-1]}' if len(p) >= 2 else n


def load_pay_rates(path):
    if not path.exists():
        return {}, {}
    today = date.today()
    cur   = {}
    for row in csv.DictReader(open(path, encoding='utf-8')):
        end = row.get('PayRateEndDate', '').strip()
        ok  = not end
        if not ok:
            try: ok = date.fromisoformat(end[:10]) > today
            except ValueError: pass
        if not ok:
            continue
        try:
            rate = float(row['PayRate']) if row['PayRate'] else 0.0
            spp  = float(row.get('salaryperpayperiod') or 0)
        except ValueError:
            continue
        is_h   = row.get('IsHourly', '').strip().upper() == 'TRUE'
        annual = rate * 2080 if is_h else spp * 26
        cur[_norm(row['EmployeeName'])] = {
            'hourly_rate':   round(rate if is_h else annual / 2080, 4),
            'annual_salary': round(annual, 2),
            'is_hourly':     is_h,
        }
    fl = {_fl(k): v for k, v in cur.items()}
    return cur, fl


def load_billing_rates(path):
    if not path.exists():
        return {}, {}, {}
    emp, pos = {}, {}
    for row in csv.DictReader(open(path, encoding='utf-8')):
        try: rate = float(row.get('BillingRate') or 0)
        except ValueError: continue
        if not rate: continue
        n = (row.get('Employee') or '').strip()
        p = (row.get('EmployeeType') or '').strip()
        if n: emp[_norm(n)] = rate
        elif p: pos[p.lower()] = rate
    fl = {_fl(k): v for k, v in emp.items()}
    return emp, fl, pos


def _pay(name, pay, fl):
    n = _norm(name)
    return pay.get(n) or fl.get(_fl(n)) or {}

def _bill_rate(name, etype, emp, fl, pos):
    n = _norm(name)
    return emp.get(n) or fl.get(_fl(n)) or pos.get((etype or '').lower(), 0.0)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _xf(c):  return PatternFill('solid', fgColor=c)
def _xb():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _xcell(ws, row, col, value, *, fill=None, bold=False, italic=False,
           size=10, color='000000', fmt=None, halign='center', wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
    if fill: c.fill = _xf(fill)
    c.border    = _xb()
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    if fmt: c.number_format = fmt
    return c

def _hdr(ws, row, col, text):
    return _xcell(ws, row, col, text, fill=_BLUE, bold=True, color=_WHITE, wrap=True)

def _title(ws, row, ncols, text, height=30):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14, color=_WHITE, name='Calibri')
    c.fill = _xf(_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = height

def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col) if isinstance(col, int) else col].width = w


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def write_excel(employees, period_str, period_label, lq_label, out_path,
                pay_lookup, pay_fl, emp_rates, emp_fl, pos_rates):

    ytd_frac = 1.0   # trailing 12 months = full year window

    # Build per-employee row data
    rows = []
    for e in employees:
        if e['ytd_hours']['total'] <= 0:
            continue
        ytd    = e['ytd_hours']
        lq     = e['lq_hours']
        cur    = e['current_hours']
        target = e['target_pct'] or DEFAULT_TARGET
        ytd_pct = ytd['billable'] / ytd['total'] * 100 if ytd['total'] else 0.0
        lq_pct  = lq['billable']  / lq['total']  * 100 if lq['total']  else 0.0
        cur_pct = cur['billable'] / cur['total'] * 100 if cur['total'] else 0.0
        vs_tgt  = ytd_pct - target

        pay    = _pay(e['name'], pay_lookup, pay_fl)
        annual = pay.get('annual_salary') or 0.0
        hrly   = pay.get('hourly_rate')   or 0.0
        ytd_sal = annual  # full year
        ytd_bil = ytd['billable'] * hrly

        br      = _bill_rate(e['name'], e['type'], emp_rates, emp_fl, pos_rates)
        ytd_rev = ytd['billable'] * br if br else 0.0
        margin  = ytd_rev - ytd_bil if br else None
        margin_pct = margin / ytd_rev * 100 if ytd_rev else None
        recovery   = ytd_bil / ytd_sal * 100 if ytd_sal else None
        mult       = ytd_bil / ytd_sal        if ytd_sal else None

        def dlr(h): return h * hrly if hrly else 0.0

        rows.append({
            'name': e['name'], 'type': e['type'], 'status': e['status'],
            'target': target, 'ytd_h': ytd['total'], 'bill_h': ytd['billable'],
            'bill_d': ytd_bil, 'ytd_pct': ytd_pct, 'lq_pct': lq_pct,
            'lq_h': lq['billable'], 'lq_total_h': lq['total'], 'cur_pct': cur_pct,
            'vs_tgt': vs_tgt, 'annual': annual, 'hrly': hrly,
            'ytd_sal': ytd_sal, 'recovery': recovery, 'mult': mult,
            'bill_rate': br, 'ytd_rev': ytd_rev, 'margin_d': margin, 'margin_pct': margin_pct,
            **{f'{k}_h': ytd[k]      for k in CAT_KEYS},
            **{f'{k}_d': dlr(ytd[k]) for k in CAT_KEYS},
        })

    n          = len(rows)
    tot_h      = sum(r['ytd_h']      for r in rows)
    tot_bill_h = sum(r['bill_h']     for r in rows)
    tot_bill_d = sum(r['bill_d']     for r in rows)
    tot_sal    = sum(r['ytd_sal']    for r in rows if r['ytd_sal'])
    tot_rev    = sum(r['ytd_rev']    for r in rows if r['ytd_rev'])
    tot_margin = tot_rev - tot_bill_d if tot_rev else 0.0
    firm_pct   = tot_bill_h / tot_h  * 100 if tot_h   else 0.0
    firm_rec   = tot_bill_d / tot_sal * 100 if tot_sal else 0.0
    firm_mult  = tot_bill_d / tot_sal       if tot_sal else 0.0
    firm_marg_pct = tot_margin / tot_rev * 100 if tot_rev else 0.0
    avg_tgt    = sum(r['target'] for r in rows) / n if n else DEFAULT_TARGET
    below_10   = sorted([r for r in rows if r['vs_tgt'] < -10], key=lambda r: r['vs_tgt'])
    above_tgt  = [r for r in rows if r['vs_tgt'] >= 0]
    cat_h      = {k: sum(r[f'{k}_h'] for r in rows) for k in CAT_KEYS}
    cat_d      = {k: sum(r[f'{k}_d'] for r in rows) for k in CAT_KEYS}
    tot_ind_h  = sum(cat_h.values())
    tot_ind_d  = sum(cat_d.values())
    has_billing = tot_rev > 0

    # Last-quarter aggregates
    lq_bill_h  = sum(r['lq_h']       for r in rows)
    lq_total_h = sum(r['lq_total_h'] for r in rows)
    lq_bill_d  = sum(r['lq_h'] * (r['hrly'] or 0) for r in rows)
    lq_rev     = sum(r['lq_h'] * r['bill_rate'] for r in rows if r['bill_rate'])
    lq_margin  = lq_rev - lq_bill_d if lq_rev else 0.0
    lq_pct_f   = lq_bill_h / lq_total_h * 100 if lq_total_h else 0.0
    lq_sal_est = tot_sal / 4  # one quarter of annual payroll
    lq_rec     = lq_bill_d / lq_sal_est * 100 if lq_sal_est else 0.0
    lq_marg_pct = lq_margin / lq_rev * 100 if lq_rev else 0.0

    wb = Workbook()

    # ── Tab 1: Executive Summary ─────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Executive Summary'
    ws.sheet_view.showGridLines = False
    _title(ws, 1, 11, f'Reztark Design Studio  —  Employee Utilization  |  {period_str}', height=34)

    def _kpi_row(ws, label_row, val_row, note_row, kpis):
        ws.row_dimensions[label_row].height = 20
        ws.row_dimensions[val_row].height   = 44
        ws.row_dimensions[note_row].height  = 16
        for c1, c2, lbl, val, note in kpis:
            ws.merge_cells(start_row=label_row, start_column=c1, end_row=label_row, end_column=c2)
            _xcell(ws, label_row, c1, lbl, fill=_BLUE, bold=True, color=_WHITE)
            ws.merge_cells(start_row=val_row, start_column=c1, end_row=val_row, end_column=c2)
            c = _xcell(ws, val_row, c1, val, fill=_LBLUE, bold=True, color=_DARK, size=20)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=note_row, start_column=c1, end_row=note_row, end_column=c2)
            _xcell(ws, note_row, c1, note, fill=_LBLUE, italic=True, color='595959', size=9)

    # Section label row helper
    def _section_lbl(ws, row, text):
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(bold=True, size=10, color=_WHITE, name='Calibri')
        c.fill = _xf(_DARK)
        c.alignment = Alignment(horizontal='left', vertical='center')

    # ── Trailing 12-month KPIs (rows 2–6) ────────────────────────────────
    ws.row_dimensions[2].height = 6
    _section_lbl(ws, 3, f'  Trailing 12 Months  —  {period_str}')
    if has_billing:
        t12_kpis = [
            (2,3,  'Active Staff',     str(n),                      period_label),
            (4,5,  'Firm Billable %',  f'{firm_pct:.1f}%',          f'Target {avg_tgt:.0f}%'),
            (6,7,  'Billing Revenue',  f'${tot_rev:,.0f}',          f'{tot_bill_h:,.0f} billable hrs'),
            (8,9,  'Gross Margin',     f'${tot_margin:,.0f}',       f'{firm_marg_pct:.1f}% margin'),
            (10,11,'Cost Recovery',    f'{firm_rec:.1f}%',          f'${tot_bill_d:,.0f} / ${tot_sal:,.0f}'),
        ]
    else:
        t12_kpis = [
            (2,3,  'Active Staff',     str(n),                      period_label),
            (4,5,  'Firm Billable %',  f'{firm_pct:.1f}%',          f'Target {avg_tgt:.0f}%'),
            (6,7,  'Billable Hours',   f'{tot_bill_h:,.0f} h',      f'of {tot_h:,.0f} total'),
            (8,9,  'Billable $ (est)', f'${tot_bill_d:,.0f}',       'from pay rates'),
            (10,11,'Cost Recovery',    f'{firm_rec:.1f}%',          f'est. vs ${tot_sal:,.0f} payroll'),
        ]
    _kpi_row(ws, 4, 5, 6, t12_kpis)

    # ── Last-quarter KPIs (rows 7–11) ────────────────────────────────────
    ws.row_dimensions[7].height = 6
    _section_lbl(ws, 8, f'  {lq_label}  —  Last Full Quarter')
    if has_billing:
        lq_kpis = [
            (2,3,  'Active Staff',     str(n),                      lq_label),
            (4,5,  'Firm Billable %',  f'{lq_pct_f:.1f}%',         f'T12: {firm_pct:.1f}%'),
            (6,7,  'Billing Revenue',  f'${lq_rev:,.0f}',           f'{lq_bill_h:,.0f} billable hrs'),
            (8,9,  'Gross Margin',     f'${lq_margin:,.0f}',        f'{lq_marg_pct:.1f}% margin'),
            (10,11,'Cost Recovery',    f'{lq_rec:.1f}%',            f'vs ~${lq_sal_est:,.0f} qtr payroll'),
        ]
    else:
        lq_kpis = [
            (2,3,  'Active Staff',     str(n),                      lq_label),
            (4,5,  'Firm Billable %',  f'{lq_pct_f:.1f}%',         f'T12: {firm_pct:.1f}%'),
            (6,7,  'Billable Hours',   f'{lq_bill_h:,.0f} h',       f'of {lq_total_h:,.0f} total'),
            (8,9,  'Billable $ (est)', f'${lq_bill_d:,.0f}',        'from pay rates'),
            (10,11,'Cost Recovery',    f'{lq_rec:.1f}%',            f'vs ~${lq_sal_est:,.0f} qtr payroll'),
        ]
    _kpi_row(ws, 9, 10, 11, lq_kpis)

    _widths(ws, {i: 16 for i in range(1, 12)})
    ws.column_dimensions['A'].width = 2

    # ── Narrative (rows 12–14) ────────────────────────────────────────────
    ws.row_dimensions[12].height = 10
    ws.row_dimensions[13].height = 18
    ws.merge_cells('B13:K13')
    c = ws.cell(row=13, column=2, value='Executive Summary')
    c.font = Font(bold=True, size=13, color=_DARK, name='Calibri')

    n_below   = len(below_10)
    top3      = sorted(above_tgt, key=lambda r: -r['vs_tgt'])[:3]
    top_names = ', '.join(r['name'].split()[0] for r in top3) if top3 else 'N/A'
    watch_s   = (
        'All employees are meeting or exceeding their billable target.' if n_below == 0
        else (f'One employee warrants attention: {below_10[0]["name"]} '
              f'({below_10[0]["ytd_pct"]:.1f}% vs {below_10[0]["target"]:.0f}% target).' if n_below == 1
        else f'{n_below} employees are more than 10 points below target: '
             f'{", ".join(r["name"] for r in below_10[:4])}'
             + (f' and {n_below-4} others.' if n_below > 4 else '.'))
    )
    narrative = (
        f'Trailing 12 months through {period_str.split("–")[-1].strip()}: '
        f'Reztark logged {tot_h:,.0f} total hours — {tot_bill_h:,.0f} ({firm_pct:.1f}%) billable '
        f'against a firm standard target of {avg_tgt:.0f}%. '
        + (f'Billing revenue is ${tot_rev:,.0f} with a gross margin of ${tot_margin:,.0f} ({firm_marg_pct:.1f}%). '
           if has_billing else f'Billable cost (estimated from pay rates) is ${tot_bill_d:,.0f}. ')
        + f'Cost recovery against estimated annual payroll of ${tot_sal:,.0f} is {firm_rec:.1f}%. '
        + f'In {lq_label}, the firm achieved {lq_pct_f:.1f}% billable ({lq_bill_h:,.0f} of {lq_total_h:,.0f} hours). '
        + f'{watch_s} Top performers above target: {top_names}.'
    )
    ws.row_dimensions[14].height = 80
    ws.merge_cells('B14:K14')
    c = ws.cell(row=14, column=2, value=narrative)
    c.font = Font(size=11, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ── Watch list (row 15+) ──────────────────────────────────────────────
    if below_10:
        ws.row_dimensions[15].height = 10
        ws.row_dimensions[16].height = 16
        ws.merge_cells('B16:K16')
        c = ws.cell(row=16, column=2,
                    value=f'Watch List  —  {n_below} employee(s) more than 10 points below target')
        c.font = Font(bold=True, size=12, color=_RFONT, name='Calibri')
        ws.row_dimensions[17].height = 20
        for ci, h in enumerate(['Employee','Type','T12 Bill%',f'{lq_label} Bill%','Target %','Gap','T12 Hours'], start=2):
            _hdr(ws, 17, ci, h)
        for ri, r in enumerate(below_10, start=18):
            ws.row_dimensions[ri].height = 16
            for ci, (v, fmt, ha) in enumerate([
                (r['name'],    None,              'left'),
                (r['type'],    None,              'left'),
                (r['ytd_pct'],'0.0"%"',          'center'),
                (r['lq_pct'], '0.0"%"',          'center'),
                (r['target'], '0"%"',             'center'),
                (r['vs_tgt'], '+0.0"%";-0.0"%"', 'center'),
                (r['ytd_h'],  '#,##0.0',         'right'),
            ], start=2):
                _xcell(ws, ri, ci, v, fill=_RED, color=_RFONT, fmt=fmt, halign=ha)

    # ── Tab 2: Utilization by Employee ───────────────────────────────────────
    ws2 = wb.create_sheet('Utilization by Employee')
    ws2.sheet_view.showGridLines = False
    _title(ws2, 1, 9, f'Employee Utilization  —  Trailing 12 Months  {period_str}')

    hdrs2 = ['Employee','Type','Status',
             f'{period_label}\nBill %', f'{lq_label}\nBill %', f'{lq_label}\nBill h',
             'Trailing 12m\nBill %','Target\n%','Gap vs\nTarget',
             'Bill h\n(12m)','Total h\n(12m)']
    ws2.row_dimensions[2].height = 28
    for ci, h in enumerate(hdrs2, 1): _hdr(ws2, 2, ci, h)

    for ri, r in enumerate(sorted(rows, key=lambda r: r['ytd_pct']), start=3):
        ws2.row_dimensions[ri].height = 16
        fill = _GREEN if r['vs_tgt'] >= 0 else (_AMBER if r['vs_tgt'] >= -10 else _RED)
        for ci, (v, fmt, ha) in enumerate([
            (r['name'],        None,               'left'),
            (r['type'],        None,               'left'),
            (r['status'],      None,               'center'),
            (r['cur_pct'],     '0.0"%"',           'center'),
            (r['lq_pct'],      '0.0"%"',           'center'),
            (r['lq_h'],        '#,##0.0',          'right'),
            (r['ytd_pct'],     '0.0"%"',           'center'),
            (r['target'],      '0"%"',             'center'),
            (r['vs_tgt'],      '+0.0"%";-0.0"%"',  'center'),
            (r['bill_h'],      '#,##0.0',          'right'),
            (r['ytd_h'],       '#,##0.0',          'right'),
        ], start=1):
            _xcell(ws2, ri, ci, v, fill=fill, fmt=fmt, halign=ha)

    _widths(ws2, {1:26, 2:22, 3:10, 4:12, 5:12, 6:12, 7:15, 8:10, 9:12, 10:13, 11:13})
    leg = len(rows) + 4
    ws2.merge_cells(f'A{leg}:K{leg}')
    c = ws2.cell(row=leg, column=1,
                 value='Green = at/above target   |   Amber = within 10 pts below   |   Red = >10 pts below')
    c.font = Font(italic=True, color='595959', size=9, name='Calibri')

    cd = leg + 2
    ws2.cell(row=cd, column=1, value='Employee')
    ws2.cell(row=cd, column=2, value='Trailing 12m Bill %')
    ws2.cell(row=cd, column=3, value='Target %')
    util_sorted = sorted(rows, key=lambda r: r['ytd_pct'])
    for i, r in enumerate(util_sorted, start=cd+1):
        parts = r['name'].split()
        ws2.cell(row=i, column=1, value=f"{parts[-1]}, {parts[0]}")
        ws2.cell(row=i, column=2, value=round(r['ytd_pct'], 1))
        ws2.cell(row=i, column=3, value=round(r['target'], 0))

    bar = BarChart()
    bar.type = 'bar'; bar.grouping = 'clustered'
    bar.title = 'Trailing 12-Month Billable % vs Target'
    bar.x_axis.title = 'Billable %'; bar.style = 10
    bar.width = 20; bar.height = max(12, len(rows) * 0.52)
    d = Reference(ws2, min_col=2, max_col=3, min_row=cd, max_row=cd+len(rows))
    bar.add_data(d, titles_from_data=True)
    bar.set_categories(Reference(ws2, min_col=1, min_row=cd+1, max_row=cd+len(rows)))
    bar.series[0].graphicalProperties.solidFill = _BLUE
    bar.series[0].graphicalProperties.line.solidFill = _BLUE
    if len(bar.series) > 1:
        bar.series[1].graphicalProperties.solidFill = 'BFBFBF'
        bar.series[1].graphicalProperties.line.solidFill = 'BFBFBF'
    ws2.add_chart(bar, 'M3')

    # ── Tab 3: Salary & Cost Recovery ────────────────────────────────────────
    ws3 = wb.create_sheet('Salary & Cost Recovery')
    ws3.sheet_view.showGridLines = False
    _title(ws3, 1, 8, f'Salary & Billable Cost Recovery  —  Trailing 12 Months  {period_str}')

    ws3.row_dimensions[2].height = 10
    ws3.merge_cells('A3:H3')
    c = ws3.cell(row=3, column=1,
                 value='Billable $ estimated from pay rates × billable hours. '
                       'Annual salary used as 12-month cost baseline.')
    c.font = Font(italic=True, color='595959', size=9, name='Calibri')
    c.alignment = Alignment(horizontal='left')
    ws3.row_dimensions[3].height = 14

    hdrs3 = ['Employee','Type','Hourly Rate','Annual Salary',
             'Est. Annual Cost','Billable $ (est)','Recovery %','Multiplier']
    ws3.row_dimensions[4].height = 28
    for ci, h in enumerate(hdrs3, 1): _hdr(ws3, 4, ci, h)

    sal_rows = sorted([r for r in rows if r['ytd_sal'] > 0], key=lambda r: r['recovery'] or 0)
    no_rate  = [r for r in rows if r['ytd_sal'] == 0]

    for ri, r in enumerate(sal_rows, start=5):
        ws3.row_dimensions[ri].height = 16
        rec  = r['recovery'] or 0
        fill = _GREEN if rec >= 80 else (_AMBER if rec >= 50 else _RED)
        for ci, (v, fmt, ha) in enumerate([
            (r['name'],    None,        'left'),
            (r['type'],    None,        'left'),
            (r['hrly'],    '$#,##0.00', 'right'),
            (r['annual'],  '$#,##0',    'right'),
            (r['ytd_sal'], '$#,##0',    'right'),
            (r['bill_d'],  '$#,##0',    'right'),
            (r['recovery'],'0.0"%"',    'center'),
            (r['mult'],    '0.00"x"',   'center'),
        ], start=1):
            _xcell(ws3, ri, ci, v or 0, fill=fill, fmt=fmt, halign=ha)

    tr = len(sal_rows) + 5
    ws3.row_dimensions[tr].height = 18
    for ci, (v, fmt, ha) in enumerate([
        ('FIRM TOTAL', None, 'left'), ('', None, 'left'), ('', None, 'right'),
        (sum(r['annual'] for r in sal_rows), '$#,##0', 'right'),
        (tot_sal,  '$#,##0',  'right'),
        (tot_bill_d, '$#,##0', 'right'),
        (firm_rec, '0.0"%"',  'center'),
        (firm_mult,'0.00"x"', 'center'),
    ], start=1):
        _xcell(ws3, tr, ci, v, fill=_LBLUE, bold=True, fmt=fmt, halign=ha)

    if no_rate:
        nr = tr + 2
        ws3.merge_cells(f'A{nr}:H{nr}')
        names = ', '.join(r['name'] for r in no_rate[:6])
        if len(no_rate) > 6: names += f' (+{len(no_rate)-6} more)'
        c = ws3.cell(row=nr, column=1, value=f'Pay rate not on file: {names}')
        c.font = Font(italic=True, color='595959', size=9, name='Calibri')

    _widths(ws3, {1:26, 2:22, 3:13, 4:15, 5:15, 6:15, 7:13, 8:12})

    rc_row = tr + (4 if no_rate else 3)
    ws3.cell(row=rc_row, column=1, value='Employee')
    ws3.cell(row=rc_row, column=2, value='Recovery %')
    ws3.cell(row=rc_row, column=3, value='80% Threshold')
    for i, r in enumerate(sal_rows, start=rc_row+1):
        parts = r['name'].split()
        ws3.cell(row=i, column=1, value=f"{parts[-1]}, {parts[0]}")
        ws3.cell(row=i, column=2, value=round(r['recovery'], 1) if r['recovery'] else 0)
        ws3.cell(row=i, column=3, value=80)
    bar2 = BarChart()
    bar2.type = 'bar'; bar2.grouping = 'clustered'
    bar2.title = 'Cost Recovery % (Billable $ ÷ Annual Salary)'
    bar2.x_axis.title = 'Recovery %'; bar2.style = 10
    bar2.width = 20; bar2.height = max(12, len(sal_rows) * 0.52)
    d2 = Reference(ws3, min_col=2, max_col=3, min_row=rc_row, max_row=rc_row+len(sal_rows))
    bar2.add_data(d2, titles_from_data=True)
    bar2.set_categories(Reference(ws3, min_col=1, min_row=rc_row+1, max_row=rc_row+len(sal_rows)))
    bar2.series[0].graphicalProperties.solidFill = _BLUE
    bar2.series[0].graphicalProperties.line.solidFill = _BLUE
    if len(bar2.series) > 1:
        bar2.series[1].graphicalProperties.solidFill = 'BFBFBF'
        bar2.series[1].graphicalProperties.line.solidFill = 'BFBFBF'
    ws3.add_chart(bar2, 'J4')

    # ── Tab 4: Profitability (billing rates only) ────────────────────────────
    if has_billing:
        ws_p = wb.create_sheet('Profitability')
        ws_p.sheet_view.showGridLines = False
        _title(ws_p, 1, 10, f'Profitability by Employee  —  Trailing 12 Months  {period_str}')
        ws_p.row_dimensions[2].height = 14
        ws_p.merge_cells('A3:J3')
        c = ws_p.cell(row=3, column=1,
                      value='Revenue = billable hours × billing rate.  Cost = billable hours × pay rate.  Gross Margin = Revenue − Cost.')
        c.font = Font(italic=True, color='595959', size=9, name='Calibri')
        c.alignment = Alignment(horizontal='left')
        ws_p.row_dimensions[3].height = 14

        hdrs_p = ['Employee','Type','Cost Rate\n($/hr)','Billing Rate\n($/hr)',
                  'Markup','Bill h','Revenue','Cost $','Gross Margin','Margin %']
        ws_p.row_dimensions[4].height = 28
        for ci, h in enumerate(hdrs_p, 1): _hdr(ws_p, 4, ci, h)

        prof_rows = sorted([r for r in rows if r['bill_rate'] > 0], key=lambda r: r['margin_pct'] or 0)
        no_bill   = [r for r in rows if not r['bill_rate']]

        for ri, r in enumerate(prof_rows, start=5):
            ws_p.row_dimensions[ri].height = 16
            mp   = r['margin_pct'] or 0
            fill = _GREEN if mp >= 40 else (_AMBER if mp >= 20 else _RED)
            markup = r['bill_rate'] / r['hrly'] if r['hrly'] else None
            for ci, (v, fmt, ha) in enumerate([
                (r['name'],      None,        'left'),
                (r['type'],      None,        'left'),
                (r['hrly'],      '$#,##0.00', 'right'),
                (r['bill_rate'], '$#,##0.00', 'right'),
                (markup,         '0.00"x"',   'center'),
                (r['bill_h'],    '#,##0.0',   'right'),
                (r['ytd_rev'],   '$#,##0',    'right'),
                (r['bill_d'],    '$#,##0',    'right'),
                (r['margin_d'],  '$#,##0',    'right'),
                (r['margin_pct'],'0.0"%"',    'center'),
            ], start=1):
                _xcell(ws_p, ri, ci, v or 0, fill=fill, fmt=fmt, halign=ha)

        ptr = len(prof_rows) + 5
        for ci, (v, fmt, ha) in enumerate([
            ('FIRM TOTAL',None,'left'),('',None,'left'),('',None,'right'),('',None,'right'),
            (tot_rev/tot_bill_d if tot_bill_d else 0,'0.00"x"','center'),
            (tot_bill_h,'#,##0.0','right'),(tot_rev,'$#,##0','right'),
            (tot_bill_d,'$#,##0','right'),(tot_margin,'$#,##0','right'),
            (firm_marg_pct,'0.0"%"','center'),
        ], start=1):
            _xcell(ws_p, ptr, ci, v, fill=_LBLUE, bold=True, fmt=fmt, halign=ha)

        if no_bill:
            nb = ptr + 2
            ws_p.merge_cells(f'A{nb}:J{nb}')
            names = ', '.join(r['name'] for r in no_bill[:6])
            c = ws_p.cell(row=nb, column=1, value=f'Billing rate not on file: {names}')
            c.font = Font(italic=True, color='595959', size=9, name='Calibri')

        _widths(ws_p, {1:26, 2:22, 3:13, 4:14, 5:10, 6:11, 7:14, 8:14, 9:15, 10:11})

    # ── Tab 5: Indirect Cost Breakdown ───────────────────────────────────────
    ws4 = wb.create_sheet('Indirect Cost Breakdown')
    ws4.sheet_view.showGridLines = False
    _title(ws4, 1, 7, f'Indirect Time & Cost Breakdown  —  Trailing 12 Months  {period_str}')

    ws4.row_dimensions[2].height = 10
    ws4.merge_cells('A3:G3')
    c = ws4.cell(row=3, column=1,
                 value=(f'Total indirect: {tot_ind_h:,.1f} h  |  '
                        f'Est. cost: ${tot_ind_d:,.0f}  |  '
                        f'Dollar values estimated using each employee\'s effective hourly rate.'))
    c.font = Font(italic=True, color='595959', size=10, name='Calibri')
    c.alignment = Alignment(horizontal='left')
    ws4.row_dimensions[3].height = 16

    hdrs4 = ['Category','Hours','% of Indirect h','% of Total h','Est. Cost $','% of Indirect $','Notes']
    ws4.row_dimensions[4].height = 20
    for ci, h in enumerate(hdrs4, 1): _hdr(ws4, 4, ci, h)

    cat_order  = sorted(CAT_KEYS, key=lambda k: -cat_h[k])
    pie_colors = ['2E75B6','70AD47','ED7D31','FFC000','FF0000','4472C4','A9D18E','F4B183']

    for ri, k in enumerate(cat_order, start=5):
        ws4.row_dimensions[ri].height = 18
        h    = cat_h[k]; d = cat_d[k]
        fill = _LBLUE if ri % 2 == 0 else _WHITE
        for ci, (v, fmt, ha) in enumerate([
            (CAT_LABELS[k], None, 'left'),
            (h,  '#,##0.0', 'right'),
            (h / tot_ind_h * 100 if tot_ind_h else 0, '0.0"%"', 'center'),
            (h / tot_h      * 100 if tot_h      else 0, '0.0"%"', 'center'),
            (d,  '$#,##0',  'right'),
            (d / tot_ind_d * 100 if tot_ind_d else 0, '0.0"%"', 'center'),
            ('', None, 'left'),
        ], start=1):
            _xcell(ws4, ri, ci, v, fill=fill, fmt=fmt, halign=ha)

    tot4 = len(cat_order) + 5
    ws4.row_dimensions[tot4].height = 18
    for ci, (v, fmt, ha) in enumerate([
        ('TOTAL', None, 'left'), (tot_ind_h, '#,##0.0', 'right'),
        (100.0, '0"%"', 'center'),
        (tot_ind_h / tot_h * 100 if tot_h else 0, '0.0"%"', 'center'),
        (tot_ind_d, '$#,##0', 'right'), (100.0, '0"%"', 'center'), ('', None, 'left'),
    ], start=1):
        _xcell(ws4, tot4, ci, v, fill=_LBLUE, bold=True, fmt=fmt, halign=ha)

    _widths(ws4, {1:20, 2:13, 3:17, 4:15, 5:14, 6:18, 7:22})

    pd = tot4 + 3
    ws4.cell(row=pd, column=1, value='Category')
    ws4.cell(row=pd, column=2, value='Est. Cost $')
    for i, k in enumerate(cat_order, start=pd+1):
        ws4.cell(row=i, column=1, value=CAT_LABELS[k])
        ws4.cell(row=i, column=2, value=round(cat_d[k], 0))

    pie = PieChart()
    pie.title = 'Indirect Cost Distribution'; pie.style = 10
    pie.width = 16; pie.height = 12
    pie.add_data(Reference(ws4, min_col=2, min_row=pd, max_row=pd+len(cat_order)), titles_from_data=True)
    pie.set_categories(Reference(ws4, min_col=1, min_row=pd+1, max_row=pd+len(cat_order)))
    for i, color in enumerate(pie_colors[:len(cat_order)]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    ws4.add_chart(pie, 'I4')

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f'  Excel -> {out_path}')


# ---------------------------------------------------------------------------
# Console report
# ---------------------------------------------------------------------------

def print_report(employees, period_str, lq_label, pay_lookup, pay_fl):
    with_h = [e for e in employees if e['ytd_hours']['total'] > 0]
    sep    = '=' * 80

    tot_h    = sum(e['ytd_hours']['total']    for e in with_h)
    tot_b    = sum(e['ytd_hours']['billable'] for e in with_h)
    lq_tot_b = sum(e['lq_hours']['billable']  for e in with_h)
    lq_tot_h = sum(e['lq_hours']['total']     for e in with_h)
    firm_pct = tot_b    / tot_h    * 100 if tot_h    else 0
    lq_pct   = lq_tot_b / lq_tot_h * 100 if lq_tot_h else 0

    print(f'\n{sep}')
    print('REZTARK DESIGN STUDIO  —  EMPLOYEE UTILIZATION')
    print(f'Trailing 12m : {period_str}   |   Last full quarter : {lq_label}')
    print(f'{sep}')
    print(f'\n  Trailing 12m — Total: {tot_h:,.1f}h   Billable: {tot_b:,.1f}h   Firm: {firm_pct:.1f}%')
    print(f'  {lq_label:<12} — Total: {lq_tot_h:,.1f}h   Billable: {lq_tot_b:,.1f}h   Firm: {lq_pct:.1f}%\n')

    targeted = sorted(
        [e for e in with_h if e['target_pct'] > 0],
        key=lambda e: (e['ytd_hours']['billable'] / e['ytd_hours']['total'] * 100 - e['target_pct'])
                      if e['ytd_hours']['total'] else -999,
    )
    print(f'UTILIZATION vs TARGET  ({len(targeted)} employees)')
    print(f'  {"Employee":<28}  {"Type":<22}  {"12m%":>5}  {lq_label:>7}  {"Tgt":>5}  {"Gap":>6}  {"12m h":>6}')
    print(f'  {"-"*28}  {"-"*22}  {"-"*5}  {"-"*7}  {"-"*5}  {"-"*6}  {"-"*6}')
    for e in targeted:
        ytd  = e['ytd_hours']
        lq   = e['lq_hours']
        pct  = ytd['billable'] / ytd['total'] * 100 if ytd['total'] else 0
        lqp  = lq['billable']  / lq['total']  * 100 if lq['total']  else 0
        gap  = pct - e['target_pct']
        flag = ' !' if gap < -10 else (' ~' if gap < 0 else '  ')
        print(f'{flag} {e["name"]:<28}  {e["type"]:<22}  {pct:>4.1f}%  {lqp:>6.1f}%  '
              f'{e["target_pct"]:>4.0f}%  {gap:>+5.1f}%  {ytd["total"]:>5.1f}h')

    no_target = sorted([e for e in with_h if e['target_pct'] == 0],
                       key=lambda e: -e['ytd_hours']['total'])
    if no_target:
        print(f'\nNO TARGET SET  ({len(no_target)} employees)')
        print(f'  {"Employee":<28}  {"Type":<22}  {"12m%":>5}  {lq_label:>7}  {"12m h":>6}')
        for e in no_target:
            ytd = e['ytd_hours']
            lq  = e['lq_hours']
            pct = ytd['billable'] / ytd['total'] * 100 if ytd['total'] else 0
            lqp = lq['billable']  / lq['total']  * 100 if lq['total']  else 0
            print(f'   {e["name"]:<28}  {e["type"]:<22}  {pct:>4.1f}%  {lqp:>6.1f}%  {ytd["total"]:>5.1f}h')
    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    load_dotenv(HERE / 'ajera.env')
    api_url  = os.environ['AJERA_API_URL']
    username = os.environ['AJERA_USERNAME']
    password = os.environ['AJERA_PASSWORD']

    trail_start, cur_month_start, today = _trail_dates()
    lq_start, lq_end, lq_label         = _last_quarter_dates()
    period_label = cur_month_start.strftime('%b%Y')
    period_str   = (f'{trail_start.strftime("%b %Y")} – '
                    f'{today.strftime("%b %Y")}')

    print(f'Period: {period_str}   |   Last quarter: {lq_label} ({lq_start} – {lq_end})')

    token = _connect(api_url, username, password)
    try:
        details = fetch_details(api_url, token, trail_start)
    finally:
        _disconnect(api_url, token)

    employees = parse_details(details, cur_month_start, lq_start, lq_end)
    active    = sum(1 for e in employees if e['status'] == 'Active')
    print(f'  {len(employees)} employees ({active} active)')

    pay_lookup, pay_fl = load_pay_rates(PAY_RATES_CSV)
    emp_rates, emp_fl, pos_rates = load_billing_rates(BILLING_RATES_CSV)

    out_path = OUTPUT_DIR / f'cincinnati_utilization_api_{period_label}.xlsx'
    write_excel(employees, period_str, period_label, lq_label, out_path,
                pay_lookup, pay_fl, emp_rates, emp_fl, pos_rates)

    print_report(employees, period_str, lq_label, pay_lookup, pay_fl)


if __name__ == '__main__':
    main()
