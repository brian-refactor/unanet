"""
Export COA preview workbook for stakeholder review.

Tabs:
  1. Master COA       — proposed unified chart of accounts
  2. Minnesota        — current QBO COA
  3. Cincinnati       — current Ajera COA
  4. Dallas           — current QB Desktop COA
  5. Orlando          — current QB Desktop COA

Output: output/COA_Preview.xlsx

Usage:
    python etl/export_coa_preview.py
"""

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv(Path(__file__).parent / ".env")

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

OUTPUT_PATH = Path("output/COA_Preview.xlsx")

OFFICES = ["minnesota", "cincinnati", "dallas", "orlando"]
OFFICE_LABELS = {
    "minnesota": "Minnesota",
    "cincinnati": "Cincinnati",
    "dallas": "Dallas",
    "orlando": "Orlando",
}

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark navy
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SECTION_FILL  = PatternFill("solid", fgColor="D6E4F0")   # light blue
SECTION_FONT  = Font(bold=True, name="Calibri", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
ALT_FILL      = PatternFill("solid", fgColor="F5F5F5")   # zebra stripe
THIN_BORDER   = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)

MASTER_SECTION_ORDER = [
    "Assets",
    "Liabilities",
    "Equity",
    "Revenue",
    "Direct Labor",
    "Direct Expenses",
    "Overhead",
    "G&A",
    "Other Income/Expense",
]

MASTER_SECTION_COLORS = {
    "Assets":               "E8F4FD",
    "Liabilities":          "FEF9E7",
    "Equity":               "EAFAF1",
    "Revenue":              "F9EBEA",
    "Direct Labor":         "FDF2F8",
    "Direct Expenses":      "FDF2F8",
    "Overhead":             "F4ECF7",
    "G&A":                  "EBF5FB",
    "Other Income/Expense": "FDFEFE",
}

def section_sort_key(section: str | None) -> tuple:
    try:
        return (MASTER_SECTION_ORDER.index(section), )
    except (ValueError, TypeError):
        return (len(MASTER_SECTION_ORDER), )

# Accounting-order sort key for financial types
FIN_TYPE_ORDER = {
    "Asset":        0,
    "Liability":    1,
    "Equity":       2,
    "Revenue":      3,
    "Expense":      4,
    "Other":        5,
}

def fin_sort_key(val: str | None) -> int:
    if val is None:
        return 99
    for k, v in FIN_TYPE_ORDER.items():
        if k.lower() in val.lower():
            return v
    return 98


def fetch_paginated(sb: Client, table: str, **filters) -> list[dict]:
    page, offset, rows = 1000, 0, []
    while True:
        q = sb.table(table).select("*")
        for col, val in filters.items():
            q = q.eq(col, val)
        batch = q.range(offset, offset + page - 1).execute().data
        rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return rows


def apply_header(ws, row_idx: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze(ws, cell: str) -> None:
    ws.freeze_panes = cell


def write_master_tab(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Master COA")

    headers = [
        "Master Code", "Account Name", "Section", "Financial Type",
        "Subledger Type", "Metric Type", "Cost Type", "PM Type",
        "Active", "1099", "Subcontractor", "Description", "Notes",
    ]
    db_cols = [
        "master_code", "master_name", "section", "financial_type",
        "subledger_type", "metric_type", "cost_type", "pm_type",
        "is_active", "is_1099", "is_subcontractor", "description", "notes",
    ]

    apply_header(ws, 1, headers)
    freeze(ws, "A2")

    rows_sorted = sorted(rows, key=lambda r: (section_sort_key(r.get("section")), r.get("sort_order", 9999)))

    current_section = None
    data_row = 2
    for i, row in enumerate(rows_sorted):
        section = row.get("section", "")
        if section != current_section:
            current_section = section
            # Section header row
            cell = ws.cell(row=data_row, column=1, value=section.upper())
            cell.font = SECTION_FONT
            cell.fill = PatternFill("solid", fgColor=MASTER_SECTION_COLORS.get(section, "EEEEEE"))
            ws.merge_cells(
                start_row=data_row, start_column=1,
                end_row=data_row, end_column=len(headers)
            )
            data_row += 1

        fill_color = MASTER_SECTION_COLORS.get(section, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color) if i % 2 == 0 else ALT_FILL

        for c, col in enumerate(db_cols, 1):
            val = row.get(col)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=data_row, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER

        data_row += 1

    set_col_widths(ws, [14, 38, 22, 16, 16, 14, 14, 12, 8, 8, 14, 40, 40])
    ws.row_dimensions[1].height = 18

    print(f"  Master COA: {len(rows_sorted)} accounts")


def write_office_tab(wb: openpyxl.Workbook, office: str, rows: list[dict]) -> None:
    label = OFFICE_LABELS[office]
    ws = wb.create_sheet(label)

    # Column order matches Unanet 02-COA_Fusion.xlsx template exactly
    headers = [
        "Account Code", "Account Name", "Description",
        "Active", "1099", "Subcontractor",
        "Financial Type", "Subledger Type", "Metric Type",
        "Cost Type", "PM Type", "Labor Rev Type", "Expense Rev Type",
    ]
    db_cols = [
        "base_code", "base_name", "description",
        "is_active", "is_1099", "is_subcontractor",
        "financial_type", "subledger_type", "metric_type",
        "cost_type", "pm_type", "labor_revenue_type", "expense_revenue_type",
    ]

    apply_header(ws, 1, headers)
    freeze(ws, "A2")

    rows_sorted = sorted(rows, key=lambda r: (fin_sort_key(r.get("financial_type")), str(r.get("base_code", ""))))

    current_fin = None
    data_row = 2
    for i, row in enumerate(rows_sorted):
        fin = row.get("financial_type", "")
        if fin != current_fin:
            current_fin = fin
            cell = ws.cell(row=data_row, column=1, value=(fin or "Other").upper())
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            ws.merge_cells(
                start_row=data_row, start_column=1,
                end_row=data_row, end_column=len(headers)
            )
            data_row += 1

        row_fill = PatternFill("solid", fgColor="FFFFFF") if i % 2 == 0 else ALT_FILL

        for c, col in enumerate(db_cols, 1):
            val = row.get(col)
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=data_row, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER

        data_row += 1

    set_col_widths(ws, [16, 38, 40, 8, 8, 14, 16, 16, 14, 14, 12, 16, 16])
    ws.row_dimensions[1].height = 18

    print(f"  {label}: {len(rows_sorted)} accounts")


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    print("Fetching data from Supabase...")
    master_rows = fetch_paginated(sb, "coa_master")
    master_rows.sort(key=lambda r: r.get("sort_order", 9999))

    office_rows = {}
    for office in OFFICES:
        office_rows[office] = fetch_paginated(sb, "coa_resolved", office=office)

    print("Building workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_master_tab(wb, master_rows)
    for office in OFFICES:
        write_office_tab(wb, office, office_rows[office])

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
