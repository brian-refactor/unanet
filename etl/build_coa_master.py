"""
Build COA_Master_Mapping.xlsx — two tabs:
  1. Master COA: one row per new account, Unanet fields + source mapping per office
  2. Needs Mapping: accounts with OPEN or N/A source entries
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import OrderedDict

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_BLUE  = '1F4E79'
MED_BLUE   = '2E75B6'
WHITE      = 'FFFFFF'

SECTION_COLORS = {
    'Assets':      'EBF3FB',
    'Liabilities': 'FFF2CC',
    'Equity':      'E2EFDA',
    'Capital':     'E2EFDA',
    'Revenue':     'F4F4F4',
    'Income':      'F4F4F4',
    'Expense':     'FFFFFF',
}

def hdr(ws, row, col, value, bg=MED_BLUE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=WHITE, size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    return c

def cell(ws, row, col, value, bg=None, bold=False, color='000000'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=9, bold=bold, color=color)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(vertical='top')
    return c

# ── Load source data ───────────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(r'E:\Unanet\output\COA_Working File-6-2-26.xlsx')
ws_matrix = wb_src['COA MATRIX']
src_rows = list(ws_matrix.iter_rows(min_row=2, values_only=True))

# ── Load authoritative Unanet field values from the populated Fusion template ─
# MetricType, CostType, PMType, LaborRevenueType were set by populate_coa_metrics.py
# The COA MATRIX doesn't have these — use the Fusion file as the source of truth.
wb_fusion = openpyxl.load_workbook(r'E:\Unanet\output\02-COA_Fusion 6-2-26_updated.xlsx')
ws_fusion = wb_fusion['Chart of Accounts']
# BaseCode(1), BaseName(2), Desc(3), IsActive(4), Is1099(5), IsSubcon(6),
# FinancialType(7), SubledgerType(8), MetricType(9), CostType(10), PMType(11),
# LaborRevType(12), ExpenseRevType(13)
fusion_lookup = {}  # str(BaseCode) -> dict of Unanet fields
for row in ws_fusion.iter_rows(min_row=4, values_only=True):
    if not row[0]:
        continue
    fusion_lookup[str(row[0]).strip()] = {
        'fin_type':  str(row[6]  or '').strip(),
        'subledger': str(row[7]  or '').strip(),
        'metric':    str(row[8]  or '').strip(),
        'cost':      str(row[9]  or '').strip(),
        'pm':        str(row[10] or '').strip(),
        'labor_rev': str(row[11] or '').strip(),
        'active':    str(row[3]  or '').strip(),
        'is_1099':   str(row[4]  or '').strip(),
        'subcon':    str(row[5]  or '').strip(),
    }

OFFICES = ['DAL', 'MSP', 'CIN', 'ORL', 'CORP']

acct_map  = OrderedDict()   # key -> {office: (src_num, src_name, src_erp)}
acct_meta = OrderedDict()   # key -> field dict

# Also load per-office full detail rows for Active/Inactive
office_detail = {}  # (key, office) -> full row

for r in src_rows:
    new_code = r[2]
    if not new_code or str(new_code).strip() == '':
        continue
    key = str(new_code).strip()
    if key not in acct_map:
        # Prefer Fusion template values (authoritative); fall back to working file
        fus = fusion_lookup.get(str(r[2]).strip(), {})

        ft_raw  = fus.get('fin_type') or str(r[10] or '').strip()
        ft_norm = {'Equity': 'Capital', 'Revenue': 'Income'}.get(ft_raw, ft_raw)

        acct_map[key]  = {}
        acct_meta[key] = {
            'rollup_code': r[0],   'rollup_name': r[1],
            'new_code':    r[2],   'new_name':    r[3],
            'section':     r[9],   'fin_type':    ft_norm,
            'subledger':   fus.get('subledger') or str(r[11] or '').strip(),
            'metric':      fus.get('metric')    or str(r[12] or '').strip(),
            'cost':        fus.get('cost')      or str(r[13] or '').strip(),
            'pm':          fus.get('pm')        or str(r[14] or '').strip(),
            'active':      fus.get('active')    or str(r[15] or '').strip(),
            'is_1099':     fus.get('is_1099')   or str(r[16] or '').strip(),
            'subcon':      fus.get('subcon')    or str(r[17] or '').strip(),
            'description': str(r[18] or '').strip(),
            'notes':       r[19],
        }
    office = str(r[7] or '').strip()
    if office in OFFICES:
        acct_map[key][office] = (str(r[4] or '').strip(), str(r[5] or '').strip(), str(r[6] or '').strip())
        office_detail[(key, office)] = r
    elif office.upper().startswith('OPEN'):
        acct_map[key]['OPEN'] = True

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: Master COA
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Master COA'
ws1.freeze_panes = 'A3'

# Row 1: group headers
group_row = [
    (1,  3,  'ROLLUP',             DARK_BLUE),
    (4,  14, 'UNANET COA FIELDS',  MED_BLUE),
    (15, 16, 'DALLAS',             '4472C4'),
    (17, 18, 'MINNEAPOLIS',        '4472C4'),
    (19, 20, 'CINCINNATI',         '4472C4'),
    (21, 22, 'ORLANDO',            '4472C4'),
    (23, 24, 'FLAGS',              '595959'),
]
for start, end, label, color in group_row:
    ws1.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    c = ws1.cell(row=1, column=start, value=label)
    c.font  = Font(bold=True, color=WHITE, size=9)
    c.fill  = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 16

# Row 2: column headers
col_hdrs = [
    # Rollup (1-3)
    'Rollup Code', 'Rollup Account Name', 'Section',
    # Unanet fields (4-14)
    'BaseCode', 'BaseName', 'FinancialType', 'SubledgerType', 'MetricType',
    'CostType', 'PMType', 'IsActive', 'Is1099', 'IsSubcontractor', 'Description',
    # Sources (15-22)
    'DAL Src #', 'DAL Src Name',
    'MSP Src #', 'MSP Src Name',
    'CIN Src #', 'CIN Src Name',
    'ORL Src #', 'ORL Src Name',
    # Flags (23-24)
    'Has OPEN?', 'Notes',
]
for col, h in enumerate(col_hdrs, 1):
    hdr(ws1, 2, col, h, bg='404040')
ws1.row_dimensions[2].height = 28

# Data rows
for key, meta in acct_meta.items():
    sources  = acct_map[key]
    section  = str(meta['section'] or '').strip()
    bg       = SECTION_COLORS.get(section, 'FFFFFF')
    has_open = sources.get('OPEN', False)

    dal = sources.get('DAL', ('', ''))
    msp = sources.get('MSP', ('', ''))
    cin = sources.get('CIN', ('', ''))
    orl = sources.get('ORL', ('', ''))

    # Default boolean flags if blank (OPEN placeholder rows etc.)
    active = meta['active'] or 'Yes'
    is1099 = meta['is_1099'] or 'No'
    subcon = meta['subcon']  or 'No'

    vals = [
        meta['rollup_code'], meta['rollup_name'], section,
        meta['new_code'], meta['new_name'], meta['fin_type'], meta['subledger'],
        meta['metric'], meta['cost'], meta['pm'],
        active, is1099, subcon, meta['description'],
        dal[0], dal[1],
        msp[0], msp[1],
        cin[0], cin[1],
        orl[0], orl[1],
        'YES' if has_open else '',
        meta['notes'],
    ]
    rn = ws1.max_row + 1
    for col, val in enumerate(vals, 1):
        c = cell(ws1, rn, col, val, bg=bg)
    if has_open:
        c23 = ws1.cell(row=rn, column=23)
        c23.fill  = PatternFill('solid', fgColor='FCE4D6')
        c23.font  = Font(size=9, bold=True, color='C00000')

ws1.auto_filter.ref = f'A2:{get_column_letter(len(col_hdrs))}2'

# ── Highlight any remaining blank required cells (highlight only, no placeholder text) ─
RED_FILL = PatternFill('solid', fgColor='FCE4D6')
RED_FONT = Font(size=9, color='C00000', bold=True)

for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row):
    basecode  = str(row[3].value or '').strip()
    basename  = str(row[4].value or '').strip()
    fin_type  = str(row[5].value or '').strip()
    subledger = str(row[6].value or '').strip()
    metric    = str(row[7].value or '').strip()

    if not basecode:
        continue

    if not fin_type:
        row[5].fill = RED_FILL
        row[5].font = RED_FONT

    if not basename:
        row[4].fill = RED_FILL
        row[4].font = RED_FONT

    # Non-project income accounts intentionally have no MetricType
    NON_PROJECT = {43501, 44001}
    if fin_type in ('Income', 'Expense') and not metric and not subledger:
        try:
            if int(str(basecode)) not in NON_PROJECT:
                row[7].fill = RED_FILL
                row[7].font = Font(size=9, color='C00000')
        except (ValueError, TypeError):
            row[7].fill = RED_FILL
            row[7].font = Font(size=9, color='C00000')

col_widths = [12, 36, 12, 10, 36, 14, 18, 12, 10, 10, 8, 7, 13, 32,
              12, 30, 12, 30, 12, 30, 12, 30, 10, 28]
for col, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w

# ── Inject the 26 manually resolved mappings into Master COA source columns ───
# These are source accounts that weren't in the original COA MATRIX but have
# now been mapped to a master code. We append them to the correct master row.
MANUAL_MAPPINGS = [
    ('DAL', 'Design_Income-Accrued',                    'Design Income-Accrued',              42501),
    ('DAL', 'Fusion_AE_-_FVC_Bank',                     'Fusion AE - FVC Bank',               10302),
    ('DAL', 'Fusion_AE_-_Payroll',                      'Fusion AE - Payroll',                10201),
    ('DAL', 'ID_Studio4,_LLC_-_Sweep',                  'ID Studio4, LLC - Sweep',            10301),
    ('DAL', 'Investment_in_Reztark_ACi',                'Investment in Reztark ACi',          15001),
    ('DAL', 'NWC_Escrow_-_Reztark',                     'NWC Escrow - Reztark',               25001),
    ('DAL', 'Preferred_Equity_Payments',                'Preferred Equity Payments',          81701),
    ('MSP', '401',                                       'Billed fee revenue',                 40001),
    ('MSP', '767',                                       'Drafting Expenses',                  63101),
    ('MSP', '721',                                       'Equipment Rental',                   72601),
    ('MSP', '980',                                       'MN Min Fee Tax',                     81303),
    ('MSP', '186',                                       'Organization Expenses',              12401),
    ('MSP', '172',                                       'Preferred Equity Payments-MN',       81701),
    ('MSP', '550',                                       'Reimbursable Expenses',              61001),
    ('MSP', '981',                                       'Sales Tax',                          81301),
    ('MSP', '982',                                       'WI Fee Tax',                         81303),
    ('CIN', 'AJ-183',                                    'Federal Income Tax Withholding',     20303),
    ('CIN', 'AJ-205',                                    'Intercompany Due From',              12201),
    ('CIN', 'AJ-206',                                    'Intercompany Due To',                21203),
    ('CIN', 'AJ-207',                                    'Intercompany Other Income',          90101),
    ('CIN', 'AJ-90',                                     'Nonbillable Consultant Expenses',    60201),
    ('CIN', 'AJ-107',                                    'Professional Registration & Dues',   73101),
    ('ORL', 'Accrued_Revenue',                           'Accrued Revenue',                    20704),
    ('ORL', 'Insurance_Auto_Insurance',                  'Auto Insurance',                     74003),
    ('ORL', 'Travel_&_Entertainment_Entertainment',      'Entertainment',                      76001),
    ('ORL', 'Car/Truck_Expense_Registration_&_License',  'Registration & License',             73201),
]

OFFICE_SRC_COL = {'DAL': (15, 16), 'MSP': (17, 18), 'CIN': (19, 20), 'ORL': (21, 22)}

# Build a row-index lookup: master_code -> row number in ws1
master_row_idx = {}
for rn in range(3, ws1.max_row + 1):
    code = ws1.cell(row=rn, column=4).value
    if code is not None:
        master_row_idx[str(code).strip()] = rn

for office, src_code, src_name, master_code in MANUAL_MAPPINGS:
    key = str(master_code).strip()
    rn  = master_row_idx.get(key)
    if not rn:
        print(f'  WARNING: master code {master_code} not found in sheet')
        continue
    num_col, name_col = OFFICE_SRC_COL[office]
    existing_num  = str(ws1.cell(row=rn, column=num_col).value  or '').strip()
    existing_name = str(ws1.cell(row=rn, column=name_col).value or '').strip()
    # Append if already has a value (multiple sources → same master)
    ws1.cell(row=rn, column=num_col).value  = f'{existing_num} / {src_code}'.lstrip(' / ')   if existing_num  else src_code
    ws1.cell(row=rn, column=name_col).value = f'{existing_name} / {src_name}'.lstrip(' / ')  if existing_name else src_name

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: Needs Mapping
# Source accounts that have NOT been matched to any master COA account.
# One row per unmapped source account. The reviewer picks the master COA
# account it should roll up to.
# ══════════════════════════════════════════════════════════════════════════════

# Build source account lists and identify unmapped ones
OFFICE_TABS = {'DAL': 'Dallas', 'MSP': 'Minnesota', 'CIN': 'Cincinnati', 'ORL': 'Orlando'}
SKIP_CODES  = {'ASSET','LIABILITY','LIABILITIES','EQUITY','REVENUE','EXPENSE','INCOME',
               'OVERHEAD','COST','G&A','INTERCOMPANY','DIRECT COSTS','ASSETS','CAPITAL',''}

source_accts = {}
for off, tab in OFFICE_TABS.items():
    ws_off = wb_src[tab]
    accts  = []
    for r in ws_off.iter_rows(min_row=2, values_only=True):
        code = str(r[0] or '').strip()
        name = str(r[1] or '').strip()
        if code and code.upper() not in SKIP_CODES and name:
            accts.append({'code': code, 'name': name, 'active': r[3], 'fin_type': r[6]})
    source_accts[off] = accts

# Build matrix lookup: office -> {code -> master, name -> master}
mapped_by_code = {'DAL': {}, 'MSP': {}, 'CIN': {}, 'ORL': {}}
mapped_by_name = {'DAL': {}, 'MSP': {}, 'CIN': {}, 'ORL': {}}
for r in src_rows:
    office   = str(r[7] or '').strip()
    src_code = str(r[4] or '').strip()
    src_name = str(r[5] or '').strip()
    new_code = r[2]
    new_name = r[3]
    if office not in mapped_by_code or not new_code:
        continue
    if src_code and src_code.upper() not in ('', 'OPEN', 'N/A'):
        mapped_by_code[office][src_code] = (new_code, new_name)
    if src_name and src_name.upper() not in ('', 'OPEN', 'N/A'):
        mapped_by_name[office][src_name.lower()] = (new_code, new_name)

# Build set of manually resolved source codes so we can exclude them
MANUAL_RESOLVED = {(office, src_code) for office, src_code, _, _ in MANUAL_MAPPINGS}

# Collect unmapped source accounts
unmapped_rows = []
for off, accts in source_accts.items():
    for a in accts:
        if (off, a['code']) in MANUAL_RESOLVED:
            continue
        match = (mapped_by_code[off].get(a['code']) or
                 mapped_by_name[off].get(a['name'].lower()))
        if not match:
            unmapped_rows.append({
                'office':   off,
                'src_code': a['code'],
                'src_name': a['name'],
                'active':   a['active'],
                'fin_type': a['fin_type'],
            })

# Sort by office then name
OFFICE_ORDER = {'DAL': 0, 'MSP': 1, 'CIN': 2, 'ORL': 3}
unmapped_rows.sort(key=lambda x: (OFFICE_ORDER.get(x['office'], 9), x['src_name']))

ws2 = wb.create_sheet('Needs Mapping')
ws2.freeze_panes = 'A2'

nm_hdrs = [
    'Office', 'Source Account #', 'Source Account Name', 'Financial Type',
    'Active?', 'Map to Master COA Code', 'Map to Master COA Name', 'Notes',
]
for col, h in enumerate(nm_hdrs, 1):
    hdr(ws2, 1, col, h)
ws2.row_dimensions[1].height = 28

OFFICE_COLORS = {'DAL': 'EBF3FB', 'MSP': 'E2EFDA', 'CIN': 'FFF2CC', 'ORL': 'F4EAFB'}

for rn, u in enumerate(unmapped_rows, 2):
    bg = OFFICE_COLORS.get(u['office'], 'FFFFFF')
    vals = [
        u['office'], u['src_code'], u['src_name'], u['fin_type'],
        u['active'], '', '', '',
    ]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=rn, column=col, value=val)
        c.font = Font(size=9)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(vertical='top')
    # Highlight the two "Map to" columns so it's obvious they need filling in
    for col in (6, 7):
        ws2.cell(row=rn, column=col).fill = PatternFill('solid', fgColor='FCE4D6')

ws2.auto_filter.ref = f'A1:{get_column_letter(len(nm_hdrs))}1'

nm_widths = [8, 22, 42, 16, 8, 22, 42, 30]
for col, w in enumerate(nm_widths, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = r'E:\Unanet\output\COA_Master_Mapping_v3.xlsx'
wb.save(out_path)

print(f'Saved: {out_path}')
print(f'  Tab 1 (Master COA):    {len(acct_meta)} accounts')
print(f'  Tab 2 (Needs Mapping): {len(unmapped_rows)} source accounts not yet mapped to master COA')
by_office = {}
for u in unmapped_rows:
    by_office.setdefault(u["office"], 0)
    by_office[u["office"]] += 1
for off, cnt in sorted(by_office.items()):
    print(f'    {off}: {cnt}')
