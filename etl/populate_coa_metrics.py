"""
Populate MetricType, CostType, PMType, and LaborRevenueType on the Fusion COA template
for P&L accounts that currently have those fields blank.

Rules applied:
  40xxx  Design/Fee Revenue      → MT=Billed Revenue, PM=Labor,                  LaborRevType=Direct
  41xxx  Reimbursable Revenue    → MT=Billed Revenue, PM=Other Direct Charges
  42001  Subconsultant Revenue   → MT=Billed Revenue, PM=In-Contract Charges
  42501  WIP Adj                 → MT=Work In Progress, PM=Labor
  43001  Writeoffs               → MT=Bad Debt
  43501  Interest Income         → (non-project, leave blank)
  44001  Other Income            → (non-project, leave blank)
  60101  Subcon In-Contract      → MT=Cost, CT=Direct, PM=In-Contract Charges
  60201  Subcon Out-of-Contract  → MT=Cost, CT=Direct, PM=Out-of-Contract Charges
  60301  Other Consultants       → MT=Cost, CT=Direct, PM=In-Contract Charges
  62xxx  COGS                    → MT=Cost, CT=Direct, PM=Other Direct Charges
  63xxx  Project Direct Costs    → MT=Cost, CT=Direct, PM=Other Direct Charges
  70xxx  Overhead / Indirect     → MT=Cost, CT=Indirect
  71xxx  Fringe                  → MT=Cost, CT=Indirect
  72xxx  Facilities/General OH   → MT=Cost, CT=Indirect
  73xxx  Prof Dev OH             → MT=Cost, CT=Indirect
  74xxx  Insurance OH            → MT=Cost, CT=Indirect
  75xxx  Depreciation            → MT=Cost, CT=Indirect
  76xxx  M&E                     → MT=Cost, CT=Indirect
  77xxx-79xxx Misc OH            → MT=Cost, CT=Indirect
  80xxx  G&A                     → MT=Cost, CT=Indirect
  81xxx  G&A                     → MT=Cost, CT=Indirect
  90101  Intercompany Revenue    → MT=Billed Revenue, PM=Labor, LaborRevType=Direct
  90201  Intercompany Expense    → MT=Cost, CT=Direct
  90502  Billable Time           → MT=Cost, CT=Direct, PM=Labor
"""
import openpyxl
from pathlib import Path

SRC  = Path(r'E:\Unanet\output\02-COA_Fusion 6-2-26.xlsx')
DEST = Path(r'E:\Unanet\output\02-COA_Fusion 6-2-26_updated.xlsx')

wb = openpyxl.load_workbook(SRC)
ws = wb['Chart of Accounts']

# Column positions (1-based) — row 3 has the field names
# BaseCode=1, BaseName=2, Desc=3, IsActive=4, Is1099=5, IsSubcon=6,
# FinancialType=7, SubledgerType=8, MetricType=9, CostType=10, PMType=11,
# LaborRevenueType=12, ExpenseRevenueType=13
C_CODE  = 1
C_NAME  = 2
C_FT    = 7
C_SL    = 8
C_MT    = 9
C_CT    = 10
C_PM    = 11
C_LABRT = 12

def series(code):
    try:
        n = int(str(code).strip())
        return (n // 100) * 100
    except (ValueError, TypeError):
        return None

def apply(ws, row_idx, mt=None, ct=None, pm=None, labrt=None):
    if mt    is not None: ws.cell(row=row_idx, column=C_MT).value    = mt
    if ct    is not None: ws.cell(row=row_idx, column=C_CT).value    = ct
    if pm    is not None: ws.cell(row=row_idx, column=C_PM).value    = pm
    if labrt is not None: ws.cell(row=row_idx, column=C_LABRT).value = labrt

changed = 0
skipped_non_project = []

for rn in range(4, ws.max_row + 1):
    code    = ws.cell(row=rn, column=C_CODE).value
    name    = str(ws.cell(row=rn, column=C_NAME).value or '').strip()
    ft      = str(ws.cell(row=rn, column=C_FT).value or '').strip()
    sl      = str(ws.cell(row=rn, column=C_SL).value or '').strip()
    mt_cur  = str(ws.cell(row=rn, column=C_MT).value or '').strip()

    if not code or not name:
        continue
    if ft not in ('Income', 'Expense'):
        continue
    if mt_cur or sl:
        continue  # already populated

    s = series(code)

    # ── Revenue ──────────────────────────────────────────────────────────────
    if s == 40000:
        apply(ws, rn, mt='Billed Revenue', pm='Labor', labrt='Direct')
    elif s == 41000:
        apply(ws, rn, mt='Billed Revenue', pm='Other Direct Charges')
    elif code == 42001:
        apply(ws, rn, mt='Billed Revenue', pm='In-Contract Charges')
    elif code == 42501:
        apply(ws, rn, mt='Work In Progress', pm='Labor')
    elif code == 43001:
        apply(ws, rn, mt='Bad Debt')
    elif code in (43501, 44001):
        skipped_non_project.append(f'{code}  {name}')
        continue

    # ── Direct Costs ─────────────────────────────────────────────────────────
    elif code == 60101:
        apply(ws, rn, mt='Cost', ct='Direct', pm='In-Contract Charges')
    elif code == 60201:
        apply(ws, rn, mt='Cost', ct='Direct', pm='Out-of-Contract Charges')
    elif code == 60301:
        apply(ws, rn, mt='Cost', ct='Direct', pm='In-Contract Charges')
    elif 62000 <= (s or 0) <= 63900:
        apply(ws, rn, mt='Cost', ct='Direct', pm='Other Direct Charges')

    # ── Overhead (70xxx–79xxx) ────────────────────────────────────────────────
    elif 70000 <= (s or 0) <= 79900:
        apply(ws, rn, mt='Cost', ct='Indirect')

    # ── G&A (80xxx–81xxx) ────────────────────────────────────────────────────
    elif 80000 <= (s or 0) <= 81900:
        apply(ws, rn, mt='Cost', ct='Indirect')

    # ── Intercompany (90xxx) ─────────────────────────────────────────────────
    elif code == 90101:
        apply(ws, rn, mt='Billed Revenue', pm='Labor', labrt='Direct')
    elif code == 90201:
        apply(ws, rn, mt='Cost', ct='Direct')
    elif code == 90502:
        apply(ws, rn, mt='Cost', ct='Direct', pm='Labor')

    else:
        skipped_non_project.append(f'{code}  {name}  (no rule matched)')
        continue

    changed += 1

wb.save(DEST)
print(f'Updated {changed} accounts.')
print(f'Saved: {DEST}')
if skipped_non_project:
    print(f'\nLeft blank (non-project / no rule):')
    for s in skipped_non_project:
        print(f'  {s}')
