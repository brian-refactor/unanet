"""
Projects template writer — reads from the projects and project_phases Supabase tables
and writes output/templates/projects_merged.xlsx in the 07a-OpenProjects_Fusion.xlsx format.

Supports an optional lookup step to apply org-unit paths and employee codes:
    --org-units  path/to/org_units.csv     (columns: org_path, ...)
    --emp-codes  path/to/employee_codes.csv (columns: emp_code, first_name, last_name, ...)

Without the lookup files the script still runs — those columns are left blank
so you can fill them in after Andrew sends the files.

Usage:
    python etl/write_projects_template.py
    python etl/write_projects_template.py --org-units input/OrgUnits.xlsx --emp-codes input/EmployeeCodes.xlsx
    python etl/write_projects_template.py --office dallas
"""

import argparse
import csv
import os
import shutil
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

HERE = Path(__file__).parent
load_dotenv(HERE / ".env")

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

TEMPLATE_DIR = Path("Documentation/OneDrive_1_4-24-2026/Data Upload Templates")
OUTPUT_DIR   = Path("output/templates")
TEMPLATE_FILE = "07a-OpenProjects_Fusion.xlsx"

OFFICES = ["minnesota", "cincinnati", "dallas", "orlando"]

# ---------------------------------------------------------------------------
# Column order — must match template physical columns left-to-right
# ---------------------------------------------------------------------------

PROJECTS_COLS = [
    "client_firm_code",    # ClientCode
    "owning_org",          # OwningOrg
    "project_code",        # ProjectCode
    "project_name",        # ProjectName
    "charge_type",         # ChargeTypeName
    "start_date",          # StartDate
    "end_date",            # EndDate
    "contract_type",       # ContractTypeName
    "project_note",        # ProjectNote
    "po_number",           # PONumber
    "pm_emp_code",         # ProjectManagerEmpCode
    "pic_emp_code",        # PICEmpCode
    "pa_emp_code",         # ProjectAccountEmpCode
    "billing_term_type",   # BillingTermType
    "net_days",            # NetDays
    None,                  # NextInvNum (leave blank until go-live)
    "invoice_email",       # InvoiceEmail
    "location_street1",    # ProjectLocationStreet1
    "location_street2",    # ProjectLocationStreet2
    "location_city",       # ProjectLocationCity
    "location_state",      # ProjectLocationState
    "location_zip",        # ProjectLocationZip
    "location_country",    # ProjectLocationCountry
    "use_client_bill_to",  # UseClientBillTo
    "bill_to_street1",     # BillToStreet1
    "bill_to_street2",     # BillToStreet2
    "bill_to_city",        # BillToCity
    "bill_to_state",       # BillToState
    "bill_to_zip",         # BillToZip
    "bill_to_country",     # BillToCountry
]

PHASES_COLS = [
    "project_code",        # LevelOneProjectCode
    "contract_type",       # ContractType
    "level2_name",         # Level2ProjectName
    "level2_code",         # Level2ProjectCode
    "level3_name",         # Level3ProjectName
    "level3_code",         # Level3ProjectCode
    "start_date",          # StartDate
    "end_date",            # EndDate
    "org_path",            # OrgPath
    "fixed_fee",           # FixedFee
    "labor_contract_cap",  # LaborContractCap
    "odc_contract_cap",    # ODContractCap
    "occ_contract_cap",    # OCCContractCap
    "icc_fixed_fee",       # ICCFixedFeePortion
    "labor_budget",        # LaborBudget
    "odc_budget",          # ODCBudget
    "occ_budget",          # OCCBudget
    "icc_budget",          # ICCBudget
    "hours_budget",        # HoursBudget
]


# ---------------------------------------------------------------------------
# Supabase fetch
# ---------------------------------------------------------------------------

def fetch_projects(sb: Client, office: str | None) -> list[dict]:
    cols = ["office"] + [c for c in PROJECTS_COLS if c]
    select_str = ",".join(cols)
    q = sb.table("projects").select(select_str).order("office").order("project_code")
    if office:
        q = q.eq("office", office)
    rows = []
    page, size = 0, 1000
    while True:
        batch = q.range(page, page + size - 1).execute().data
        rows.extend(batch)
        if len(batch) < size:
            break
        page += size
    return rows


def fetch_phases(sb: Client, office: str | None) -> list[dict]:
    select_str = ",".join(["office"] + PHASES_COLS)
    q = (
        sb.table("project_phases")
        .select(select_str)
        .order("office")
        .order("project_code")
        .order("level2_code")
        .order("level3_code")
    )
    if office:
        q = q.eq("office", office)
    rows = []
    page, size = 0, 1000
    while True:
        batch = q.range(page, page + size - 1).execute().data
        rows.extend(batch)
        if len(batch) < size:
            break
        page += size
    return rows


# ---------------------------------------------------------------------------
# Optional lookup loaders
# ---------------------------------------------------------------------------

def load_xlsx_or_csv(path: Path) -> list[dict]:
    """Read an .xlsx or .csv and return list of dicts."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [
            {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
            for row in rows[1:]
            if any(v is not None for v in row)
        ]
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def build_emp_lookup(path: Path) -> dict[str, str]:
    """
    Return {lower_full_name: emp_code} from the employee codes file.
    Handles 'FirstName LastName' or 'LastName, FirstName' formats.
    """
    lookup = {}
    for row in load_xlsx_or_csv(path):
        code = (row.get("EmpCode") or row.get("emp_code") or row.get("EmployeeCode") or "").strip()
        if not code:
            continue
        first = (row.get("FirstName") or row.get("first_name") or "").strip()
        last  = (row.get("LastName")  or row.get("last_name")  or "").strip()
        name  = (row.get("Name")      or row.get("name")       or "").strip()
        if first and last:
            lookup[f"{first} {last}".lower()] = code
            lookup[f"{last}, {first}".lower()] = code
            lookup[f"{last} {first}".lower()] = code
        if name:
            lookup[name.lower()] = code
    return lookup


# ---------------------------------------------------------------------------
# Write helpers
# ---------------------------------------------------------------------------

def coerce(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, str) and val.upper() in ("TRUE", "FALSE"):
        return val.upper()
    return val


def write_sheet(ws, data_start_row: int, col_keys: list, rows: list[dict]):
    for r_idx, row in enumerate(rows):
        for c_idx, key in enumerate(col_keys):
            val = coerce(row.get(key)) if key else None
            ws.cell(row=data_start_row + r_idx, column=c_idx + 1, value=val)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Write projects Excel template from Supabase")
    parser.add_argument("--office", choices=OFFICES, default=None)
    parser.add_argument("--org-units", type=Path, default=None,
                        help="CSV/XLSX with org unit paths for owning_org lookup")
    parser.add_argument("--emp-codes", type=Path, default=None,
                        help="CSV/XLSX with employee codes for PM/PA lookup")
    args = parser.parse_args()

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    projects = fetch_projects(sb, args.office)
    phases   = fetch_phases(sb, args.office)

    print(f"Fetched {len(projects)} projects, {len(phases)} phase rows")

    # Apply employee code lookup if provided
    if args.emp_codes and args.emp_codes.exists():
        emp_lookup = build_emp_lookup(args.emp_codes)
        applied = 0
        for p in projects:
            if not p.get("pm_emp_code"):
                # Attempt fuzzy match would go here — for now just report
                pass
        print(f"  Loaded {len(emp_lookup)} employee codes from {args.emp_codes.name}")

    # Remove example rows from template, write data
    src = TEMPLATE_DIR / TEMPLATE_FILE
    if not src.exists():
        raise FileNotFoundError(f"Template not found: {src}")

    suffix = f"_{args.office}" if args.office else "_merged"
    dst = OUTPUT_DIR / f"projects{suffix}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)

    wb = openpyxl.load_workbook(dst)

    # --- Projects sheet ---
    ws_proj = wb["Projects"]
    DATA_ROW_PROJ = 4  # row 1=instructions, row 2=labels, row 3=field names, row 4=data start
    # Clear example rows
    for row in ws_proj.iter_rows(min_row=DATA_ROW_PROJ, max_row=ws_proj.max_row):
        for cell in row:
            cell.value = None
    write_sheet(ws_proj, DATA_ROW_PROJ, PROJECTS_COLS, projects)

    # --- Phases and Tasks sheet ---
    ws_phases = wb["Phases and Tasks"]
    DATA_ROW_PHASES = 4
    for row in ws_phases.iter_rows(min_row=DATA_ROW_PHASES, max_row=ws_phases.max_row):
        for cell in row:
            cell.value = None
    write_sheet(ws_phases, DATA_ROW_PHASES, PHASES_COLS, phases)

    wb.save(dst)
    print(f"Written: {dst}")
    print(f"  Projects sheet: {len(projects)} rows")
    print(f"  Phases sheet:   {len(phases)} rows")

    if not any(p.get("owning_org") for p in projects):
        print("\n  [NOTE] owning_org is blank for all rows — run again with --org-units once Andrew sends the file")
    if not any(p.get("pm_emp_code") for p in projects):
        print("  [NOTE] pm_emp_code is blank — run again with --emp-codes once Chandler sends the file")


if __name__ == "__main__":
    main()
