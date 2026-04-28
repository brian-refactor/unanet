"""
Ajera Employee Utilization Report Parser
Parses the "Employee Utilization" XLS export from Ajera and produces
a clean CSV plus a formatted console report.

Usage:
    python etl/ajera_utilization.py
    python etl/ajera_utilization.py --input "path/to/report.xls"

Output:
    output/cincinnati/cincinnati_utilization.csv
"""

import argparse
import csv
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUTPUT_DIR = HERE.parent / 'output' / 'cincinnati'
PAY_RATES_CSV     = OUTPUT_DIR / 'cincinnati_Employees.csv'
BILLING_RATES_CSV = OUTPUT_DIR / 'cincinnati_billing_rates.csv'

DEFAULT_INPUT = (
    Path.home()
    / 'OneDrive - Willow Creek Partners'
    / 'Shared Documents - Fusion-AE'
    / 'Financial'
    / 'Reztark'
    / 'sample time sheet.xls'
)

DEFAULT_TARGET = 85.0   # applied when Ajera has no target set for an employee

# Column indices as they appear in the Ajera utilization export
COL = {
    'billable':     3,
    'billable_pct': 4,
    'indirect':     5,
    'total':        7,
    'marketing':    17,
    'meetings':     19,
    'vacation':     20,
    'holiday':      21,
    'sick':         22,
    'cont_ed':      23,
    'admin':        24,
    'other':        25,
}


# ---------------------------------------------------------------------------
# Pay rate lookup
# ---------------------------------------------------------------------------

def _norm_name(s: str) -> str:
    """Normalize a name for matching: strip *, remove accents, lowercase, collapse spaces."""
    s = s.lstrip('*').strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')  # strip accents
    s = re.sub(r'\.', '', s)          # remove periods (middle initials)
    s = re.sub(r'\s+', ' ', s).lower().strip()
    return s


def _first_last(norm: str) -> str:
    """Return 'firstname lastname' from a normalized full name."""
    parts = norm.split()
    return f'{parts[0]} {parts[-1]}' if len(parts) >= 2 else norm


def load_pay_rates(csv_path: Path, as_of: date = None) -> dict:
    """
    Load current pay rates from cincinnati_Employees.csv.
    Returns dict keyed by normalized name -> {'hourly_rate', 'annual_salary', 'is_hourly'}.
    'Current' = no end date, or end date after as_of.
    """
    if not csv_path.exists():
        return {}
    as_of = as_of or date.today()

    current = {}
    for row in csv.DictReader(open(csv_path, encoding='utf-8')):
        name = row['EmployeeName'].strip()
        end  = row['PayRateEndDate'].strip()
        is_current = not end
        if not is_current:
            try:
                is_current = date.fromisoformat(end[:10]) > as_of
            except ValueError:
                pass
        if not is_current:
            continue

        try:
            pay_rate = float(row['PayRate']) if row['PayRate'] else 0.0
            sal_pp   = float(row['salaryperpayperiod']) if row['salaryperpayperiod'] else 0.0
        except ValueError:
            continue

        is_hourly = row.get('IsHourly', '').strip().upper() == 'TRUE'
        annual    = pay_rate * 2080 if is_hourly else sal_pp * 26

        current[_norm_name(name)] = {
            'hourly_rate':    round(pay_rate, 4),
            'annual_salary':  round(annual, 2),
            'is_hourly':      is_hourly,
        }

    # Build first+last fallback index
    fl_index = {_first_last(k): v for k, v in current.items()}
    return current, fl_index


def lookup_pay(name: str, pay_lookup: dict, fl_index: dict) -> dict:
    """Find pay data for a utilization employee name."""
    norm = _norm_name(name)
    if norm in pay_lookup:
        return pay_lookup[norm]
    fl = _first_last(norm)
    return fl_index.get(fl, {})


def load_billing_rates(csv_path: Path) -> tuple[dict, dict, dict]:
    """
    Load billing rates from cincinnati_billing_rates.csv.
    Returns (emp_rates, emp_fl, pos_rates):
      emp_rates  — normalized employee name -> rate
      emp_fl     — first+last fallback index
      pos_rates  — lowercase employee type -> rate
    """
    if not csv_path.exists():
        return {}, {}, {}
    emp_rates = {}
    pos_rates = {}
    for row in csv.DictReader(open(csv_path, encoding='utf-8')):
        try:
            rate = float(row.get('BillingRate') or 0)
        except ValueError:
            continue
        if not rate:
            continue
        name = (row.get('Employee') or '').strip()
        pos  = (row.get('EmployeeType') or '').strip()
        if name:
            emp_rates[_norm_name(name)] = rate
        elif pos:
            pos_rates[pos.lower()] = rate
    emp_fl = {_first_last(k): v for k, v in emp_rates.items()}
    return emp_rates, emp_fl, pos_rates


def lookup_billing_rate(name: str, emp_type: str,
                        emp_rates: dict, emp_fl: dict, pos_rates: dict) -> float:
    """Employee-specific rate first, then position fallback."""
    norm = _norm_name(name)
    if norm in emp_rates:
        return emp_rates[norm]
    fl = _first_last(norm)
    if fl in emp_fl:
        return emp_fl[fl]
    return pos_rates.get((emp_type or '').lower(), 0.0)


# ---------------------------------------------------------------------------
# Report parsing
# ---------------------------------------------------------------------------

def _val(row, col, default=0.0):
    try:
        v = row[col]
        return float(v) if v != '' else default
    except (IndexError, TypeError, ValueError):
        return default


def _parse_data_row(row) -> dict:
    return {k: _val(row, c) for k, c in COL.items()}


def _parse_block(rows) -> dict:
    return {
        'current_hours':   _parse_data_row(rows[0]) if len(rows) > 0 else {},
        'current_amounts': _parse_data_row(rows[1]) if len(rows) > 1 else {},
        'ytd_hours':       _parse_data_row(rows[2]) if len(rows) > 2 else {},
        'ytd_amounts':     _parse_data_row(rows[3]) if len(rows) > 3 else {},
        'target_pct':      _val(rows[4], COL['billable_pct']) if len(rows) > 4 else 0.0,
    }


def parse_report(path: Path) -> tuple[dict, list, dict]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    meta = {}
    for i in range(14):
        row = ws.row_values(i)
        for v in row:
            if not isinstance(v, str):
                continue
            if 'period range from:' in v:
                meta['period'] = v.split('from:')[-1].strip()
            elif 'year-to-date range from:' in v.lower():
                meta['ytd_range'] = v.split('from:')[-1].strip()

    employees = []
    totals    = None
    i = 15

    while i < ws.nrows:
        row   = ws.row_values(i)
        label = str(row[0]).strip() if row[0] else ''

        if label == 'Report Totals':
            data_rows = [ws.row_values(i + k) for k in range(1, 6) if i + k < ws.nrows]
            totals = _parse_block(data_rows)
            break

        if label and isinstance(row[10], str) and 'Employee Type:' in row[10]:
            data_rows = [ws.row_values(i + k) for k in range(1, 6) if i + k < ws.nrows]
            emp = {
                'name':      label,
                'type':      row[10].replace('Employee Type: ', '').strip(),
                'hire_date': str(row[22]).replace('Hire Date: ', '').strip() if row[22] else '',
                'status':    str(row[24]).replace('Status: ', '').strip() if row[24] else '',
            }
            emp.update(_parse_block(data_rows))
            employees.append(emp)
            i += 6
            continue

        i += 1

    return meta, employees, totals


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def _ytd_fraction(meta: dict) -> float:
    """Fraction of the year elapsed based on current period end date."""
    try:
        end_str    = meta.get('period', '').split(' to ')[-1].strip()
        end_dt     = datetime.strptime(end_str, '%m/%d/%Y').date()
        year_start = date(end_dt.year, 1, 1)
        year_end   = date(end_dt.year, 12, 31)
        return (end_dt - year_start).days / (year_end - year_start).days
    except Exception:
        today = date.today()
        return today.timetuple().tm_yday / 365


def write_csv(employees: list, path: Path, period_label: str,
              pay_lookup: dict, fl_index: dict, ytd_frac: float):
    rows = []
    for e in [e for e in employees if e['ytd_hours']['total'] > 0]:
        ytd   = e['ytd_hours']
        cur   = e['current_hours']
        ytd_a = e['ytd_amounts']

        ytd_pct = round(ytd['billable'] / ytd['total'] * 100, 1) if ytd['total'] else 0.0
        cur_pct = round(cur['billable'] / cur['total'] * 100, 1) if cur['total'] else 0.0
        target  = e['target_pct'] or DEFAULT_TARGET
        vs_tgt  = round(ytd_pct - target, 1)

        pay = lookup_pay(e['name'], pay_lookup, fl_index)
        annual     = pay.get('annual_salary', '')
        hrly_rate  = pay.get('hourly_rate', 0.0)
        ytd_sal    = round(annual * ytd_frac, 2) if annual else ''
        ytd_bill   = ytd_a['billable']
        recovery   = round(ytd_bill / ytd_sal * 100, 1) if ytd_sal else ''
        multiplier = round(ytd_bill / ytd_sal, 2)        if ytd_sal else ''

        def dollars(hours):
            return round(hours * hrly_rate, 2) if hrly_rate else ''

        rows.append({
            'Employee':               e['name'],
            'Type':                   e['type'],
            'Status':                 e['status'],
            'HireDate':               e['hire_date'],
            'Target_%':               target,
            'HourlyRate':             hrly_rate if hrly_rate else '',
            'AnnualSalary':           annual,
            'YTD_SalaryEst':          ytd_sal,
            'YTD_Total_h':            ytd['total'],
            'YTD_Total_$':            dollars(ytd['total']),
            'YTD_Billable_h':         ytd['billable'],
            'YTD_Billable_$':         dollars(ytd['billable']),
            'YTD_Billable_%':         ytd_pct,
            'vs_Target':              vs_tgt,
            'YTD_CostRecovery_%':     recovery,
            'BillingMultiplier':      multiplier,
            'YTD_Indirect_h':         ytd['indirect'],
            'YTD_Indirect_$':         dollars(ytd['indirect']),
            'YTD_Marketing_h':        ytd['marketing'],
            'YTD_Marketing_$':        dollars(ytd['marketing']),
            'YTD_Meetings_h':         ytd['meetings'],
            'YTD_Meetings_$':         dollars(ytd['meetings']),
            'YTD_Vacation_h':         ytd['vacation'],
            'YTD_Vacation_$':         dollars(ytd['vacation']),
            'YTD_Holiday_h':          ytd['holiday'],
            'YTD_Holiday_$':          dollars(ytd['holiday']),
            'YTD_Sick_h':             ytd['sick'],
            'YTD_Sick_$':             dollars(ytd['sick']),
            'YTD_ContEd_h':           ytd['cont_ed'],
            'YTD_ContEd_$':           dollars(ytd['cont_ed']),
            'YTD_Admin_h':            ytd['admin'],
            'YTD_Admin_$':            dollars(ytd['admin']),
            'YTD_Other_h':            ytd['other'],
            'YTD_Other_$':            dollars(ytd['other']),
            f'{period_label}_Total_h':    cur['total'],
            f'{period_label}_Billable_h': cur['billable'],
            f'{period_label}_Billable_%': cur_pct,
            f'{period_label}_Billable_$': dollars(cur['billable']),
        })

    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)
    print(f'  Wrote {len(rows)} rows -> {path}')


def print_report(employees: list, totals: dict, meta: dict,
                 pay_lookup: dict, fl_index: dict, ytd_frac: float):
    with_hours = [e for e in employees if e['ytd_hours']['total'] > 0]
    sep = '=' * 72

    print(f'\n{sep}')
    print('REZTARK DESIGN STUDIO  —  EMPLOYEE UTILIZATION')
    print(f'Period :  {meta.get("period", "n/a")}')
    print(f'YTD    :  {meta.get("ytd_range", "n/a")}')
    print(sep)

    # Firm-wide summary
    if totals:
        t   = totals['ytd_hours']
        ta  = totals['ytd_amounts']
        ct  = totals['current_hours']
        pct  = round(t['billable']  / t['total']  * 100, 1) if t['total']  else 0
        cpct = round(ct['billable'] / ct['total'] * 100, 1) if ct['total'] else 0

        print('\nFIRM-WIDE SUMMARY')
        print(f'  {"":28}  {"Current period":>16}   {"YTD":>16}')
        print(f'  {"Total hours":<28}  {ct["total"]:>14,.1f}h   {t["total"]:>14,.1f}h')
        print(f'  {"Billable hours":<28}  {ct["billable"]:>14,.1f}h   {t["billable"]:>14,.1f}h')
        print(f'  {"Billable %":<28}  {cpct:>14.1f}%   {pct:>14.1f}%')
        print(f'  {"Billable $ (at cost)":<28}  {"":>16}   ${ta["billable"]:>14,.2f}')
        print(f'  {"Total cost $":<28}  {"":>16}   ${ta["total"]:>14,.2f}')

    # Utilization vs target
    targeted = sorted(
        [e for e in with_hours if e['target_pct'] > 0],
        key=lambda e: (e['ytd_hours']['billable'] / e['ytd_hours']['total'] * 100 - e['target_pct'])
        if e['ytd_hours']['total'] else -999,
    )
    print(f'\nUTILIZATION vs TARGET  ({len(targeted)} employees)')
    print(f'  {"Employee":<28}  {"Type":<24}  {"YTD%":>5}  {"Tgt":>5}  {"Gap":>6}  {"YTD h":>6}')
    print(f'  {"-"*28}  {"-"*24}  {"-"*5}  {"-"*5}  {"-"*6}  {"-"*6}')
    for e in targeted:
        ytd  = e['ytd_hours']
        pct  = ytd['billable'] / ytd['total'] * 100 if ytd['total'] else 0
        gap  = pct - e['target_pct']
        flag = ' !' if gap < -10 else (' ~' if gap < 0 else '  ')
        print(f'{flag} {e["name"]:<28}  {e["type"]:<24}  {pct:>4.1f}%  {e["target_pct"]:>4.0f}%  {gap:>+5.1f}%  {ytd["total"]:>5.1f}h')

    # No target
    no_target = sorted(
        [e for e in with_hours if e['target_pct'] == 0],
        key=lambda e: -e['ytd_hours']['total'],
    )
    print(f'\nNO TARGET SET  ({len(no_target)} employees — principals, directors, admin)')
    print(f'  {"Employee":<28}  {"Type":<24}  {"YTD Bill%":>9}  {"YTD h":>6}')
    print(f'  {"-"*28}  {"-"*24}  {"-"*9}  {"-"*6}')
    for e in no_target:
        ytd = e['ytd_hours']
        pct = ytd['billable'] / ytd['total'] * 100 if ytd['total'] else 0
        print(f'   {e["name"]:<28}  {e["type"]:<24}  {pct:>8.1f}%  {ytd["total"]:>5.1f}h')

    # Salary vs billable recovery
    with_pay = [(e, lookup_pay(e['name'], pay_lookup, fl_index))
                for e in with_hours if lookup_pay(e['name'], pay_lookup, fl_index)]
    with_pay = [(e, p) for e, p in with_pay if p.get('annual_salary', 0) > 0]

    if with_pay:
        with_pay.sort(key=lambda ep: (
            ep[0]['ytd_amounts']['billable'] / (ep[1]['annual_salary'] * ytd_frac)
            if ep[1]['annual_salary'] else 0
        ))

        total_annual  = sum(p['annual_salary'] for _, p in with_pay)
        total_ytd_sal = total_annual * ytd_frac
        total_ytd_bil = sum(e['ytd_amounts']['billable'] for e, _ in with_pay)

        print(f'\nSALARY vs BILLABLE RECOVERY  (YTD thru {round(ytd_frac*100,0):.0f}% of year)')
        print(f'  {"Employee":<28}  {"Annual Sal":>10}  {"YTD Sal Est":>11}  {"YTD Bill $":>10}  {"Recovery":>9}  {"Multiplier":>10}')
        print(f'  {"-"*28}  {"-"*10}  {"-"*11}  {"-"*10}  {"-"*9}  {"-"*10}')
        for e, p in with_pay:
            ytd_bil = e['ytd_amounts']['billable']
            ytd_sal = p['annual_salary'] * ytd_frac
            rec     = ytd_bil / ytd_sal * 100 if ytd_sal else 0
            mult    = ytd_bil / ytd_sal if ytd_sal else 0
            flag    = ' !' if rec < 50 else (' ~' if rec < 80 else '  ')
            print(f'{flag} {e["name"]:<28}  ${p["annual_salary"]:>9,.0f}  ${ytd_sal:>10,.0f}  ${ytd_bil:>9,.0f}  {rec:>8.1f}%  {mult:>10.2f}x')

        print(f'  {"":28}  {"----------":>10}  {"-----------":>11}  {"----------":>10}')
        firm_rec = total_ytd_bil / total_ytd_sal * 100 if total_ytd_sal else 0
        firm_mul = total_ytd_bil / total_ytd_sal if total_ytd_sal else 0
        print(f'   {"TOTAL / FIRM AVG":<28}  ${total_annual:>9,.0f}  ${total_ytd_sal:>10,.0f}  ${total_ytd_bil:>9,.0f}  {firm_rec:>8.1f}%  {firm_mul:>10.2f}x')

    # Indirect breakdown — aggregate dollars using each person's hourly rate
    if totals:
        t   = totals['ytd_hours']
        ind = t['indirect']

        # Sum dollars per category across all employees with hours
        cat_keys = ['admin', 'marketing', 'vacation', 'meetings', 'holiday', 'sick', 'cont_ed', 'other']
        cat_labels = {
            'admin':    'Administration',
            'marketing':'Marketing',
            'vacation': 'Vacation',
            'meetings': 'Meetings',
            'holiday':  'Holiday',
            'sick':     'Sick',
            'cont_ed':  'Cont. Education',
            'other':    'Other',
        }
        cat_dollars = {k: 0.0 for k in cat_keys}
        for e in with_hours:
            pay = lookup_pay(e['name'], pay_lookup, fl_index)
            rate = pay.get('hourly_rate', 0.0)
            if not rate:
                continue
            for k in cat_keys:
                cat_dollars[k] += e['ytd_hours'].get(k, 0.0) * rate

        total_indirect_cost = sum(cat_dollars.values())

        print(f'\nINDIRECT TIME BREAKDOWN  (YTD firm-wide, {ind:,.1f}h  /  ${total_indirect_cost:,.0f})')
        print(f'  {"Category":<20}  {"Hours":>7}  {"Cost $":>10}  {"% Indirect":>10}  {"% Total h":>9}')
        print(f'  {"-"*20}  {"-"*7}  {"-"*10}  {"-"*10}  {"-"*9}')
        for k in sorted(cat_keys, key=lambda k: -t[k]):
            hours = t[k]
            if hours <= 0:
                continue
            cost    = cat_dollars[k]
            pct_ind = hours / ind          * 100 if ind          else 0
            pct_tot = hours / t['total']   * 100 if t['total']   else 0
            print(f'  {cat_labels[k]:<20}  {hours:>7.1f}h  ${cost:>9,.0f}  {pct_ind:>9.1f}%  {pct_tot:>8.1f}%')

    print()


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

_DARK   = '1F4E79'
_BLUE   = '2E75B6'
_LBLUE  = 'DEEAF1'
_GREEN  = 'E2EFDA'
_AMBER  = 'FFF2CC'
_RED    = 'FCE4D6'
_RFONT  = 'C00000'
_GRAY   = 'F2F2F2'
_WHITE  = 'FFFFFF'

_CAT_KEYS = ['admin', 'marketing', 'vacation', 'meetings',
             'holiday', 'sick', 'cont_ed', 'other']
_CAT_LABELS = {
    'admin': 'Administration', 'marketing': 'Marketing',
    'vacation': 'Vacation',    'meetings': 'Meetings',
    'holiday': 'Holiday',      'sick': 'Sick',
    'cont_ed': 'Cont. Education', 'other': 'Other',
}


def _xf(c):
    return PatternFill('solid', fgColor=c)

def _xb():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _xcell(ws, row, col, value, *, fill=None, bold=False, italic=False,
           size=10, color='000000', fmt=None, halign='center', wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
    if fill:
        c.fill = _xf(fill)
    c.border = _xb()
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    return c

def _hdr(ws, row, col, text):
    return _xcell(ws, row, col, text, fill=_BLUE, bold=True, color=_WHITE, wrap=True)

def _title(ws, row, ncols, text, height=30):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14, color=_WHITE, name='Calibri')
    c.fill = _xf(_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = height

def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col) if isinstance(col, int) else col].width = w


def write_excel(employees: list, totals: dict, meta: dict,
                pay_lookup: dict, fl_index: dict, ytd_frac: float,
                out_path: Path, period_label: str,
                emp_rates: dict = None, emp_fl: dict = None, pos_rates: dict = None):

    period_str = meta.get('period', '')
    emp_rates = emp_rates or {}
    emp_fl    = emp_fl    or {}
    pos_rates = pos_rates or {}

    # ── per-employee data ─────────────────────────────────────────────────
    rows = []
    for e in employees:
        if e['ytd_hours']['total'] <= 0:
            continue
        ytd   = e['ytd_hours']
        cur   = e['current_hours']
        ytd_a = e['ytd_amounts']
        target    = e['target_pct'] or DEFAULT_TARGET
        ytd_pct   = ytd['billable'] / ytd['total'] * 100 if ytd['total'] else 0.0
        cur_pct   = cur['billable'] / cur['total'] * 100 if cur['total'] else 0.0
        vs_tgt    = ytd_pct - target
        pay       = lookup_pay(e['name'], pay_lookup, fl_index)
        annual    = pay.get('annual_salary') or 0.0
        hrly      = pay.get('hourly_rate') or 0.0
        ytd_sal   = annual * ytd_frac if annual else 0.0
        ytd_bil   = ytd_a['billable']               # cost $ from Ajera report
        recovery  = ytd_bil / ytd_sal * 100 if ytd_sal else None
        mult      = ytd_bil / ytd_sal if ytd_sal else None

        bill_rate = lookup_billing_rate(e['name'], e['type'], emp_rates, emp_fl, pos_rates)
        ytd_rev   = ytd['billable'] * bill_rate if bill_rate else 0.0
        margin_d  = ytd_rev - ytd_bil if bill_rate else None
        margin_pct = margin_d / ytd_rev * 100 if ytd_rev else None

        def dlr(h): return h * hrly if hrly else 0.0

        rows.append({
            'name':      e['name'],
            'type':      e['type'],
            'status':    e['status'],
            'target':    target,
            'ytd_h':     ytd['total'],
            'bill_h':    ytd['billable'],
            'bill_d':    ytd_bil,
            'ytd_pct':   ytd_pct,
            'cur_pct':   cur_pct,
            'vs_tgt':    vs_tgt,
            'annual':    annual,
            'hrly':      hrly,
            'ytd_sal':   ytd_sal,
            'recovery':  recovery,
            'mult':      mult,
            'bill_rate': bill_rate,
            'ytd_rev':   ytd_rev,
            'margin_d':  margin_d,
            'margin_pct': margin_pct,
            **{f'{k}_h': ytd[k]      for k in _CAT_KEYS},
            **{f'{k}_d': dlr(ytd[k]) for k in _CAT_KEYS},
        })

    # ── firm totals ───────────────────────────────────────────────────────
    n          = len(rows)
    tot_h      = sum(r['ytd_h']   for r in rows)
    tot_bill_h = sum(r['bill_h']  for r in rows)
    tot_bill_d = sum(r['bill_d']  for r in rows)
    tot_sal    = sum(r['ytd_sal'] for r in rows if r['ytd_sal'])
    tot_rev    = sum(r['ytd_rev'] for r in rows if r['ytd_rev'])
    tot_margin = tot_rev - tot_bill_d if tot_rev else 0.0
    firm_pct   = tot_bill_h / tot_h   * 100 if tot_h   else 0.0
    firm_rec   = tot_bill_d / tot_sal * 100 if tot_sal else 0.0
    firm_mult  = tot_bill_d / tot_sal       if tot_sal else 0.0
    firm_margin_pct = tot_margin / tot_rev * 100 if tot_rev else 0.0
    avg_tgt    = sum(r['target'] for r in rows) / n if n else DEFAULT_TARGET
    below_10   = sorted([r for r in rows if r['vs_tgt'] < -10], key=lambda r: r['vs_tgt'])
    above_tgt  = [r for r in rows if r['vs_tgt'] >= 0]
    cat_h      = {k: sum(r[f'{k}_h'] for r in rows) for k in _CAT_KEYS}
    cat_d      = {k: sum(r[f'{k}_d'] for r in rows) for k in _CAT_KEYS}
    tot_ind_h  = sum(cat_h.values())
    tot_ind_d  = sum(cat_d.values())
    has_billing = tot_rev > 0

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Tab 1 — Executive Summary
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Executive Summary'
    ws.sheet_view.showGridLines = False

    _title(ws, 1, 11, f'Reztark Design Studio  —  Employee Utilization  |  {period_str}', height=34)

    # KPI tiles: two merged columns each → 5 tiles across B:K
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 44
    ws.row_dimensions[5].height = 16

    if has_billing:
        kpis = [
            (2, 3,  'Active Staff',       str(n),                          period_label),
            (4, 5,  'Firm Billable %',    f'{firm_pct:.1f}%',              f'Target {avg_tgt:.0f}%'),
            (6, 7,  'YTD Billing Rev',    f'${tot_rev:,.0f}',              f'{tot_bill_h:,.0f} billable hours'),
            (8, 9,  'YTD Gross Margin',   f'${tot_margin:,.0f}',           f'{firm_margin_pct:.1f}% margin'),
            (10, 11,'Cost Recovery',      f'{firm_rec:.1f}%',              f'${tot_bill_d:,.0f} / ${tot_sal:,.0f}'),
        ]
    else:
        kpis = [
            (2, 3,  'Active Staff',        str(n),                           period_label),
            (4, 5,  'Firm Billable %',     f'{firm_pct:.1f}%',               f'Target {avg_tgt:.0f}%'),
            (6, 7,  'YTD Billable Hours',  f'{tot_bill_h:,.0f} h',           f'of {tot_h:,.0f} total'),
            (8, 9,  'YTD Billable $',      f'${tot_bill_d:,.0f}',            'at cost rates'),
            (10, 11,'Cost Recovery',       f'{firm_rec:.1f}%',               f'${tot_bill_d:,.0f} / ${tot_sal:,.0f}'),
        ]
    for c1, c2, lbl, val, note in kpis:
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
        _xcell(ws, 3, c1, lbl, fill=_BLUE, bold=True, color=_WHITE)
        ws.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
        c = _xcell(ws, 4, c1, val, fill=_LBLUE, bold=True, color=_DARK, size=20)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
        _xcell(ws, 5, c1, note, fill=_LBLUE, italic=True, color='595959', size=9)

    _widths(ws, {1: 2, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16, 9: 16, 10: 16, 11: 16})

    # Section heading
    ws.row_dimensions[6].height = 10
    ws.row_dimensions[7].height = 18
    ws.merge_cells('B7:K7')
    c = ws.cell(row=7, column=2, value='Executive Summary')
    c.font = Font(bold=True, size=13, color=_DARK, name='Calibri')

    # Narrative
    n_below   = len(below_10)
    top3      = sorted(above_tgt, key=lambda r: -r['vs_tgt'])[:3]
    top_names = ', '.join(r['name'].split()[0] for r in top3) if top3 else 'N/A'

    if n_below == 0:
        watch_s = 'All employees are meeting or exceeding their 85% billable target.'
    elif n_below == 1:
        r0 = below_10[0]
        watch_s = (f'One employee warrants attention: {r0["name"]} '
                   f'({r0["ytd_pct"]:.1f}% actual vs {r0["target"]:.0f}% target).')
    else:
        names = ', '.join(r['name'] for r in below_10[:4])
        if n_below > 4:
            names += f' and {n_below - 4} others'
        watch_s = f'{n_below} employees are more than 10 points below target: {names}.'

    if has_billing:
        narrative = (
            f'Through {period_str}, Reztark logged {tot_h:,.0f} total hours — {tot_bill_h:,.0f} '
            f'({firm_pct:.1f}%) billable against a firm standard target of {avg_tgt:.0f}%. '
            f'The firm is approximately {round(ytd_frac * 100)}% through the fiscal year. '
            f'YTD billing revenue is ${tot_rev:,.0f} against a cost of ${tot_bill_d:,.0f}, '
            f'producing a gross margin of ${tot_margin:,.0f} ({firm_margin_pct:.1f}%). '
            f'Cost recovery against estimated YTD payroll of ${tot_sal:,.0f} is {firm_rec:.1f}%. '
            f'{watch_s} '
            f'Top performers above target: {top_names}.'
        )
    else:
        narrative = (
            f'Through {period_str}, Reztark logged {tot_h:,.0f} total hours — {tot_bill_h:,.0f} '
            f'({firm_pct:.1f}%) billable against a firm standard target of {avg_tgt:.0f}%. '
            f'The firm is approximately {round(ytd_frac * 100)}% through the fiscal year. '
            f'YTD billable revenue at cost is ${tot_bill_d:,.0f}, a cost recovery rate of '
            f'{firm_rec:.1f}% against estimated YTD payroll of ${tot_sal:,.0f}. '
            f'{watch_s} '
            f'Top performers above target: {top_names}.'
        )
    ws.row_dimensions[8].height = 72
    ws.merge_cells('B8:K8')
    c = ws.cell(row=8, column=2, value=narrative)
    c.font = Font(size=11, name='Calibri')
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Watch list
    if below_10:
        ws.row_dimensions[9].height = 10
        ws.row_dimensions[10].height = 16
        ws.merge_cells('B10:K10')
        c = ws.cell(row=10, column=2,
                    value=f'Watch List  —  {n_below} employee(s) more than 10 points below target')
        c.font = Font(bold=True, size=12, color=_RFONT, name='Calibri')

        ws.row_dimensions[11].height = 20
        for ci, h in enumerate(['Employee', 'Type', 'YTD Bill %', 'Target %', 'Gap', 'YTD Hours'], start=2):
            _hdr(ws, 11, ci, h)

        for ri, r in enumerate(below_10, start=12):
            ws.row_dimensions[ri].height = 16
            for ci, (v, fmt, ha) in enumerate([
                (r['name'],    None,               'left'),
                (r['type'],    None,               'left'),
                (r['ytd_pct'], '0.0"%"',           'center'),
                (r['target'],  '0"%"',             'center'),
                (r['vs_tgt'],  '+0.0"%";-0.0"%"',  'center'),
                (r['ytd_h'],   '#,##0.0',          'right'),
            ], start=2):
                _xcell(ws, ri, ci, v, fill=_RED, color=_RFONT, fmt=fmt, halign=ha)

    # ════════════════════════════════════════════════════════════════════════
    # Tab 2 — Utilization by Employee
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Utilization by Employee')
    ws2.sheet_view.showGridLines = False

    _title(ws2, 1, 9, f'Employee Utilization  —  YTD {period_str}')

    hdrs2 = ['Employee', 'Type', 'Status',
             f'{period_label}\nBill %', 'YTD\nBill %', 'Target\n%', 'Gap vs\nTarget',
             'YTD\nBill h', 'YTD\nTotal h']
    ws2.row_dimensions[2].height = 28
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 2, ci, h)

    util_sorted = sorted(rows, key=lambda r: r['ytd_pct'])
    for ri, r in enumerate(util_sorted, start=3):
        ws2.row_dimensions[ri].height = 16
        fill = _GREEN if r['vs_tgt'] >= 0 else (_AMBER if r['vs_tgt'] >= -10 else _RED)
        for ci, (v, fmt, ha) in enumerate([
            (r['name'],    None,               'left'),
            (r['type'],    None,               'left'),
            (r['status'],  None,               'center'),
            (r['cur_pct'], '0.0"%"',           'center'),
            (r['ytd_pct'], '0.0"%"',           'center'),
            (r['target'],  '0"%"',             'center'),
            (r['vs_tgt'],  '+0.0"%";-0.0"%"',  'center'),
            (r['bill_h'],  '#,##0.0',          'right'),
            (r['ytd_h'],   '#,##0.0',          'right'),
        ], start=1):
            _xcell(ws2, ri, ci, v, fill=fill, fmt=fmt, halign=ha)

    _widths(ws2, {1: 26, 2: 22, 3: 10, 4: 12, 5: 12, 6: 10, 7: 12, 8: 13, 9: 13})

    leg = len(util_sorted) + 4
    ws2.merge_cells(f'A{leg}:I{leg}')
    c = ws2.cell(row=leg, column=1,
                 value='Green = at or above 85% target   |   Amber = within 10 pts below   |   Red = more than 10 pts below')
    c.font = Font(italic=True, color='595959', size=9, name='Calibri')

    # Chart data (below legend)
    cd = leg + 2
    ws2.cell(row=cd, column=1, value='Employee')
    ws2.cell(row=cd, column=2, value='YTD Bill %')
    ws2.cell(row=cd, column=3, value='Target %')
    for i, r in enumerate(util_sorted, start=cd + 1):
        parts = r['name'].split()
        ws2.cell(row=i, column=1, value=f"{parts[-1]}, {parts[0]}")
        ws2.cell(row=i, column=2, value=round(r['ytd_pct'], 1))
        ws2.cell(row=i, column=3, value=round(r['target'], 0))

    bar = BarChart()
    bar.type = 'bar'
    bar.grouping = 'clustered'
    bar.title = 'YTD Billable % vs 85% Target'
    bar.x_axis.title = 'Billable %'
    bar.style = 10
    bar.width  = 20
    bar.height = max(12, len(util_sorted) * 0.52)
    d = Reference(ws2, min_col=2, max_col=3, min_row=cd, max_row=cd + len(util_sorted))
    cats = Reference(ws2, min_col=1, min_row=cd + 1, max_row=cd + len(util_sorted))
    bar.add_data(d, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = _BLUE
    bar.series[0].graphicalProperties.line.solidFill = _BLUE
    if len(bar.series) > 1:
        bar.series[1].graphicalProperties.solidFill = 'BFBFBF'
        bar.series[1].graphicalProperties.line.solidFill = 'BFBFBF'
    ws2.add_chart(bar, 'K3')

    # ════════════════════════════════════════════════════════════════════════
    # Tab 3 — Salary & Cost Recovery
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Salary & Cost Recovery')
    ws3.sheet_view.showGridLines = False

    _title(ws3, 1, 8, f'Salary & Billable Cost Recovery  —  YTD  ({round(ytd_frac * 100):.0f}% of year elapsed)')

    hdrs3 = ['Employee', 'Type', 'Hourly Rate', 'Annual Salary',
             'YTD Salary Est', 'YTD Billable $\n(cost)', 'Recovery %', 'Multiplier']
    ws3.row_dimensions[2].height = 28
    for ci, h in enumerate(hdrs3, 1):
        _hdr(ws3, 2, ci, h)

    sal_rows = sorted([r for r in rows if r['ytd_sal'] > 0],
                      key=lambda r: (r['recovery'] or 0))
    no_rate  = [r for r in rows if r['ytd_sal'] == 0]

    for ri, r in enumerate(sal_rows, start=3):
        ws3.row_dimensions[ri].height = 16
        rec  = r['recovery'] or 0
        fill = _GREEN if rec >= 80 else (_AMBER if rec >= 50 else _RED)
        for ci, (v, fmt, ha) in enumerate([
            (r['name'],     None,        'left'),
            (r['type'],     None,        'left'),
            (r['hrly'],     '$#,##0.00', 'right'),
            (r['annual'],   '$#,##0',    'right'),
            (r['ytd_sal'],  '$#,##0',    'right'),
            (r['bill_d'],   '$#,##0',    'right'),
            (r['recovery'], '0.0"%"',    'center'),
            (r['mult'],     '0.00"x"',   'center'),
        ], start=1):
            _xcell(ws3, ri, ci, v or 0, fill=fill, fmt=fmt, halign=ha)

    # Firm total row
    tr = len(sal_rows) + 3
    ws3.row_dimensions[tr].height = 18
    for ci, (v, fmt, ha) in enumerate([
        ('FIRM TOTAL',                                None,      'left'),
        ('',                                          None,      'left'),
        ('',                                          None,      'right'),
        (sum(r['annual'] for r in sal_rows),          '$#,##0',  'right'),
        (tot_sal,                                     '$#,##0',  'right'),
        (tot_bill_d,                                  '$#,##0',  'right'),
        (firm_rec,                                    '0.0"%"',  'center'),
        (firm_mult,                                   '0.00"x"', 'center'),
    ], start=1):
        _xcell(ws3, tr, ci, v, fill=_LBLUE, bold=True, fmt=fmt, halign=ha)

    if no_rate:
        nr_row = tr + 2
        ws3.merge_cells(f'A{nr_row}:H{nr_row}')
        names = ', '.join(r['name'] for r in no_rate[:6])
        if len(no_rate) > 6:
            names += f' (+{len(no_rate) - 6} more)'
        c = ws3.cell(row=nr_row, column=1, value=f'Pay rate not on file: {names}')
        c.font = Font(italic=True, color='595959', size=9, name='Calibri')

    _widths(ws3, {1: 26, 2: 22, 3: 13, 4: 15, 5: 15, 6: 15, 7: 13, 8: 12})

    # Recovery bar chart
    rc = tr + (4 if no_rate else 3)
    ws3.cell(row=rc, column=1, value='Employee')
    ws3.cell(row=rc, column=2, value='Recovery %')
    ws3.cell(row=rc, column=3, value='80% Threshold')
    for i, r in enumerate(sal_rows, start=rc + 1):
        parts = r['name'].split()
        ws3.cell(row=i, column=1, value=f"{parts[-1]}, {parts[0]}")
        ws3.cell(row=i, column=2, value=round(r['recovery'], 1) if r['recovery'] else 0)
        ws3.cell(row=i, column=3, value=80)

    bar2 = BarChart()
    bar2.type = 'bar'
    bar2.grouping = 'clustered'
    bar2.title = 'Cost Recovery % (YTD Billable ÷ YTD Salary)'
    bar2.x_axis.title = 'Recovery %'
    bar2.style = 10
    bar2.width  = 20
    bar2.height = max(12, len(sal_rows) * 0.52)
    d2    = Reference(ws3, min_col=2, max_col=3, min_row=rc, max_row=rc + len(sal_rows))
    cats2 = Reference(ws3, min_col=1, min_row=rc + 1, max_row=rc + len(sal_rows))
    bar2.add_data(d2, titles_from_data=True)
    bar2.set_categories(cats2)
    bar2.series[0].graphicalProperties.solidFill = _BLUE
    bar2.series[0].graphicalProperties.line.solidFill = _BLUE
    if len(bar2.series) > 1:
        bar2.series[1].graphicalProperties.solidFill = 'BFBFBF'
        bar2.series[1].graphicalProperties.line.solidFill = 'BFBFBF'
    ws3.add_chart(bar2, 'J3')

    # ════════════════════════════════════════════════════════════════════════
    # Tab 4 — Profitability by Employee  (only when billing rates available)
    # ════════════════════════════════════════════════════════════════════════
    if has_billing:
        ws_p = wb.create_sheet('Profitability')
        ws_p.sheet_view.showGridLines = False

        _title(ws_p, 1, 9,
               f'Profitability by Employee  —  YTD {period_str}  '
               f'(billing revenue vs cost at {round(ytd_frac*100):.0f}% of year)')

        ws_p.row_dimensions[2].height = 10
        ws_p.merge_cells('A3:I3')
        c = ws_p.cell(row=3, column=1,
                      value=('Billing Revenue = YTD billable hours × billing rate.   '
                             'Cost = YTD billable $ at cost from Ajera.   '
                             'Gross Margin = Revenue − Cost.'))
        c.font = Font(italic=True, color='595959', size=9, name='Calibri')
        c.alignment = Alignment(horizontal='left')
        ws_p.row_dimensions[3].height = 14
        ws_p.row_dimensions[4].height = 6

        hdrs_p = ['Employee', 'Type', 'Cost Rate\n($/hr)', 'Billing Rate\n($/hr)',
                  'Markup', 'YTD Bill h',
                  'YTD Revenue\n(billing)', 'YTD Cost\n(Ajera)', 'Gross Margin\n$', 'Margin %']
        ws_p.row_dimensions[5].height = 28
        for ci, h in enumerate(hdrs_p, 1):
            _hdr(ws_p, 5, ci, h)

        prof_rows = sorted([r for r in rows if r['bill_rate'] > 0],
                           key=lambda r: (r['margin_pct'] or 0))
        no_bill   = [r for r in rows if not r['bill_rate']]

        for ri, r in enumerate(prof_rows, start=6):
            ws_p.row_dimensions[ri].height = 16
            mp = r['margin_pct'] or 0
            fill = _GREEN if mp >= 40 else (_AMBER if mp >= 20 else _RED)
            markup = r['bill_rate'] / r['hrly'] if r['hrly'] else None
            for ci, (v, fmt, ha) in enumerate([
                (r['name'],      None,        'left'),
                (r['type'],      None,        'left'),
                (r['hrly'],      '$#,##0.00', 'right'),
                (r['bill_rate'], '$#,##0.00', 'right'),
                (markup,         '0.00"x"',   'center'),
                (r['bill_h'],    '#,##0.0',   'right'),
                (r['ytd_rev'],   '$#,##0',    'right'),
                (r['bill_d'],    '$#,##0',    'right'),
                (r['margin_d'],  '$#,##0',    'right'),
                (r['margin_pct'],'0.0"%"',    'center'),
            ], start=1):
                _xcell(ws_p, ri, ci, v or 0, fill=fill, fmt=fmt, halign=ha)

        # Firm totals
        ptr = len(prof_rows) + 6
        ws_p.row_dimensions[ptr].height = 18
        tot_markup = tot_rev / tot_bill_d if tot_bill_d else 0
        for ci, (v, fmt, ha) in enumerate([
            ('FIRM TOTAL', None,       'left'),
            ('',           None,       'left'),
            ('',           None,       'right'),
            ('',           None,       'right'),
            (tot_markup,   '0.00"x"',  'center'),
            (tot_bill_h,   '#,##0.0',  'right'),
            (tot_rev,      '$#,##0',   'right'),
            (tot_bill_d,   '$#,##0',   'right'),
            (tot_margin,   '$#,##0',   'right'),
            (firm_margin_pct, '0.0"%"','center'),
        ], start=1):
            _xcell(ws_p, ptr, ci, v, fill=_LBLUE, bold=True, fmt=fmt, halign=ha)

        if no_bill:
            nb_row = ptr + 2
            ws_p.merge_cells(f'A{nb_row}:J{nb_row}')
            names = ', '.join(r['name'] for r in no_bill[:6])
            if len(no_bill) > 6:
                names += f' (+{len(no_bill) - 6} more)'
            c = ws_p.cell(row=nb_row, column=1, value=f'Billing rate not on file: {names}')
            c.font = Font(italic=True, color='595959', size=9, name='Calibri')

        _widths(ws_p, {1: 26, 2: 22, 3: 13, 4: 14, 5: 10, 6: 11, 7: 16, 8: 14, 9: 15, 10: 11})

        # Margin % bar chart
        mc = ptr + (4 if no_bill else 3)
        ws_p.cell(row=mc, column=1, value='Employee')
        ws_p.cell(row=mc, column=2, value='Margin %')
        ws_p.cell(row=mc, column=3, value='40% Target')
        for i, r in enumerate(prof_rows, start=mc + 1):
            parts = r['name'].split()
            ws_p.cell(row=i, column=1, value=f"{parts[-1]}, {parts[0]}")
            ws_p.cell(row=i, column=2, value=round(r['margin_pct'], 1) if r['margin_pct'] else 0)
            ws_p.cell(row=i, column=3, value=40)

        barp = BarChart()
        barp.type = 'bar'
        barp.grouping = 'clustered'
        barp.title = 'Gross Margin % by Employee'
        barp.x_axis.title = 'Margin %'
        barp.style = 10
        barp.width  = 20
        barp.height = max(12, len(prof_rows) * 0.52)
        dp   = Reference(ws_p, min_col=2, max_col=3, min_row=mc, max_row=mc + len(prof_rows))
        catsp = Reference(ws_p, min_col=1, min_row=mc + 1, max_row=mc + len(prof_rows))
        barp.add_data(dp, titles_from_data=True)
        barp.set_categories(catsp)
        barp.series[0].graphicalProperties.solidFill = _BLUE
        barp.series[0].graphicalProperties.line.solidFill = _BLUE
        if len(barp.series) > 1:
            barp.series[1].graphicalProperties.solidFill = 'BFBFBF'
            barp.series[1].graphicalProperties.line.solidFill = 'BFBFBF'
        ws_p.add_chart(barp, 'L6')

    # ════════════════════════════════════════════════════════════════════════
    # Tab 4 — Indirect Cost Breakdown
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Indirect Cost Breakdown')
    ws4.sheet_view.showGridLines = False

    _title(ws4, 1, 7, f'Indirect Time & Cost Breakdown  —  YTD {period_str}')

    ws4.row_dimensions[2].height = 10
    ws4.merge_cells('A3:G3')
    c = ws4.cell(row=3, column=1,
                 value=(f'Total indirect hours: {tot_ind_h:,.1f}   |   '
                        f'Estimated indirect cost: ${tot_ind_d:,.0f}   |   '
                        f'Dollar values estimated using each employee\'s effective hourly rate.'))
    c.font = Font(italic=True, color='595959', size=10, name='Calibri')
    c.alignment = Alignment(horizontal='left')
    ws4.row_dimensions[3].height = 16
    ws4.row_dimensions[4].height = 6

    hdrs4 = ['Category', 'YTD Hours', '% of Indirect h', '% of Total h',
             'Est. Cost $', '% of Indirect $', 'Notes']
    ws4.row_dimensions[5].height = 20
    for ci, h in enumerate(hdrs4, 1):
        _hdr(ws4, 5, ci, h)

    cat_order  = sorted(_CAT_KEYS, key=lambda k: -cat_h[k])
    pie_colors = ['2E75B6', '70AD47', 'ED7D31', 'FFC000', 'FF0000', '4472C4', 'A9D18E', 'F4B183']

    for ri, k in enumerate(cat_order, start=6):
        ws4.row_dimensions[ri].height = 18
        h    = cat_h[k]
        d    = cat_d[k]
        fill = _LBLUE if ri % 2 == 0 else _WHITE
        for ci, (v, fmt, ha) in enumerate([
            (_CAT_LABELS[k], None,      'left'),
            (h,  '#,##0.0',  'right'),
            (h / tot_ind_h * 100 if tot_ind_h else 0, '0.0"%"', 'center'),
            (h / tot_h      * 100 if tot_h      else 0, '0.0"%"', 'center'),
            (d,  '$#,##0',   'right'),
            (d / tot_ind_d * 100 if tot_ind_d else 0, '0.0"%"', 'center'),
            ('', None, 'left'),
        ], start=1):
            _xcell(ws4, ri, ci, v, fill=fill, fmt=fmt, halign=ha)

    tot4 = len(cat_order) + 6
    ws4.row_dimensions[tot4].height = 18
    for ci, (v, fmt, ha) in enumerate([
        ('TOTAL',   None,      'left'),
        (tot_ind_h, '#,##0.0', 'right'),
        (100.0,     '0"%"',    'center'),
        (tot_ind_h / tot_h * 100 if tot_h else 0, '0.0"%"', 'center'),
        (tot_ind_d, '$#,##0',  'right'),
        (100.0,     '0"%"',    'center'),
        ('', None, 'left'),
    ], start=1):
        _xcell(ws4, tot4, ci, v, fill=_LBLUE, bold=True, fmt=fmt, halign=ha)

    _widths(ws4, {1: 20, 2: 13, 3: 17, 4: 15, 5: 14, 6: 18, 7: 22})

    # Pie chart
    pd = tot4 + 3
    ws4.cell(row=pd, column=1, value='Category')
    ws4.cell(row=pd, column=2, value='Est. Cost $')
    for i, k in enumerate(cat_order, start=pd + 1):
        ws4.cell(row=i, column=1, value=_CAT_LABELS[k])
        ws4.cell(row=i, column=2, value=round(cat_d[k], 0))

    pie = PieChart()
    pie.title  = 'Indirect Cost Distribution'
    pie.style  = 10
    pie.width  = 16
    pie.height = 12
    pie.add_data(Reference(ws4, min_col=2, min_row=pd, max_row=pd + len(cat_order)),
                 titles_from_data=True)
    pie.set_categories(Reference(ws4, min_col=1, min_row=pd + 1, max_row=pd + len(cat_order)))
    for i, color in enumerate(pie_colors[:len(cat_order)]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    ws4.add_chart(pie, 'I5')

    # ── save ─────────────────────────────────────────────────────────────
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f'  Excel report -> {out_path}')


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default=str(DEFAULT_INPUT),
                        help='Path to Ajera Employee Utilization XLS export')
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        raise SystemExit(f'File not found: {path}')

    print(f'Parsing {path.name}...')
    meta, employees, totals = parse_report(path)
    active_count = sum(1 for e in employees if e['status'] == 'Active')
    print(f'  {len(employees)} employees ({active_count} active)')

    pay_lookup, fl_index = load_pay_rates(PAY_RATES_CSV)
    matched = sum(1 for e in employees if lookup_pay(e['name'], pay_lookup, fl_index))
    print(f'  Pay rates matched: {matched}/{len(employees)}')

    emp_rates, emp_fl, pos_rates = load_billing_rates(BILLING_RATES_CSV)
    bill_matched = sum(
        1 for e in employees
        if lookup_billing_rate(e['name'], e['type'], emp_rates, emp_fl, pos_rates)
    )
    print(f'  Billing rates matched: {bill_matched}/{len(employees)}')

    ytd_frac = _ytd_fraction(meta)

    period_label = 'Current'
    try:
        start_str    = meta.get('period', '').split(' to ')[0].strip()
        period_label = datetime.strptime(start_str, '%m/%d/%Y').strftime('%b%Y')
    except Exception:
        pass

    write_csv(employees, OUTPUT_DIR / 'cincinnati_utilization.csv',
              period_label, pay_lookup, fl_index, ytd_frac)

    xl_path = OUTPUT_DIR / f'cincinnati_utilization_{period_label}.xlsx'
    write_excel(employees, totals, meta, pay_lookup, fl_index, ytd_frac,
                xl_path, period_label, emp_rates, emp_fl, pos_rates)

    print_report(employees, totals, meta, pay_lookup, fl_index, ytd_frac)


if __name__ == '__main__':
    main()
