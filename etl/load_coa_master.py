"""
Load consolidated COA data into Supabase:
  1. coa_master  — 389 master accounts from the Fusion template + working file
  2. coa_crosswalk — source office account → master code mappings

Truncates both tables before loading (full replace).

Usage:
    python etl/load_coa_master.py
"""
import os
from pathlib import Path
from collections import OrderedDict

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(Path(__file__).parent / '.env')
sb = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_KEY'])

OFFICE_MAP = {'DAL': 'dallas', 'MSP': 'minnesota', 'CIN': 'cincinnati', 'ORL': 'orlando'}

def to_bool(val):
    if val is None:
        return False
    return str(val).strip().upper() in ('TRUE', 'YES', '1')

def blank(val):
    v = str(val or '').strip()
    return v if v else None

# ── Load Fusion template (authoritative Unanet fields) ────────────────────────
wb_fusion = openpyxl.load_workbook(r'E:\Unanet\output\02-COA_Fusion 6-2-26_updated.xlsx')
ws_fusion = wb_fusion['Chart of Accounts']
# Row 3: BaseCode(1) BaseName(2) Desc(3) IsActive(4) Is1099(5) IsSubcon(6)
#        FinType(7) Subledger(8) Metric(9) CostType(10) PMType(11) LaborRev(12) ExpRev(13)
fusion = {}
for row in ws_fusion.iter_rows(min_row=4, values_only=True):
    if not row[0]:
        continue
    fusion[str(row[0]).strip()] = {
        'base_name':          blank(row[1]),
        'description':        blank(row[2]),
        'is_active':          to_bool(row[3]),
        'is_1099':            to_bool(row[4]),
        'is_subcontractor':   to_bool(row[5]),
        'financial_type':     blank(row[6]),
        'subledger_type':     blank(row[7]),
        'metric_type':        blank(row[8]),
        'cost_type':          blank(row[9]),
        'pm_type':            blank(row[10]),
        'labor_revenue_type': blank(row[11]),
        'expense_revenue_type': blank(row[12]),
    }

# ── Load working file COA MATRIX (section, rollup, notes) ────────────────────
wb_src = openpyxl.load_workbook(r'E:\Unanet\output\COA_Working File-6-2-26.xlsx')
ws_matrix = wb_src['COA MATRIX']
src_rows = list(ws_matrix.iter_rows(min_row=2, values_only=True))

FT_NORM = {'Equity': 'Capital', 'Revenue': 'Income'}
OFFICES = set(OFFICE_MAP.keys())

# Gather section and notes per master code from the matrix
meta = OrderedDict()   # str(code) -> {section, notes, sort_order}
sort_order = 0
for r in src_rows:
    code = r[2]
    if not code:
        continue
    key = str(code).strip()
    if key not in meta:
        sort_order += 1
        meta[key] = {
            'section':    blank(r[9]),
            'notes':      blank(r[19]),
            'sort_order': sort_order,
        }

# ── Build coa_master rows ─────────────────────────────────────────────────────
master_rows = []
for key, m in meta.items():
    fus = fusion.get(key, {})
    master_rows.append({
        'master_code':          key,
        'master_name':          fus.get('base_name') or key,
        'description':          fus.get('description'),
        'section':              m['section'],
        'financial_type':       FT_NORM.get(fus.get('financial_type', ''), fus.get('financial_type', '')) or '',
        'subledger_type':       fus.get('subledger_type'),
        'metric_type':          fus.get('metric_type'),
        'cost_type':            fus.get('cost_type'),
        'pm_type':              fus.get('pm_type'),
        'labor_revenue_type':   fus.get('labor_revenue_type'),
        'expense_revenue_type': fus.get('expense_revenue_type'),
        'is_active':            fus.get('is_active', True),
        'is_1099':              fus.get('is_1099', False),
        'is_subcontractor':     fus.get('is_subcontractor', False),
        'sort_order':           m['sort_order'],
        'notes':                m['notes'],
    })

# ── Build coa_crosswalk rows from COA MATRIX ──────────────────────────────────
crosswalk_rows = []
seen_crosswalk = set()

for r in src_rows:
    master_code = r[2]
    office_short = str(r[7] or '').strip()
    src_code = str(r[4] or '').strip()
    src_name = str(r[5] or '').strip()

    if not master_code or office_short not in OFFICES:
        continue

    # For QBD offices (DAL/ORL), source code is often N/A — fall back to name
    if src_code.upper() in ('', 'OPEN'):
        continue
    if src_code.upper() == 'N/A':
        if not src_name or src_name.upper() in ('', 'OPEN', 'N/A'):
            continue
        src_code = src_name  # use name as the identifier

    office_full = OFFICE_MAP[office_short]
    key = (str(master_code).strip(), office_full, src_code)
    if key in seen_crosswalk:
        continue
    seen_crosswalk.add(key)

    crosswalk_rows.append({
        'master_code':      str(master_code).strip(),
        'office':           office_full,
        'source_base_code': src_code,
        'source_base_name': blank(src_name),
        'mapped_by':        'etl',
        'notes':            None,
    })

# ── Add the 26 manual mappings ────────────────────────────────────────────────
MANUAL = [
    ('DAL', 'Design_Income-Accrued',                   'Design Income-Accrued',              42501),
    ('DAL', 'Fusion_AE_-_FVC_Bank',                    'Fusion AE - FVC Bank',               10302),
    ('DAL', 'Fusion_AE_-_Payroll',                     'Fusion AE - Payroll',                10201),
    ('DAL', 'ID_Studio4,_LLC_-_Sweep',                 'ID Studio4, LLC - Sweep',            10301),
    ('DAL', 'Investment_in_Reztark_ACi',               'Investment in Reztark ACi',          15001),
    ('DAL', 'NWC_Escrow_-_Reztark',                    'NWC Escrow - Reztark',               25001),
    ('DAL', 'Preferred_Equity_Payments',               'Preferred Equity Payments',          81701),
    ('MSP', '401',                                      'Billed fee revenue',                 40001),
    ('MSP', '767',                                      'Drafting Expenses',                  63101),
    ('MSP', '721',                                      'Equipment Rental',                   72601),
    ('MSP', '980',                                      'MN Min Fee Tax',                     81303),
    ('MSP', '186',                                      'Organization Expenses',              12401),
    ('MSP', '172',                                      'Preferred Equity Payments-MN',       81701),
    ('MSP', '550',                                      'Reimbursable Expenses',              61001),
    ('MSP', '981',                                      'Sales Tax',                          81301),
    ('MSP', '982',                                      'WI Fee Tax',                         81303),
    ('CIN', 'AJ-183',                                   'Federal Income Tax Withholding',     20303),
    ('CIN', 'AJ-205',                                   'Intercompany Due From',              12201),
    ('CIN', 'AJ-206',                                   'Intercompany Due To',                21203),
    ('CIN', 'AJ-207',                                   'Intercompany Other Income',          90101),
    ('CIN', 'AJ-90',                                    'Nonbillable Consultant Expenses',    60201),
    ('CIN', 'AJ-107',                                   'Professional Registration & Dues',   73101),
    ('ORL', 'Accrued_Revenue',                          'Accrued Revenue',                    20704),
    ('ORL', 'Insurance_Auto_Insurance',                 'Auto Insurance',                     74003),
    ('ORL', 'Travel_&_Entertainment_Entertainment',     'Entertainment',                      76001),
    ('ORL', 'Car/Truck_Expense_Registration_&_License', 'Registration & License',             73201),
]

for office_short, src_code, src_name, master_code in MANUAL:
    office_full = OFFICE_MAP[office_short]
    key = (str(master_code), office_full, src_code)
    if key not in seen_crosswalk:
        seen_crosswalk.add(key)
        crosswalk_rows.append({
            'master_code':      str(master_code),
            'office':           office_full,
            'source_base_code': src_code,
            'source_base_name': src_name,
            'mapped_by':        'manual',
            'notes':            'Manually mapped — no direct source account in original COA MATRIX',
        })

# ── Load into Supabase ────────────────────────────────────────────────────────
def chunk(lst, n=500):
    for i in range(0, len(lst), n):
        yield lst[i:i+n]

print('Truncating coa_crosswalk...')
sb.table('coa_crosswalk').delete().neq('id', 0).execute()

print('Truncating coa_master...')
sb.table('coa_master').delete().neq('id', 0).execute()

print(f'Loading {len(master_rows)} rows into coa_master...')
for batch in chunk(master_rows):
    sb.table('coa_master').insert(batch).execute()
print(f'  Done.')

# Deduplicate on (office, source_base_code) — keep first occurrence
deduped = {}
for row in crosswalk_rows:
    key = (row['office'], row['source_base_code'])
    if key not in deduped:
        deduped[key] = row
crosswalk_rows = list(deduped.values())

print(f'Loading {len(crosswalk_rows)} rows into coa_crosswalk (after dedup)...')
for batch in chunk(crosswalk_rows):
    sb.table('coa_crosswalk').upsert(
        batch, on_conflict='office,source_base_code'
    ).execute()
print(f'  Done.')

print(f'\nSummary:')
print(f'  coa_master:    {len(master_rows)} accounts loaded')
print(f'  coa_crosswalk: {len(crosswalk_rows)} source-to-master mappings loaded')
